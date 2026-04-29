import json
import re
import io
from datetime import date, datetime

import config
from config import (
    EXPENSE_CATEGORIES, EXPENSE_CARDS, TIMEZONE, overseas_state,
    YOUR_CHAT_ID,
)
from clients import client, spreadsheet, drive_service
from sheets import expenses_sheet, merchant_map_sheet, cards_sheet


# --- Emoji ---

EXPENSE_CATEGORY_EMOJI = {
    "FnB": "🍽️", "Transport": "🚗", "Entertainment": "🎬",
    "Personal": "🪞", "Family": "👨‍👩‍👧", "Work": "💼",
    "Shopping": "🛍️", "Household": "🏠", "Travel": "✈️",
}

EXPENSE_MERCHANT_OVERRIDES = {
    "FnB": [
        (["coffee", "starbucks", "kopitiam", "ya kun", "toast box", "kopi", "cafe", "espresso", "latte"], "☕"),
        (["ramen", "ichiran", "ippudo", "japanese", "sushi", "yakitori", "donburi", "izakaya"], "🍜"),
        (["mcdonald", "burger king", "wendy", "burger", "kfc", "popeyes", "fast food"], "🍔"),
        (["bar", "beer", "wine", "drinks", "cocktail", "pub", "taproom", "brewery"], "🍺"),
    ],
    "Transport": [
        (["flight", "airline", "air asia", "scoot", "singapore airlines", "cathay", "emirates", "jetstar"], "✈️"),
        (["grab", "gojek", "taxi", "uber", "ryde", "tada"], "🚕"),
        (["mrt", "bus", "transit", "ez-link", "train"], "🚌"),
    ],
    "Entertainment": [
        (["netflix", "disney", "hbo", "prime video", "apple tv", "streaming", "hulu", "mewatch"], "📺"),
        (["spotify", "apple music", "tidal", "deezer", "music"], "🎵"),
        (["cinema", "cathay", "gv", "shaw", "golden village", "movie"], "🎥"),
        (["steam", "playstation", "xbox", "nintendo", "game"], "🎮"),
    ],
    "Shopping": [
        (["guardian", "watsons", "unity", "pharmacy", "watson"], "💊"),
        (["ntuc", "giant", "cold storage", "fairprice", "supermarket", "grocery", "market"], "🛒"),
    ],
}

_category_emoji_cache = {}

def get_merchant_emoji(category, merchant):
    merchant_lower = merchant.lower() if merchant else ""
    overrides = EXPENSE_MERCHANT_OVERRIDES.get(category, [])
    for keywords, emoji in overrides:
        if any(kw in merchant_lower for kw in keywords):
            return emoji
    if category in EXPENSE_CATEGORY_EMOJI:
        return EXPENSE_CATEGORY_EMOJI[category]
    if category:
        if category in _category_emoji_cache:
            return _category_emoji_cache[category]
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6", max_tokens=5,
                messages=[{"role": "user", "content": f"Return a single emoji that best represents the expense category '{category}'. Return ONLY the emoji, nothing else."}]
            )
            inferred = resp.content[0].text.strip()
            _category_emoji_cache[category] = inferred
            return inferred
        except Exception:
            pass
    return "💳"


# --- Merchant Memory ---

def _get_merchant_records():
    if config._merchant_cache is None:
        try:
            config._merchant_cache = merchant_map_sheet().get_all_records()
        except Exception as e:
            print(f"_get_merchant_records error: {e}")
            return []
    return config._merchant_cache

def _invalidate_merchant_cache():
    config._merchant_cache = None

def get_merchant_memory(merchant):
    try:
        records = _get_merchant_records()
        merchant_lower = merchant.lower().strip()
        for r in records:
            if r.get("Merchant", "").lower() == merchant_lower:
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
        for r in records:
            known = r.get("Merchant", "").lower().strip()
            if known and (known in merchant_lower or merchant_lower in known):
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
        merchant_words = [w for w in merchant_lower.split() if len(w) > 2]
        for r in records:
            known = r.get("Merchant", "").lower().strip()
            known_words = [w for w in known.split() if len(w) > 2]
            if merchant_words and known_words and merchant_words[0] == known_words[0]:
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
    except Exception as e:
        print(f"get_merchant_memory error for '{merchant}': {e}")
    return None, None, None

def save_merchant_memory(merchant, category, card):
    try:
        sheet = merchant_map_sheet()
        sheet.append_row([merchant, category, card])
        _invalidate_merchant_cache()
        return True
    except Exception as e:
        print(f"save_merchant_memory failed for '{merchant}': {e}")
        return False

def delete_merchant(merchant_name):
    try:
        sheet = merchant_map_sheet()
        records = sheet.get_all_values()
        if len(records) <= 1:
            return "No merchants saved yet."
        matches = [(i, row[0]) for i, row in enumerate(records[1:], start=2)
                   if row and merchant_name.lower() in row[0].lower()]
        if not matches:
            return f"No merchant found matching '{merchant_name}'."
        if len(matches) == 1:
            row_idx, found_name = matches[0]
            sheet.delete_rows(row_idx)
            _invalidate_merchant_cache()
            return f"Deleted '{found_name}' from merchant memory ✅\nNext time you log there, I'll ask for category and card again."
        lines = [f"Found {len(matches)} merchants matching '{merchant_name}' — which one?"]
        for i, (_, name) in enumerate(matches, 1):
            lines.append(f"{i}. {name}")
        lines.append("\nReply 'delete merchant [exact name]' to remove a specific one.")
        return "\n".join(lines)
    except Exception as e:
        return f"Error deleting merchant: {e}"


# --- Cards ---

def get_cards_live():
    if config._card_names_cache is not None:
        return config._card_names_cache
    try:
        ws = cards_sheet()
        config._card_names_cache = ws.get_all_records()
        return config._card_names_cache
    except Exception as e:
        print(f"get_cards_live error: {e}")
        return []

def get_card_names_live():
    cards = get_cards_live()
    return [c.get("Card Name", "") for c in cards if c.get("Card Name")]

def get_card_default_for_category(category):
    try:
        cards = get_cards_live()
        for c in cards:
            default_cat = c.get("Default Category", "").strip().lower()
            if default_cat and default_cat == category.strip().lower():
                return c.get("Card Name", "")
        cat_lower = category.strip().lower()
        if cat_lower in ["fnb", "food", "dining", "f&b"]:
            return "Maybank"
        if cat_lower in ["grab", "transport"]:
            return "Amex"
        return "Citi"
    except Exception as e:
        print(f"get_card_default_for_category error: {e}")
        return "Citi"

def get_card_by_last4(last4):
    try:
        cards = get_cards_live()
        for c in cards:
            if str(c.get("Last 4", "")).strip() == str(last4).strip():
                return c.get("Card Name", "")
    except Exception as e:
        print(f"get_card_by_last4 error: {e}")
    return None

def set_card_default_category(card_name, category):
    try:
        ws = cards_sheet()
        records = ws.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("Card Name", "").lower() == card_name.lower():
                ws.update_cell(i, 3, category)
                return f"Updated — {card_name} will now default to {category} ✅"
        return f"Card '{card_name}' not found. Your cards: {', '.join(get_card_names_live())}"
    except Exception as e:
        return f"Couldn't update card default: {str(e)}"

def fuzzy_match_card(text):
    cards = get_card_names_live()
    text_lower = text.strip().lower()
    for c in cards:
        if c.lower() == text_lower:
            return c, True
    for c in cards:
        if text_lower in c.lower() or c.lower() in text_lower:
            return c, True
    for c in cards:
        if len(text_lower) >= 3 and c.lower().startswith(text_lower[:3]):
            return c, False
    return None, False

def fuzzy_match_category(text):
    text_lower = text.strip().lower()
    for c in EXPENSE_CATEGORIES:
        if c.lower() == text_lower:
            return c, True
    synonyms = {
        "food": "FnB", "dining": "FnB", "restaurant": "FnB", "eat": "FnB", "f&b": "FnB",
        "fun": "Entertainment", "movie": "Entertainment", "movies": "Entertainment",
        "gym": "Personal", "health": "Personal",
        "taxi": "Transport", "grab": "Transport", "uber": "Transport", "mrt": "Transport", "bus": "Transport",
        "clothes": "Shopping", "shop": "Shopping",
        "trip": "Travel", "holiday": "Travel", "flight": "Travel",
        "office": "Work", "business": "Work",
    }
    if text_lower in synonyms:
        return synonyms[text_lower], True
    for c in EXPENSE_CATEGORIES:
        if text_lower in c.lower() or c.lower() in text_lower:
            return c, False
    return None, False

def list_cards():
    try:
        ws = cards_sheet()
        records = ws.get_all_records()
        if not records:
            return "No cards set up yet."
        lines = ["Your cards:\n"]
        for r in records:
            name = r.get("Card Name", "")
            last4 = r.get("Last 4", "")
            default_cat = r.get("Default Category", "")
            line = f"• {name}"
            if last4:
                line += f" (*{last4})"
            if default_cat:
                line += f" — defaults to {default_cat}"
            lines.append(line)
        return "\n".join(lines)
    except Exception as e:
        return f"Error listing cards: {str(e)}"


# --- Expense Parsing ---

def parse_expense_text_v2(text):
    overseas_currency = overseas_state["currency"] if overseas_state["active"] else "SGD"
    live_cats = ", ".join(EXPENSE_CATEGORIES)
    live_cards = ", ".join(get_card_names_live())
    prompt = (
        f"Extract expense details from this message: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- merchant: string (brand name only, strip legal suffixes)\n"
        f"- amount: number (just the number, no currency symbol)\n"
        f"- currency: string (3-letter ISO code. If not mentioned, default to '{overseas_currency}')\n"
        f"- category: string (one of: {live_cats}) or empty if unclear\n"
        f"- card: string (one of: {live_cards}) or empty if not mentioned\n\n"
        f"Return ONLY the JSON."
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
        parsed = json.loads(raw)
        if "amount" in parsed:
            try:
                parsed["amount"] = float(parsed["amount"])
            except (ValueError, TypeError):
                parsed["amount"] = None
        return parsed
    except Exception as e:
        print(f"parse_expense_text Claude error: {e}")
        raise

def parse_expense_text(text):
    return parse_expense_text_v2(text)


# --- Expense Logging ---

def log_expense(merchant, amount, currency, sgd_amount, category, card, receipt_link="", reconciled="No", notes=""):
    try:
        today = date.today().strftime("%d/%m/%Y")
        sheet = expenses_sheet()
        sheet.append_row([today, merchant, amount, currency, sgd_amount, category, card, receipt_link, reconciled, notes])
        return True
    except Exception as e:
        print(f"log_expense failed: {e}")
        return False


# --- Format helpers ---

def format_expense_confirmation(merchant, amount, currency, category, card, sgd_amount=None,
                                receipt_saved=False, last4=None, high_amount=False,
                                use_manual_rate=False, missing_fields=None):
    emoji = get_merchant_emoji(category, merchant)
    lines = [f"{emoji} {merchant}"]
    if currency != "SGD" and sgd_amount is not None:
        amount_str = f"${sgd_amount:.2f} ({amount:,.0f} {currency})"
    else:
        amount_str = f"${amount:.2f}" if amount else "⚠️ Amount?"
    cat_str = f"⚠️ Category?" if not category else category
    if not card:
        card_str = "⚠️ Card?"
    elif last4:
        card_str = f"{card} (*{last4})"
    else:
        card_str = card
    lines.append(f"{amount_str} | {cat_str} | {card_str}")
    if high_amount:
        lines.append("⚠️ Amount looks high — is this correct?")
    if use_manual_rate:
        lines.append("⚠️ Using manual rate — verify if needed")
    if receipt_saved:
        lines.append("🧾 Receipt saved to Drive")
    if missing_fields:
        if len(missing_fields) >= 3:
            prompt = "yes / enter missing values / skip"
        elif len(missing_fields) == 2:
            prompt = f"yes / enter {' & '.join(f.title() for f in missing_fields)} / skip"
        else:
            prompt = f"yes / enter {missing_fields[0].title()} / skip"
        lines.append(f"\nLog this? ({prompt})")
    else:
        lines.append("\nLog this? (yes / edit name · amount · category · card / skip)")
    return "\n".join(lines)


def format_expense_logged(merchant, amount, currency, category, card, sgd_amount=None, last4=None):
    emoji = get_merchant_emoji(category, merchant)
    if currency != "SGD" and sgd_amount is not None:
        amount_str = f"${sgd_amount:.2f} ({amount:,.0f} {currency})"
    else:
        amount_str = f"${amount:.2f}"
    card_str = f"{card} (*{last4})" if last4 else card
    return f"{emoji} {merchant}\n{amount_str} | {category} | {card_str}\n\nLogged ✅"


# --- Summary / Reporting ---

def get_monthly_summary(month=None, year=None):
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No expenses logged yet."
        today = date.today()
        target_month = month or today.month
        target_year = year or today.year
        month_records = []
        for r in records:
            d = r.get("Date", "")
            if not d:
                continue
            try:
                dt = datetime.strptime(d, "%d/%m/%Y")
                if dt.month == target_month and dt.year == target_year:
                    month_records.append(r)
            except Exception as e:
                print(f"get_monthly_summary: bad date: {e}")
        if not month_records:
            return f"No expenses for {datetime(target_year, target_month, 1).strftime('%B %Y')}."
        total = sum(float(r.get("SGD Amount", 0) or r.get("Amount", 0)) for r in month_records)
        cat_totals = {c: 0 for c in EXPENSE_CATEGORIES}
        card_totals = {c: 0 for c in EXPENSE_CARDS}
        for r in month_records:
            cat = r.get("Category", "")
            if cat in cat_totals:
                amt = float(r.get("SGD Amount", 0) or r.get("Amount", 0))
                cat_totals[cat] += amt
            card = r.get("Card", "")
            if card in card_totals:
                amt = float(r.get("SGD Amount", 0) or r.get("Amount", 0))
                card_totals[card] += amt
        month_label = datetime(target_year, target_month, 1).strftime("%B %Y")
        lines = [f"Monthly Summary — {month_label}\n"]
        for cat, amt in cat_totals.items():
            if amt > 0:
                lines.append(f"{cat}: ${amt:.2f}")
        lines.append(f"\nTotal: ${total:.2f}")
        lines.append("\nBy card:")
        for card, amt in card_totals.items():
            if amt > 0:
                lines.append(f"{card}: ${amt:.2f}")
        return "\n".join(lines)
    except Exception as e:
        return f"Error generating summary: {str(e)}"

def get_expense_report(report_type="monthly"):
    return get_monthly_summary()

def get_trip_summary():
    try:
        trip_start = overseas_state.get("trip_start", "")
        destinations = overseas_state.get("trip_destinations", [])
        if not trip_start:
            return "No active trip — start overseas mode first."
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        start_dt = datetime.strptime(trip_start, "%d/%m/%Y").date()
        trip_records = []
        for r in records:
            d = r.get("Date", "")
            if not d:
                continue
            try:
                dt = datetime.strptime(d, "%d/%m/%Y").date()
                if dt >= start_dt:
                    trip_records.append(r)
            except Exception:
                continue
        if not trip_records:
            return "No expenses logged for this trip yet."
        total_sgd = sum(float(r.get("SGD Amount") or r.get("Amount") or 0) for r in trip_records)
        cat_totals = {}
        currency_totals = {}
        for r in trip_records:
            cat = r.get("Category", "Other")
            amt = float(r.get("SGD Amount") or r.get("Amount") or 0)
            cat_totals[cat] = cat_totals.get(cat, 0) + amt
            curr = r.get("Currency", "SGD")
            orig = float(r.get("Amount") or 0)
            if curr != "SGD":
                currency_totals[curr] = currency_totals.get(curr, 0) + orig
        dest_str = " → ".join(destinations) if destinations else "trip"
        lines = [f"Trip summary — {dest_str}", f"From {trip_start}\n"]
        for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
            lines.append(f"{cat}: SGD ${amt:.2f}")
        lines.append(f"\nTotal: SGD ${total_sgd:.2f}")
        if currency_totals:
            lines.append("\nSpend by currency:")
            for curr, amt in currency_totals.items():
                lines.append(f"{curr}: {amt:,.0f}")
        return "\n".join(lines)
    except Exception as e:
        return f"Error generating trip summary: {str(e)}"


def rename_category(old_name, new_name):
    matched = next((c for c in EXPENSE_CATEGORIES if c.lower() == old_name.lower()), None)
    if not matched:
        return f"Category '{old_name}' not found. Current categories: {', '.join(EXPENSE_CATEGORIES)}"
    new_clean = new_name.strip()
    config.EXPENSE_CATEGORIES[:] = [new_clean if c == matched else c for c in EXPENSE_CATEGORIES]
    updated = 0
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if all_values and "Category" in all_values[0]:
            col_idx = all_values[0].index("Category")
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) > col_idx and row[col_idx] == matched:
                    sheet.update_cell(i, col_idx + 1, new_clean)
                    updated += 1
    except Exception as e:
        print(f"rename_category expenses error: {e}")
    try:
        sheet = merchant_map_sheet()
        all_values = sheet.get_all_values()
        if all_values and "Category" in all_values[0]:
            col_idx = all_values[0].index("Category")
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) > col_idx and row[col_idx] == matched:
                    sheet.update_cell(i, col_idx + 1, new_clean)
    except Exception as e:
        print(f"rename_category merchant map error: {e}")
    try:
        ws = cards_sheet()
        records = ws.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("Default Category", "").strip().lower() == matched.lower():
                ws.update_cell(i, 3, new_clean)
    except Exception as e:
        print(f"rename_category cards error: {e}")
    return f"Renamed '{matched}' to '{new_clean}'. Updated {updated} expense(s)."


# --- Edit / Delete ---

EDIT_FIELD_SYNONYMS = {
    "merchant": "Merchant", "shop": "Merchant", "store": "Merchant", "place": "Merchant",
    "amount": "Amount", "price": "Amount", "total": "Amount", "cost": "Amount",
    "currency": "Currency",
    "category": "Category", "cat": "Category",
    "card": "Card", "payment": "Card",
    "notes": "Notes", "note": "Notes",
    "sgd": "SGD Amount",
}

def parse_multi_field_edit(text, field_keywords=None):
    if field_keywords is None:
        field_keywords = {
            "merchant": "merchant", "shop": "merchant", "store": "merchant", "place": "merchant",
            "amount": "amount", "price": "amount", "total": "amount", "cost": "amount",
            "currency": "currency",
            "category": "category", "cat": "category",
            "card": "card", "payment": "card", "pay": "card",
        }
    result = {}
    text = re.sub(r'"([^"]+)"', lambda m: m.group(0).replace(" ", "_SPACE_"), text)
    tokens = text.strip().split()
    current_field = None
    current_value_parts = []

    def flush():
        if current_field and current_value_parts:
            val = " ".join(current_value_parts).replace("_SPACE_", " ").strip('"')
            result[current_field] = val

    i = 0
    while i < len(tokens):
        token_lower = tokens[i].lower()
        if token_lower in field_keywords:
            flush()
            current_field = field_keywords[token_lower]
            current_value_parts = []
        elif current_field:
            current_value_parts.append(tokens[i])
        i += 1
    flush()
    return result

def get_last_expense():
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        if not records:
            return None
        return records[-1]
    except Exception as e:
        print(f"get_last_expense error: {e}")
        return None

def show_last_expense():
    r = get_last_expense()
    if not r:
        return "No expenses logged yet."
    currency = r.get("Currency", "SGD")
    amount = r.get("Amount", "")
    sgd = r.get("SGD Amount", "")
    _merchant = r.get('Merchant', '')
    _cat = r.get('Category', '')
    _emoji = get_merchant_emoji(_cat, _merchant)
    lines = [f"Last expense: {_emoji} *{_merchant}*"]
    if currency != "SGD" and sgd:
        lines.append(f"${float(sgd):.2f} SGD ({float(amount):,.0f} {currency})")
    elif amount:
        lines.append(f"${float(amount):.2f}")
    if _cat:
        lines.append(f"Category: {_cat}")
    card = r.get("Card", "")
    if card:
        lines.append(f"Card: {card}")
    date_str = r.get("Date", "")
    if date_str:
        lines.append(f"Date: {date_str}")
    return "\n".join(lines)

def edit_last_expense(edit_text):
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return "No expenses to edit."
        headers = all_values[0]
        last_row_idx = len(all_values)
        last_row = all_values[last_row_idx - 1]
        edits = parse_multi_field_edit(edit_text, field_keywords=EDIT_FIELD_SYNONYMS)
        if not edits:
            return "Couldn't parse that edit. Try: 'edit merchant Starbucks category FnB'"
        applied = []
        merchant_changed = False
        for field_key, new_value in edits.items():
            col_name = EDIT_FIELD_SYNONYMS.get(field_key.lower())
            if not col_name or col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            if col_name == "Amount":
                try:
                    new_value = str(float(re.sub(r"[^\d.]", "", new_value)))
                except ValueError:
                    continue
            elif col_name == "Category":
                matched_cat, _ = fuzzy_match_category(new_value)
                if not matched_cat:
                    continue
                new_value = matched_cat
            elif col_name == "Card":
                matched_card, _ = fuzzy_match_card(new_value)
                if not matched_card:
                    continue
                new_value = matched_card
            elif col_name == "Merchant":
                merchant_changed = True
            sheet.update_cell(last_row_idx, col_idx, new_value)
            applied.append((col_name, new_value))
        if not applied:
            return "Couldn't match any valid fields to edit."
        if merchant_changed:
            new_merchant = next((v for k, v in edits.items() if EDIT_FIELD_SYNONYMS.get(k) == "Merchant"), "")
            if new_merchant:
                known_cat, known_card, canonical = get_merchant_memory(new_merchant)
                if canonical:
                    m_col = headers.index("Merchant") + 1
                    sheet.update_cell(last_row_idx, m_col, canonical)
                if known_cat and "Category" not in [c for c, _ in applied]:
                    c_col = headers.index("Category") + 1
                    sheet.update_cell(last_row_idx, c_col, known_cat)
                    applied.append(("Category", known_cat))
                if known_card and "Card" not in [c for c, _ in applied]:
                    cd_col = headers.index("Card") + 1
                    sheet.update_cell(last_row_idx, cd_col, known_card)
                    applied.append(("Card", known_card))
        updated_values = sheet.get_all_values()
        updated_row_data = updated_values[last_row_idx - 1] if len(updated_values) >= last_row_idx else last_row
        updated = {headers[i]: updated_row_data[i] if i < len(updated_row_data) else "" for i in range(len(headers))}
        merchant = updated.get("Merchant", "")
        amount = updated.get("Amount", "")
        currency = updated.get("Currency", "SGD")
        sgd_amount = updated.get("SGD Amount", amount)
        category = updated.get("Category", "")
        card = updated.get("Card", "")
        return format_expense_logged(merchant, float(amount) if amount else 0,
                                     currency, category, card,
                                     float(sgd_amount) if sgd_amount else None)
    except Exception as e:
        return f"Error editing expense: {str(e)}"

def delete_last_expense():
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return "No expenses to delete."
        last_row_idx = len(all_values)
        row = all_values[last_row_idx - 1]
        merchant = row[1] if len(row) > 1 else "?"
        amount = row[2] if len(row) > 2 else "?"
        sheet.delete_rows(last_row_idx)
        return f"Deleted last expense: {merchant} ${amount} ✅"
    except Exception as e:
        return f"Error deleting expense: {str(e)}"

def get_recent_expenses(n=5):
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return []
        headers = all_values[0]
        rows = all_values[1:]
        recent = []
        for i, row in enumerate(reversed(rows)):
            if i >= n:
                break
            d = {headers[j]: row[j] if j < len(row) else "" for j in range(len(headers))}
            sheet_row = len(rows) - i + 1
            recent.append((sheet_row, d))
        return recent
    except Exception as e:
        print(f"get_recent_expenses error: {e}")
        return []

def format_delete_list(expenses):
    lines = ["Which expense to delete?\n"]
    for i, (_, r) in enumerate(expenses, 1):
        merchant = r.get("Merchant", "?")
        amount = r.get("Amount", "")
        currency = r.get("Currency", "SGD")
        sgd = r.get("SGD Amount", "")
        date_str = r.get("Date", "")
        category = r.get("Category", "")
        if currency != "SGD" and sgd:
            amt_str = f"${float(sgd):.2f} ({currency} {float(amount):,.0f})"
        else:
            amt_str = f"${float(amount):.2f}" if amount else "$?"
        lines.append(f"{i}. {merchant} — {amt_str} | {category} | {date_str}")
    lines.append("\nReply with a number, or 'search [merchant]' to find another.")
    return "\n".join(lines)

def search_expenses_by_merchant(query):
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return []
        headers = all_values[0]
        rows = all_values[1:]
        q = query.lower().strip()
        matches = []
        for i, row in enumerate(rows):
            d = {headers[j]: row[j] if j < len(row) else "" for j in range(len(headers))}
            if q in d.get("Merchant", "").lower():
                matches.append((i + 2, d))
        return list(reversed(matches))[:10]
    except Exception as e:
        print(f"search_expenses_by_merchant error: {e}")
        return []

def delete_expense_by_row(sheet_row):
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if sheet_row < 2 or sheet_row > len(all_values):
            return "Couldn't find that expense."
        row = all_values[sheet_row - 1]
        merchant = row[1] if len(row) > 1 else "?"
        amount = row[2] if len(row) > 2 else "?"
        sheet.delete_rows(sheet_row)
        return f"Deleted: {merchant} ${amount}"
    except Exception as e:
        return f"Error deleting expense: {str(e)}"


# --- Detection helpers ---

def is_expense_input(text):
    from trips import extract_flight_number
    if extract_flight_number(text):
        return False
    lower = text.lower()
    if re.search(r"\d+ shares|shares of |shares @|shares at \$", lower):
        return False
    exclusions = [
        "delete expense", "remove expense", "undo expense",
        "edit expense", "edit last expense",
        "rename category", "expense report", "monthly report",
        "trip summary", "last expense", "show last expense",
        "what expense categories", "list my categories", "what categories",
        "show categories", "what are my expense", "list categories",
        "expense categories", "my categories",
        "what merchants", "my merchants", "merchant map", "list merchants",
        "show merchants", "known merchants"
    ]
    if any(lower.startswith(e) or lower == e for e in exclusions):
        return False
    triggers = ["spent", "paid", "$", "sgd", "charged", "bought", "grabbed",
                "receipt", "bill was", "cost me", "picked up",
                "recorded in", "will it be", "logged in", "track", "how much have i"]
    return any(t in lower for t in triggers)

def is_log_prefix_input(text):
    return text.lower().startswith("log ") and len(text) > 4

def is_bare_merchant_input(text):
    lower = text.lower().strip()
    if not re.search(r"\$[\d,.]+", text):
        return False
    first_word = text.strip().split()[0]
    known_cat, _, _ = get_merchant_memory(first_word)
    return bool(known_cat)

def is_bill_request(text):
    lower = text.lower()
    triggers = [
        "add bill", "new bill", "set up bill", "set bill", "bill reminder",
        "track bill", "add a bill", "credit card bill", "utility bill",
        "monthly bill", "subscription", "due on the"
    ]
    return any(t in lower for t in triggers)

def get_expense_categories():
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        cats = sorted(set(r.get("Category", "").strip() for r in records if r.get("Category", "").strip()))
        if not cats:
            cats = EXPENSE_CATEGORIES
    except Exception:
        cats = EXPENSE_CATEGORIES
    numbered = "\n".join(f"{i+1}. {c}" for i, c in enumerate(cats))
    return f"Your expense categories:\n\n{numbered}"

def get_merchant_list():
    try:
        sheet = merchant_map_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No merchants saved yet. They are added automatically when you log a new expense."
        by_cat = {}
        for r in records:
            merchant = r.get("Merchant", "").strip()
            cat = r.get("Category", "").strip() or "Uncategorised"
            card = r.get("Card", "").strip()
            if not merchant:
                continue
            by_cat.setdefault(cat, []).append(merchant + (f" ({card})" if card else ""))
        count = sum(len(v) for v in by_cat.values())
        lines = [f"Merchant Map ({count} merchants):\n"]
        for cat in sorted(by_cat):
            lines.append(f"*{cat}*")
            for m in sorted(by_cat[cat]):
                lines.append(f"  {m}")
            lines.append("")
        return "\n".join(lines).strip()
    except Exception as e:
        return f"Error fetching merchants: {e}"


# --- Bills ---

BILL_REMINDER_GREETINGS = ["Heads up", "Just a nudge", "Quick reminder", "Hey", "FYI"]

from sheets import bills_sheet

def parse_bill_request(text):
    prompt = (
        f"Extract bill details from: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- name: string\n- bank: string\n- due_day: number\n"
        f"- estimated_amount: number\n- notes: string\n\nReturn ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
    return json.loads(raw)

def add_bill(name, bank, due_day, estimated_amount, notes=""):
    sheet = bills_sheet()
    sheet.append_row([name, bank, str(due_day), str(estimated_amount), notes])

def list_bills():
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No bills set up yet."
        lines = ["Your bills:\n"]
        for r in records:
            amt = r.get("Estimated Amount", "")
            amt_str = f" — ~${amt}" if amt and str(amt) != "0" else ""
            lines.append(f"• {r.get('Name', '')} (due day {r.get('Due Date', '')}){amt_str}")
        return "\n".join(lines)
    except Exception as e:
        return f"Error listing bills: {str(e)}"

def delete_bill(name):
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if name.lower() in r.get("Name", "").lower():
                sheet.delete_rows(i + 2)
                return f"Deleted bill: {r.get('Name')}"
        return f"No bill found matching '{name}'."
    except Exception as e:
        return f"Error deleting bill: {str(e)}"

def get_cycle_expenses(card_name, due_day):
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        today = date.today()
        if today.day >= due_day:
            cycle_start = today.replace(day=due_day)
        else:
            if today.month == 1:
                cycle_start = today.replace(year=today.year - 1, month=12, day=due_day)
            else:
                cycle_start = today.replace(month=today.month - 1, day=due_day)
        total = 0
        for r in records:
            try:
                dt = datetime.strptime(r.get("Date", ""), "%d/%m/%Y").date()
                card = r.get("Card", "")
                if dt >= cycle_start and card_name.lower() in card.lower():
                    total += float(r.get("SGD Amount", 0) or r.get("Amount", 0))
            except Exception as e:
                print(f"get_cycle_expenses: bad row: {e}")
        return total
    except Exception as e:
        print(f"get_cycle_expenses error: {e}")
        return 0

async def send_bill_reminders(app):
    import random
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        today = date.today()
        for r in records:
            try:
                due_day = int(r.get("Due Date", 0))
                if not due_day:
                    continue
                if today.day <= due_day:
                    next_due = today.replace(day=due_day)
                else:
                    if today.month == 12:
                        next_due = today.replace(year=today.year + 1, month=1, day=due_day)
                    else:
                        next_due = today.replace(month=today.month + 1, day=due_day)
                days_away = (next_due - today).days
                if days_away == 7:
                    name = r.get("Name", "")
                    estimated = r.get("Estimated Amount", "")
                    cycle_total = get_cycle_expenses(name, due_day)
                    amount_str = f"${cycle_total:.2f} logged this cycle" if cycle_total > 0 else (f"~${estimated}" if estimated and str(estimated) != "0" else "amount unknown")
                    greeting = random.choice(BILL_REMINDER_GREETINGS)
                    msg = (
                        f"{greeting} — your {name} bill is due in 7 days ({next_due.strftime('%d %b')}).\n"
                        f"{amount_str}."
                    )
                    await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
            except Exception as e:
                print(f"Bill reminder error for {r.get('Name', '?')}: {e}")
    except Exception as e:
        print(f"send_bill_reminders error: {e}")

def handle_new_bill(text):
    try:
        parsed = parse_bill_request(text)
        name = parsed.get("name", "")
        bank = parsed.get("bank", "")
        due_day = parsed.get("due_day", 0)
        estimated_amount = parsed.get("estimated_amount", 0)
        notes = parsed.get("notes", "")
        if not name or not due_day:
            return "Need at least a bill name and due day — try: 'add Citi bill due on the 15th, ~$500'"
        add_bill(name, bank, due_day, estimated_amount, notes)
        due_str = f"the {due_day}{'st' if due_day == 1 else 'nd' if due_day == 2 else 'rd' if due_day == 3 else 'th'}"
        return f"Bill added — {name} is due on {due_str} each month. I'll remind you 7 days before ✅"
    except Exception as e:
        return f"Couldn't parse that — try: 'add Citi bill due on the 15th, ~$500'"


# --- Expense session handler ---

async def handle_expense_session(user_id, text, update):
    """Handle expense session: category → card → log."""
    session = config.expense_sessions.get(user_id)
    if not session:
        return False
    from state import touch_session
    touch_session(user_id)
    lower = text.strip().lower()
    if lower in ["skip", "cancel", "no"]:
        del config.expense_sessions[user_id]
        config.session_timestamps.pop(user_id, None)
        await update.message.reply_text("Skipped.")
        return True

    step = session.get("step")
    merchant = session.get("merchant", "")
    amount = session.get("amount", 0)
    currency = session.get("currency", "SGD")

    if step == "category":
        matched, _ = fuzzy_match_category(text)
        if not matched:
            await update.message.reply_text(f"Didn't recognise that — pick one: {', '.join(EXPENSE_CATEGORIES)}")
            return True
        session["category"] = matched
        session["step"] = "card"
        config.expense_sessions[user_id] = session
        cards = get_card_names_live()
        await update.message.reply_text(f"Got it — {matched}. Which card? ({' / '.join(cards)})")
        return True

    if step == "card":
        matched_card, _ = fuzzy_match_card(text)
        if not matched_card:
            cards = get_card_names_live()
            await update.message.reply_text(f"Didn't recognise that card — pick one: {', '.join(cards)}")
            return True
        category = session.get("category", "")
        from state import get_fx_rate
        sgd_amount = amount if currency == "SGD" else (amount * (get_fx_rate(currency) or 1))
        success = log_expense(merchant, amount, currency, round(sgd_amount, 2), category, matched_card)
        if success:
            save_merchant_memory(merchant, category, matched_card)
        del config.expense_sessions[user_id]
        config.session_timestamps.pop(user_id, None)
        if success:
            await update.message.reply_text(format_expense_logged(merchant, amount, currency, category, matched_card, sgd_amount if currency != "SGD" else None))
        else:
            await update.message.reply_text("⚠️ Expense session complete but couldn't write to sheet — check Em Log.")
        return True

    return False


def handle_expense_text(text, user_id, receipt_link=""):
    """Parse expense text and return (reply, needs_session, session_data)."""
    from state import get_fx_rate, save_manual_fx_rate, parse_manual_fx_input
    try:
        parsed = parse_expense_text_v2(text)
        merchant = parsed.get("merchant", "")
        amount = parsed.get("amount")
        currency = parsed.get("currency", "SGD")
        category = parsed.get("category", "")
        card = parsed.get("card", "")

        if not merchant or amount is None:
            return "Couldn't parse that — try: '45 Ichiran' or 'Grab $12.50'", False, None

        # Merchant memory
        known_cat, known_card, canonical = get_merchant_memory(merchant)
        if canonical:
            merchant = canonical
        if known_cat and not category:
            category = known_cat
        if known_card and not card:
            card = known_card

        # Card last4 lookup
        last4 = None
        if not card:
            card = get_card_default_for_category(category)
        if card:
            cards = get_cards_live()
            for c in cards:
                if c.get("Card Name", "") == card:
                    last4 = str(c.get("Last 4", "")) or None
                    break

        # FX
        sgd_amount = None
        use_manual_rate = False
        fx_rate = get_fx_rate(currency) if currency != "SGD" else 1.0
        if fx_rate is None:
            session_data = {
                "step": "fx_rate", "merchant": merchant, "amount": amount,
                "currency": currency, "category": category, "card": card,
                "receipt_link": receipt_link, "last4": last4,
                "missing_fields": ([f for f in ["category", "card"] if not locals().get(f)])
            }
            return (
                f"What's today's exchange rate? Try:\n"
                f"• `{(1/3.11):.4f}` (1 {currency} = that many SGD)\n"
                f"• `1 SGD to 3.11 {currency}`\n"
                f"• `1 {currency} = 0.32 SGD`",
                True, session_data
            )
        if currency != "SGD":
            sgd_amount = round(amount * fx_rate, 2)

        high_amount = (currency == "SGD" and amount > 500) or (currency != "SGD" and sgd_amount and sgd_amount > 500)
        missing_fields = [f for f in ["category", "card"] if not (category if f == "category" else card)]

        session_data = {
            "step": "receipt_confirm",
            "merchant": merchant, "amount": amount, "currency": currency,
            "sgd_amount": sgd_amount, "category": category, "card": card,
            "receipt_link": receipt_link, "last4": last4, "high_amount": high_amount,
            "missing_fields": missing_fields, "use_manual_rate": use_manual_rate,
        }
        reply = format_expense_confirmation(
            merchant, amount, currency, category, card, sgd_amount,
            receipt_saved=bool(receipt_link), last4=last4, high_amount=high_amount,
            use_manual_rate=use_manual_rate, missing_fields=missing_fields
        )
        return reply, True, session_data

    except Exception as e:
        return f"Couldn't parse that expense — try: '45 Ichiran' or 'Grab $12.50' ({type(e).__name__})", False, None


async def _finalise_receipt_confirm(user_id, session, update):
    """Log a confirmed expense from receipt_confirm_sessions."""
    from state import get_fx_rate
    merchant = session.get("merchant", "")
    amount = session.get("amount", 0)
    currency = session.get("currency", "SGD")
    sgd_amount = session.get("sgd_amount")
    category = session.get("category", "")
    card = session.get("card", "")
    receipt_link = session.get("receipt_link", "")
    last4 = session.get("last4")

    if currency != "SGD" and sgd_amount is None:
        rate = get_fx_rate(currency)
        if rate:
            sgd_amount = round(amount * rate, 2)

    success = log_expense(
        merchant, amount, currency,
        sgd_amount if sgd_amount is not None else amount,
        category, card, receipt_link
    )
    if success:
        save_merchant_memory(merchant, category, card)
        del config.receipt_confirm_sessions[user_id]
        config.session_timestamps.pop(user_id, None)
        reply = format_expense_logged(merchant, amount, currency, category, card,
                                      sgd_amount if currency != "SGD" else None, last4)
    else:
        reply = "⚠️ Confirmed but couldn't write to sheet — try again or check the connection."
    await update.message.reply_text(reply, parse_mode="Markdown")


async def handle_receipt_confirm_session(user_id, text, update):
    """Handle receipt/expense confirm session."""
    from state import touch_session, get_fx_rate, save_manual_fx_rate, parse_manual_fx_input
    session = config.receipt_confirm_sessions.get(user_id)
    if not session:
        return False

    touch_session(user_id)
    lower = text.strip().lower()

    if lower in ["skip", "cancel", "no", "n", "nope", "nah"]:
        del config.receipt_confirm_sessions[user_id]
        config.session_timestamps.pop(user_id, None)
        receipt_link = session.get("receipt_link", "")
        if receipt_link:
            await update.message.reply_text("Skipped — receipt saved to Drive if you need it later.")
        else:
            await update.message.reply_text("Skipped.")
        return True

    if session.get("step") == "fx_rate":
        currency = session.get("currency", "")
        parsed_rate, display = parse_manual_fx_input(text, currency)
        if parsed_rate is None:
            await update.message.reply_text(
                f"Couldn't read that rate — try one of these formats:\n"
                f"• `3.11` (1 {currency} = 3.11 SGD)\n"
                f"• `1 SGD to 3.11 {currency}`\n"
                f"• `1 {currency} = 3.11 SGD`"
            )
            return True
        num, sgd_per_unit = parsed_rate
        save_manual_fx_rate(currency, num, sgd_per_unit)
        rate = get_fx_rate(currency)
        session["sgd_amount"] = round(session["amount"] * rate, 2)
        session["step"] = "receipt_confirm"
        session["use_manual_rate"] = True
        config.receipt_confirm_sessions[user_id] = session
        await update.message.reply_text(
            f"Got it — using {display} for today.\n\n" +
            format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                last4=session.get("last4"), high_amount=session.get("high_amount", False),
                use_manual_rate=True, missing_fields=session.get("missing_fields", [])
            )
        )
        return True

    if session.get("step") == "duplicate_confirm":
        if lower in ["yes", "y", "yep", "yeah", "yup"]:
            config.receipt_confirm_sessions[user_id] = session
            await update.message.reply_text(
                format_expense_confirmation(
                    session["merchant"], session["amount"], session["currency"],
                    session.get("category", ""), session.get("card", ""),
                    session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                    last4=session.get("last4"), high_amount=session.get("high_amount", False),
                    use_manual_rate=session.get("use_manual_rate", False),
                    missing_fields=session.get("missing_fields", [])
                )
            )
        else:
            del config.receipt_confirm_sessions[user_id]
            await update.message.reply_text("Skipped duplicate.")
        return True

    if lower in ["yes", "y", "yep", "yeah", "yup"]:
        await _finalise_receipt_confirm(user_id, session, update)
        return True

    card_match, _ = fuzzy_match_card(text.strip())
    if card_match and len(text.strip().split()) == 1:
        session["card"] = card_match
        await _finalise_receipt_confirm(user_id, session, update)
        return True

    edit_field_map = {
        "edit name": "merchant", "edit merchant": "merchant",
        "edit amount": "amount", "edit price": "amount",
        "edit category": "category", "edit cat": "category",
        "edit card": "card", "edit payment": "card",
        "edit currency": "currency",
    }
    for prompt, field in edit_field_map.items():
        if lower == prompt:
            session["_editing"] = field
            config.receipt_confirm_sessions[user_id] = session
            await update.message.reply_text(f"What should the {field} be?")
            return True

    if session.get("_editing"):
        field = session.pop("_editing")
        if field == "merchant":
            session["merchant"] = text.strip()
        elif field == "amount":
            try:
                session["amount"] = float(re.sub(r"[^\d.]", "", text))
            except ValueError:
                await update.message.reply_text("Couldn't parse that amount.")
                return True
        elif field == "category":
            matched, _ = fuzzy_match_category(text)
            if matched:
                session["category"] = matched
        elif field == "card":
            matched, _ = fuzzy_match_card(text)
            if matched:
                session["card"] = matched
        elif field == "currency":
            session["currency"] = text.strip().upper()[:3]
        config.receipt_confirm_sessions[user_id] = session
        await update.message.reply_text(
            format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                last4=session.get("last4"), high_amount=session.get("high_amount", False),
                use_manual_rate=session.get("use_manual_rate", False),
                missing_fields=session.get("missing_fields", [])
            )
        )
        return True

    edits = parse_multi_field_edit(lower)
    if edits:
        for field_key, new_value in edits.items():
            if field_key == "merchant":
                session["merchant"] = new_value
            elif field_key == "amount":
                try:
                    session["amount"] = float(re.sub(r"[^\d.]", "", new_value))
                except ValueError:
                    pass
            elif field_key == "category":
                matched, _ = fuzzy_match_category(new_value)
                if matched:
                    session["category"] = matched
            elif field_key == "card":
                matched, _ = fuzzy_match_card(new_value)
                if matched:
                    session["card"] = matched
            elif field_key == "currency":
                session["currency"] = new_value.upper()[:3]
        config.receipt_confirm_sessions[user_id] = session
        await update.message.reply_text(
            format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                last4=session.get("last4"), high_amount=session.get("high_amount", False),
                use_manual_rate=session.get("use_manual_rate", False),
                missing_fields=session.get("missing_fields", [])
            )
        )
        return True

    return False


async def handle_statement_upload(file_bytes, fname, user_id, update):
    """Parse a bank statement CSV/XLSX and reconcile against logged expenses."""
    try:
        await update.message.reply_text("Got the statement, give me a sec to go through it...")
        statement_rows = []
        if fname.lower().endswith(".csv"):
            import csv
            text_data = file_bytes.decode("utf-8", errors="ignore")
            import io as _io
            reader = csv.DictReader(_io.StringIO(text_data))
            for row in reader:
                statement_rows.append(dict(row))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                statement_rows.append(dict(zip(headers, row)))

        if not statement_rows:
            await update.message.reply_text("Couldn't read any rows from that file.")
            return

        sample = json.dumps(statement_rows[:3], default=str)
        norm_prompt = (
            f"Given these bank statement rows: {sample}\n\n"
            f"Return ONLY a JSON object mapping these keys to the actual column names in the data:\n"
            f'{{\"date\": \"col\", \"description\": \"col\", \"amount\": \"col\"}}\n'
            f"If a column doesn't exist, use null."
        )
        norm_resp = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=100,
            messages=[{"role": "user", "content": norm_prompt}]
        )
        col_map = json.loads(norm_resp.content[0].text.strip().replace("```json","").replace("```","").strip())
        date_col = col_map.get("date")
        desc_col = col_map.get("description")
        amt_col = col_map.get("amount")

        if not all([date_col, desc_col, amt_col]):
            await update.message.reply_text("Couldn't identify date/description/amount columns. Try a CSV with clear headers.")
            return

        sheet = expenses_sheet()
        logged = sheet.get_all_records()
        missing = []
        corrections = []

        for srow in statement_rows:
            raw_date = str(srow.get(date_col, "")).strip()
            raw_desc = str(srow.get(desc_col, "")).strip()
            raw_amt = str(srow.get(amt_col, "")).strip().replace(",", "")
            if not raw_date or not raw_amt:
                continue
            try:
                stmt_amount = abs(float(raw_amt))
            except ValueError:
                continue

            stmt_date = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %b %y"):
                try:
                    stmt_date = datetime.strptime(raw_date, fmt).strftime("%d/%m/%Y")
                    break
                except Exception:
                    continue
            if not stmt_date:
                continue

            matched_row = None
            matched_idx = None
            for i, r in enumerate(logged):
                logged_date = r.get("Date", "")
                logged_sgd = float(r.get("SGD Amount") or r.get("Amount") or 0)
                if logged_date == stmt_date and abs(logged_sgd - stmt_amount) < 0.02:
                    matched_row = r
                    matched_idx = i
                    break

            if matched_row is None:
                missing.append(f"{stmt_date} | {raw_desc[:40]} | SGD ${stmt_amount:.2f}")
            else:
                logged_sgd = float(matched_row.get("SGD Amount") or matched_row.get("Amount") or 0)
                if abs(logged_sgd - stmt_amount) > 0.01:
                    all_values = sheet.get_all_values()
                    headers_row = all_values[0]
                    sgd_col_idx = headers_row.index("SGD Amount") + 1 if "SGD Amount" in headers_row else None
                    if sgd_col_idx:
                        sheet.update_cell(matched_idx + 2, sgd_col_idx, stmt_amount)
                    corrections.append(
                        f"{matched_row.get('Merchant', raw_desc[:20])} {stmt_date}: "
                        f"${logged_sgd:.2f} → ${stmt_amount:.2f}"
                    )

        lines = ["Reconciliation done ✅"]
        if corrections:
            lines.append(f"\nCorrected {len(corrections)} amount(s):")
            lines.extend(f"  {c}" for c in corrections)
        if missing:
            lines.append(f"\n{len(missing)} unmatched statement item(s) — couldn't find in your logs:")
            lines.extend(f"  {m}" for m in missing[:10])
            if len(missing) > 10:
                lines.append(f"  ...and {len(missing) - 10} more")
        if not corrections and not missing:
            lines.append("Everything matches up.")

        await update.message.reply_text("\n".join(lines))

    except Exception as e:
        print(f"handle_statement_upload error: {e}")
        await update.message.reply_text(f"Something went wrong parsing the statement: {str(e)}")
