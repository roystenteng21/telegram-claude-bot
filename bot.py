import os
import asyncio
import json
import re
import io
import httpx
from dotenv import load_dotenv
from anthropic import Anthropic
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes
from datetime import date, datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import pytz
import caldav

load_dotenv()

# --- Credentials ---
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
SHEET_ID = os.getenv("SHEET_ID")
ICLOUD_USERNAME = os.getenv("ICLOUD_USERNAME")
ICLOUD_PASSWORD = os.getenv("ICLOUD_PASSWORD")
AVIATIONSTACK_API_KEY = os.getenv("AVIATIONSTACK_API_KEY", "")
EXCHANGE_RATE_API_KEY = os.getenv("EXCHANGE_RATE_API_KEY", "")
ALPHA_VANTAGE_API_KEY = os.getenv("ALPHA_VANTAGE_API_KEY", "")
YOUR_CHAT_ID = int(os.getenv("YOUR_CHAT_ID", "281095850"))
RAILWAY_DEPLOYMENT_ID = os.getenv("RAILWAY_DEPLOYMENT_ID", "")

# Anthropic API health tracking
_anthropic_failure_count = 0
_anthropic_down_notified = False
ANTHROPIC_FAILURE_THRESHOLD = 3

# iCloud health tracking
_icloud_down = False
_icloud_last_notified = None

# Manual FX rate cache: { currency: { "rate": float, "date": str, "sgd_per_unit": bool } }
manual_fx_rates = {}

# Session timestamps for timeout: { user_id: datetime }
session_timestamps = {}

# --- Google Sheets + Drive Setup ---
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]
google_creds_env = os.getenv("GOOGLE_CREDENTIALS")
if google_creds_env:
    google_creds = json.loads(google_creds_env)
    creds = Credentials.from_service_account_info(google_creds, scopes=SCOPES)
else:
    creds = Credentials.from_service_account_file("credentials.json", scopes=SCOPES)

gc = gspread.authorize(creds)
spreadsheet = gc.open_by_key(SHEET_ID)
drive_service = build("drive", "v3", credentials=creds)

# --- em_profile (loaded at startup, updated by setup) ---
em_profile = {}

# --- Infrastructure Setup ---
def setup_sheets():
    """Rename Sheet1 to CRM if needed, create all required tabs.
    Single worksheets() call at start — no repeated API reads."""
    # One API call — reused throughout
    existing = [ws.title for ws in spreadsheet.worksheets()]

    CRM_HEADERS = [
        "Name", "Alias", "Birthday", "Relationship", "Context", "Notes",
        "Follow Up Date", "Follow Up Notes", "Last Updated", "Birthday Greeted",
        "Referred By", "Referral Date", "Email", "Address"
    ]

    if "CRM" not in existing:
        try:
            sheet1 = spreadsheet.worksheet("Sheet1")
            sheet1.update_title("CRM")
            existing.append("CRM")
            print("Renamed Sheet1 -> CRM")
        except Exception:
            if "CRM" not in existing:
                ws = spreadsheet.add_worksheet(title="CRM", rows=1000, cols=20)
                ws.append_row(CRM_HEADERS)
                existing.append("CRM")
                print("Created CRM tab")

    try:
        crm_ws = spreadsheet.worksheet("CRM")
        current_headers = crm_ws.row_values(1)
        if current_headers != CRM_HEADERS:
            _migrate_crm_headers(crm_ws, current_headers, CRM_HEADERS)
    except Exception as e:
        print(f"CRM header check error: {e}")

    if "Todos" not in existing:
        ws = spreadsheet.add_worksheet(title="Todos", rows=500, cols=3)
        ws.append_row(["Task", "Status", "Added"])
        existing.append("Todos")
        print("Created Todos tab")

    required_tabs = {
        "Meeting Notes": ["Event Name", "Topic", "Summary", "Action Items", "Date"],
        "Expenses": ["Date", "Merchant", "Amount", "Currency", "SGD Amount", "Category",
                     "Card", "Receipt Link", "Reconciled", "Notes"],
        "Bills": ["Name", "Bank", "Due Date", "Estimated Amount", "Notes"],
        "Merchant Map": ["Merchant", "Category", "Card"],
        "Restaurants": ["Name", "Location", "Country", "Tags", "Notes"],
        "Portfolio": ["Stock", "Quantity", "Buy Price", "Buy Date", "Notes"],
        "Trips": ["Trip ID", "Destination", "Currency", "Check In", "Check Out",
                  "Hotel Name", "Hotel Local Name", "Hotel Address", "Notes", "Status"],
        "Reminders": ["ID", "Message", "Scheduled Time", "Recurrence", "Status", "Attempts", "Contact"],
        "Settings": ["Key", "Value"],
    }

    for tab_name, headers in required_tabs.items():
        if tab_name not in existing:
            ws = spreadsheet.add_worksheet(title=tab_name, rows=500, cols=len(headers))
            ws.append_row(headers)
            existing.append(tab_name)
            print(f"Created tab: {tab_name}")

    try:
        if "Cards" not in existing:
            cards_ws = spreadsheet.add_worksheet(title="Cards", rows=100, cols=4)
            existing.append("Cards")
        else:
            cards_ws = spreadsheet.worksheet("Cards")
        cards_ws.clear()
        cards_ws.append_row(CARDS_SCHEMA)
        for card_row in INITIAL_CARDS:
            cards_ws.append_row(card_row)
        print("\u2705 Cards sheet initialised with new schema")
    except Exception as e:
        print(f"Cards sheet setup error: {e}")

    _setup_dev_notes(existing)
    _setup_em_log(existing)

    print("\u2705 Sheets setup complete")
# ── Dev Notes & Em Log ────────────────────────────────────────────────────────

DEV_NOTES_CONTENT = [
    ["Section", "Content", "Last Updated"],
    ["Coding Standard — R1", "Route by cost: exact match → regex → keyword → cached lookup → live data → Claude. Never call Claude for routing decisions.", "2026-04-26"],
    ["Coding Standard — R2", "Single pass, immediate exit. Once a handler matches, execution stops. No fallback re-runs detectors. Loops exit on first match.", "2026-04-26"],
    ["Coding Standard — R3", "One sheet read per request, at the latest possible moment. Cached in memory, invalidated only on write. No sheet read during routing.", "2026-04-26"],
    ["Coding Standard — R4", "Haiku for classification/extraction/JSON under 200 tokens. Sonnet for reasoning/conversation/multi-step only. Set min max_tokens.", "2026-04-26"],
    ["Coding Standard — R5", "Typing indicator before every external API call. Parallel calls use asyncio.gather(). Nothing blocks the event loop.", "2026-04-26"],
    ["Coding Standard — R6", "Session messages never reach main routing chain. Session data is flat dict with explicit fields. Sessions time out cleanly.", "2026-04-26"],
    ["Coding Standard — R7", "Cache hierarchy: in-memory → cached sheet read → live sheet read → external API. Each layer reached only if above misses.", "2026-04-26"],
    ["Coding Standard — R8", "Every branch sets reply or returns. Empty reply is a bug. Claude fallback for genuine unknown intent only.", "2026-04-26"],
    ["Architecture — Routing", "Primary elif chain runs once. No else block re-runs detectors. Session handlers exit before reaching main router.", "2026-04-26"],
    ["Architecture — Caches", "_merchant_cache, _card_names_cache, _system_prompt_cache. Invalidate at write site only. Warm card cache on startup.", "2026-04-26"],
    ["Architecture — Models", "Haiku: is_calendar_request, parse_expense_text_v2, classification. Sonnet: conversation fallback, reasoning, market narrative.", "2026-04-26"],
    ["Architecture — Sheets", "setup_sheets() uses single worksheets() call. No repeated API reads in setup. Em Log and Dev Notes never read in routing.", "2026-04-26"],
    ["Deploy Flow", "Download bot.py from Claude chat → python ~/deploy.py ~/Downloads/bot.py 'message' → Railway auto-deploys. deploy.py lives in repo root.", "2026-04-30"],
    ["Handoff Rule", "4 lines max: (1) last deployed commit, (2) bot.py status, (3) mid-session context if hanging, (4) Read Dev Notes + Em Log before starting.", "2026-04-26"],
    ["Rule — Ship Rule", "Nothing ships until fully wired, tested, and deployed in the same session. Plan-only sessions that produce dead files are banned.", "2026-04-30"],
    ["Rule — Modularise", "Modularisation deferred until S3+S4 done. When implemented: all files deploy via single deploy.py call — no manual upload/download of multiple files.", "2026-04-30"],
    ["Roadmap — Session 9", "DONE. Trips schema migration (new schema: Trip ID · Destination · Currency · Check In · Check Out · Hotel Name · Hotel Local Name · Hotel Address · Notes · Status) + AviationStack strip (−307 lines).", "2026-04-30"],
    ["Roadmap — Session 10", "DONE. deploy.py built — runs locally, copies bot.py into repo, commits and pushes. One-time setup: git clone repo to ~/telegram-claude-bot, place deploy.py in ~/. Repo: roystenteng21/telegram-claude-bot. Token: 90-day ghp_ (no daily reset).", "2026-04-30"],
    ["Roadmap — Session 11 (S3 Reliability)", "Error boundaries: wrap all sheet writes, notify user on failure, never delete session before write confirmed (3x 🔴). Asyncio fixes: replace time.sleep with asyncio.sleep (2x 🟠). Error logging to Em Log at runtime. Persist sessions + price alerts across Railway restarts.", "2026-04-29"],
    ["Roadmap — Session 12 (S4 Efficiency)", "CRM cache + single-pass find_row (currently 1000 iterations/lookup). Reminder cache (currently full sheet read every minute). Routing deduplication: eliminate duplicate detectors, clean elif chain. Claude call audit: Haiku vs Sonnet, min max_tokens everywhere.", "2026-04-29"],
    ["Roadmap — Session 13 (S5 Completeness)", "End-to-end feature testing across all modules. Input forgiveness: broader natural phrasing recognised. Response consistency: emoji, error format, reply length standardised across all handlers.", "2026-04-29"],
]

EM_LOG_HEADERS_BACKLOG = ["Priority", "Item", "Stage", "Notes", "Added", "Status"]
EM_LOG_HEADERS_SESSION = ["Date", "Session", "Built", "Fixed", "Pending", "Commit"]

INITIAL_BACKLOG = [
    ["🔴", "log_expense: add error handling — financial data silently lost on sheet failure", "Step 3", "Wrap append_row in try/except, notify user if write fails, do not delete session until write confirmed", "2026-04-26", "🔲 Outstanding"],
    ["🔴", "Session deleted before write confirmed — expense unrecoverable on failure", "Step 3", "Move del receipt_confirm_sessions[user_id] to after log_expense succeeds", "2026-04-26", "🔲 Outstanding"],
    ["🔴", "save_merchant_memory silent fail — merchant never learned if sheet write fails", "Step 3", "Add error handling, log failure, do not silently swallow", "2026-04-26", "🔲 Outstanding"],
    ["🟠", "sheets_call_with_retry uses time.sleep(60) — blocks entire event loop", "Step 3", "Replace with asyncio.sleep(60) inside async context", "2026-04-26", "🔲 Outstanding"],
    ["🟠", "get_calendar uses time.sleep(3) on retry — blocks event loop on every calendar request", "Step 3", "Replace with asyncio.sleep or remove retry sleep", "2026-04-26", "🔲 Outstanding"],
    ["🟠", "find_row: 5 full passes over CRM records, no cache — 1000 iterations per lookup", "Step 3", "Single-pass with match tiers, add CRM cache invalidated on write", "2026-04-26", "🔲 Outstanding"],
    ["🟠", "check_and_fire_reminders: full sheet read every minute + find_row inside loop", "Step 3", "Cache pending reminders in memory, only re-read on write. Remove find_row from loop.", "2026-04-26", "🔲 Outstanding"],
    ["🟠", "Duplicate routing: is_reminder_request 3x, is_stock_request 2x, others twice", "Step 3", "Eliminate else block, merge missing handlers into primary elif chain", "2026-04-26", "🔲 Outstanding"],
    ["🟡", "Missing env var guard at startup — cryptic crash if TELEGRAM_TOKEN or ANTHROPIC_API_KEY unset", "Step 3", "Add explicit check and clear error message before app starts", "2026-04-26", "🔲 Outstanding"],
    ["🟡", "float() cast on unvalidated Claude output in parse_expense_text_v2 — unhandled exception", "Step 3", "Validate amount field before cast, return user-friendly error if invalid", "2026-04-26", "🔲 Outstanding"],
    ["🟡", "restore_overseas_from_trips: no fallback on corrupt data — silent bad state", "Step 3", "Wrap in try/except per field, skip row if malformed, log warning", "2026-04-26", "🔲 Outstanding"],
    ["🟡", "FX rates lost on Railway restart — user must re-enter manually after every redeploy", "Step 3", "Persist cached_fx_rates to Settings sheet, load on startup", "2026-04-26", "🔲 Outstanding"],
    ["🟡", "No timeout on RSS fetch in fetch_market_rss_headlines — indefinite hang possible", "Step 3", "Add timeout=10 to requests.get call", "2026-04-26", "🔲 Outstanding"],
    ["🟢", "_finalise_expense_session labelled legacy but still wired — dead code", "Step 3", "Remove function, update any callers", "2026-04-26", "🔲 Outstanding"],
    ["🟢", "CARD_FX_FEES dict defined but never referenced anywhere", "Step 3", "Remove or wire up to FX fee display", "2026-04-26", "🔲 Outstanding"],
    ["🟢", "bare except: in format_date and calculate_age swallows all exceptions", "Step 3", "Replace with except ValueError", "2026-04-26", "🔲 Outstanding"],
    ["🟢", "Haiku for parse_expense_text_v2 and is_calendar_request — Sonnet overkill", "Step 3", "Switch model to claude-haiku-3, verify output quality unchanged", "2026-04-26", "🔲 Outstanding"],
    ["🟢", "S5: End-to-end feature testing — no systematic test coverage across modules", "Step 3", "Test all features post-modularisation: expenses, CRM, trips, reminders, stocks, restaurants, calendar, todos", "2026-04-29", "🔲 Outstanding"],
    ["🟢", "S5: Input forgiveness — narrow phrasing recognition misses natural variants", "Step 3", "Broaden detector patterns for common commands; test with varied natural language inputs", "2026-04-29", "🔲 Outstanding"],
    ["🟢", "S5: Response consistency — emoji, error format, reply length vary across handlers", "Step 3", "Audit all reply strings; standardise error prefix, emoji usage, and length conventions", "2026-04-29", "🔲 Outstanding"],
]

INITIAL_SESSION = [
    ["2026-04-26", "Session 6",
     "Flight date fix (extract_flight_dates + AviationStack date param); Perf fixes (merchant cache, card cache, system prompt cache, history cap 20); AviationStack fallback; Dev Notes + Em Log tabs; em whats pending; setup_sheets single API call",
     "Flight dates ignored (now passed to AviationStack); Merchant map re-read every expense; Card names re-read every parse; System prompt rebuilt every Claude call; setup_sheets called worksheets() 4x",
     "All 19 issues (Step 3); Stage 3 trip features",
     "TBD"],
    ["2026-04-29", "Session 7",
     "Modularisation planning; dependency hierarchy defined; config.py, clients.py, state.py, sheets.py, profile.py, crm.py, expenses.py, reminders.py, calendar.py, todos.py, stocks.py, trips.py, meetings.py, scheduler.py built",
     "Monolithic bot.py (8426 lines) split into modules",
     "routing.py, restaurants.py, bot.py (entry point) remaining; deploy pending",
     "TBD"],
    ["2026-04-29", "Session 8",
     "routing.py (943 lines), restaurants.py (498 lines), bot.py entry point (29 lines) built; all 17 modules complete; roadmap S9–S13 logged to Dev Notes and Em Log",
     "Monolithic bot.py fully replaced by 17-module structure",
     "Deploy 17 modules via git; Session 9: Trips schema migration + AviationStack strip",
     "TBD"],
]


def _setup_dev_notes(existing):
    """Create or overwrite Dev Notes tab with current coding standard and architecture rules.
    Always overwrites — rules are versioned by Last Updated column, not by appending."""
    try:
        if "Dev Notes" not in existing:
            ws = spreadsheet.add_worksheet(title="Dev Notes", rows=50, cols=3)
            existing.append("Dev Notes")
            print("Created Dev Notes tab")
        else:
            ws = spreadsheet.worksheet("Dev Notes")
        ws.clear()
        ws.update(range_name='A1', values=DEV_NOTES_CONTENT)
        # Format header row bold
        try:
            ws.format('A1:C1', {'textFormat': {'bold': True}})
        except Exception:
            pass
        print("✅ Dev Notes populated")
    except Exception as e:
        print(f"Dev Notes setup error: {e}")


def _setup_em_log(existing):
    """Create Em Log tab with Backlog and Session History sections.
    Backlog: max 10 items — enforced on every write.
    Session History: max 10 rows — oldest deleted when 11th added.
    No archive — hard delete keeps sheet lean permanently."""
    try:
        if "Em Log" not in existing:
            ws = spreadsheet.add_worksheet(title="Em Log", rows=200, cols=6)
            existing.append("Em Log")
            print("Created Em Log tab")
        else:
            ws = spreadsheet.worksheet("Em Log")
            # Only populate if empty — don't overwrite live session data
            if ws.row_values(1):
                print("Em Log already populated — skipping init")
                return

        # Section 1: Backlog
        ws.append_row(["── BACKLOG (max 10) ──", "", "", "", "", ""])
        ws.append_row(EM_LOG_HEADERS_BACKLOG)
        for row in INITIAL_BACKLOG[:10]:  # enforce cap at init
            ws.append_row(row)

        # Section divider
        ws.append_row(["", "", "", "", "", ""])
        ws.append_row(["── SESSION HISTORY (max 10) ──", "", "", "", "", ""])
        ws.append_row(EM_LOG_HEADERS_SESSION)
        for row in INITIAL_SESSION:
            ws.append_row(row)

        # Format section headers
        try:
            ws.format('A1:F1', {'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 0.9}})
        except Exception:
            pass
        print("✅ Em Log populated")
    except Exception as e:
        print(f"Em Log setup error: {e}")


def em_log_sheet():
    return get_sheet("Em Log")

def log_error_to_em_log(source: str, error: str):
    """Append a runtime error row to Em Log backlog section. Fire-and-forget — never raises."""
    try:
        ws = em_log_sheet()
        if not ws:
            return
        today = datetime.now(pytz.timezone("Asia/Kuala_Lumpur")).strftime("%Y-%m-%d %H:%M")
        # Append directly — no cap enforcement, errors are always kept
        ws.append_row(["🔴 ERROR", f"[{source}] {error}", "Runtime", today, today])
    except Exception:
        pass  # Never let error logging itself crash anything


def add_session_to_em_log(date_str, session_name, built, fixed, pending, commit):
    """Append a session row to Em Log. Enforces 10-row cap on session history — deletes oldest if exceeded."""
    try:
        ws = em_log_sheet()
        if not ws:
            return
        all_values = ws.get_all_values()

        # Find session history section header row
        session_header_row = None
        for i, row in enumerate(all_values):
            if row and "SESSION HISTORY" in str(row[0]):
                session_header_row = i + 1  # 1-indexed
                break
        if not session_header_row:
            return

        # Find all session data rows (after the column headers row)
        data_start = session_header_row + 2  # skip section header + column headers
        session_rows = [i + 1 for i, row in enumerate(all_values)
                        if i >= data_start and any(row)]

        # Enforce 10-row cap — delete oldest first
        while len(session_rows) >= 10:
            ws.delete_rows(session_rows[0])
            session_rows.pop(0)

        ws.append_row([date_str, session_name, built, fixed, pending, commit])
        print(f"Session logged to Em Log: {session_name}")
    except Exception as e:
        print(f"add_session_to_em_log error: {e}")


def add_backlog_item(priority, item, stage="", notes=""):
    """Add item to Em Log backlog. Enforces 10-item cap — lowest priority item deleted if exceeded."""
    try:
        ws = em_log_sheet()
        if not ws:
            return "Em Log sheet not found"
        all_values = ws.get_all_values()

        # Find backlog section
        backlog_header_row = None
        backlog_col_header_row = None
        for i, row in enumerate(all_values):
            if row and "BACKLOG" in str(row[0]):
                backlog_header_row = i + 1
            if backlog_header_row and row and row[0] == "Priority":
                backlog_col_header_row = i + 1
                break
        if not backlog_col_header_row:
            return "Backlog section not found in Em Log"

        # Find backlog data rows
        backlog_rows = []
        for i, row in enumerate(all_values):
            if i >= backlog_col_header_row and any(row) and "SESSION" not in str(row[0]) and row[0] != "Priority":
                backlog_rows.append(i + 1)
            if row and "SESSION HISTORY" in str(row[0]):
                break

        # Enforce 10-item cap
        if len(backlog_rows) >= 10:
            ws.delete_rows(backlog_rows[-1])  # delete lowest priority (last row)

        today = date.today().strftime("%Y-%m-%d")
        # Insert before session divider — find the blank divider row
        divider_row = None
        for i, row in enumerate(all_values):
            if i >= backlog_col_header_row and not any(row):
                divider_row = i + 1
                break

        if divider_row:
            ws.insert_row([priority, item, stage, notes, today, "🔲 Outstanding"], divider_row)
        else:
            ws.append_row([priority, item, stage, notes, today, "🔲 Outstanding"])

        return f"Added to backlog ✅"
    except Exception as e:
        return f"add_backlog_item error: {e}"


def get_pending_backlog():
    """Return formatted backlog from Em Log — Outstanding items only."""
    try:
        ws = em_log_sheet()
        if not ws:
            return "Em Log not found"
        all_values = ws.get_all_values()

        backlog_items = []
        in_backlog = False
        for row in all_values:
            if not row:
                continue
            if "BACKLOG" in str(row[0]):
                in_backlog = True
                continue
            if "SESSION HISTORY" in str(row[0]):
                break
            if in_backlog and row[0] not in ("Priority", "── BACKLOG (max 10) ──", ""):
                priority = row[0] if len(row) > 0 else ""
                item = row[1] if len(row) > 1 else ""
                stage = row[2] if len(row) > 2 else ""
                status = row[5] if len(row) > 5 else "🔲 Outstanding"
                if item and "Done" not in status:
                    backlog_items.append((priority, item, stage))

        if not backlog_items:
            return "Backlog is empty ✅"

        lines = ["*Backlog — Outstanding*\n"]
        for priority, item, stage in backlog_items:
            stage_str = f" _[{stage}]_" if stage else ""
            lines.append(f"{priority} {item}{stage_str}")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Couldn't read backlog: {e}"


def _migrate_crm_headers(ws, old_headers, new_headers):
    """Migrate CRM sheet from old column layout to new layout."""
    try:
        all_data = ws.get_all_values()
        if not all_data:
            ws.update(range_name='A1', values=[new_headers])
            print("CRM headers initialised (empty sheet)")
            return

        old_h = all_data[0]
        rows = all_data[1:]

        # Build mapping from old col name -> index
        old_idx = {h: i for i, h in enumerate(old_h)}

        def get_old(row, col_name):
            i = old_idx.get(col_name)
            return row[i] if i is not None and i < len(row) else ""

        migrated = [new_headers]
        for row in rows:
            name = get_old(row, "Name")
            if not name:
                continue
            # Map old fields to new; old "Where We Met" -> Relationship
            migrated.append([
                name,
                get_old(row, "Alias"),
                get_old(row, "Birthday"),
                get_old(row, "Where We Met"),   # becomes Relationship
                get_old(row, "Context"),
                get_old(row, "Notes"),
                get_old(row, "Follow Up Date"),
                get_old(row, "Follow Up Notes"),
                get_old(row, "Last Updated"),
                get_old(row, "Birthday Greeted"),
                get_old(row, "Referred By"),
                get_old(row, "Referral Date"),
                get_old(row, "Email"),
                get_old(row, "Address"),
            ])

        ws.clear()
        if migrated:
            ws.update(range_name='A1', values=migrated)
        print(f"✅ CRM migrated to new column layout ({len(migrated)-1} contacts)")
    except Exception as e:
        print(f"Error migrating CRM headers: {e}")

def get_or_create_drive_folder(name, parent_id=None):
    """Get a Drive folder by name (under parent), or create it."""
    query = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"
    results = drive_service.files().list(
        q=query, fields="files(id, name)",
        supportsAllDrives=True, includeItemsFromAllDrives=True
    ).execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    # Create it
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder"
    }
    if parent_id:
        meta["parents"] = [parent_id]
    folder = drive_service.files().create(
        body=meta, fields="id", supportsAllDrives=True
    ).execute()
    print(f"Created Drive folder: {name}")
    return folder["id"]

RECEIPTS_FOLDER_ID = os.getenv("RECEIPTS_FOLDER_ID", "14pG1lNPANRwehiW_xSFjoHzt-AUk5-Xb")

def setup_drive():
    """Create Em's Drive folder structure."""
    em_id = get_or_create_drive_folder("Em")
    meeting_notes_id = get_or_create_drive_folder("Meeting Notes", em_id)
    backups_id = get_or_create_drive_folder("Backups", em_id)
    settings_id = get_or_create_drive_folder("Settings", em_id)
    print("✅ Drive folders setup complete")
    return {
        "em": em_id,
        "receipts": RECEIPTS_FOLDER_ID,  # shared folder — no service account quota issues
        "meeting_notes": meeting_notes_id,
        "backups": backups_id,
        "settings": settings_id
    }

# Global Drive folder IDs (set during setup)
DRIVE_FOLDERS = {}

def setup_em_profile():
    """Load em_profile from Settings sheet, or create it if missing.
    Stored as a single row: key='em_profile', value=JSON string.
    Service accounts can't upload files to Drive, so we use Sheets instead.
    """
    global em_profile

    default_profile = {
        "version": "1.0",
        "created": date.today().strftime("%d/%m/%Y"),
        "tone": "warm, casual, never formal",
        "no_dashes_in_conversation": True,
        "emoji_style": "contextual and varied, never overdone",
        "greeting_options": ["hey", "yo", "alright", "aite", "sup", "oi"],
        "closing_style": "varied, natural, never repetitive",
        "no_repeat_phrases": True,
        "info_per_line": True,
        "silent_tracking": True,
        "language": "natural flowing sentences, no bullet dumps",
        "forbidden_phrases": ["cool cool", "certainly", "of course", "absolutely", "great question"],
        "forbidden_emojis": ["🤙"],
        "response_length": "concise by default, elaborate only when asked",
        "multi_language": True,
        "dnd_active": False,
        "dnd_held_messages": [],
        "overseas_mode": False,
        "overseas_destination": "",
        "overseas_currency": "SGD",
        "overseas_return_date": "",
        "preferences": {
            "birthday_greeting_style": "warm and casual, opens with Happy birthday",
            "expense_confirmation_emoji": "contextual and varied",
            "reminder_default_time": "09:00",
            "weekly_market_summary_day": "Monday",
            "weekly_market_summary_time": "08:00"
        },
        "learned_style": {}
    }

    try:
        settings_sheet = spreadsheet.worksheet("Settings")
        records = settings_sheet.get_all_records()

        # Look for existing em_profile row
        for i, r in enumerate(records):
            if r.get("Key") == "em_profile":
                em_profile = json.loads(r.get("Value", "{}"))
                print("✅ Loaded em_profile from Settings sheet")
                return

        # Not found — write default
        em_profile = default_profile
        settings_sheet.append_row(["em_profile", json.dumps(default_profile)])
        print("✅ Created em_profile in Settings sheet")

    except Exception as e:
        # Fallback to default in memory if sheet not ready yet
        em_profile = default_profile
        print(f"Warning: Could not load em_profile from sheet: {e}. Using defaults.")

def save_em_profile():
    """Save current em_profile back to Settings sheet."""
    try:
        settings_sheet = spreadsheet.worksheet("Settings")
        records = settings_sheet.get_all_records()
        for i, r in enumerate(records):
            if r.get("Key") == "em_profile":
                settings_sheet.update_cell(i + 2, 2, json.dumps(em_profile))
                return
        # Not found — append it
        settings_sheet.append_row(["em_profile", json.dumps(em_profile)])
    except Exception as e:
        print(f"Error saving em_profile: {e}")

def run_infrastructure_setup():
    """Run all setup steps on startup. Idempotent — safe to run every time."""
    global DRIVE_FOLDERS
    print("Running infrastructure setup...")

    health = {
        "Sheets": "✅ Connected",
        "Drive": "✅ Connected",
        "iCloud": "✅ Connected",
        "Scheduler": "✅ Running",
        "Profile": "✅ Loaded",
    }

    # Sheets
    try:
        setup_sheets()
    except Exception as e:
        health["Sheets"] = f"❌ Failed — {str(e)[:40]}"
        print(f"setup_sheets error: {e}")

    # Drive
    try:
        DRIVE_FOLDERS = setup_drive()
    except Exception as e:
        health["Drive"] = "❌ Failed — receipt uploads won't work"
        DRIVE_FOLDERS = {}
        print(f"setup_drive error: {e}")

    # Profile — attempt with retry
    try:
        setup_em_profile()
        if not em_profile:
            raise Exception("Profile empty after load")
    except Exception:
        # Brief pause before retry — acceptable at startup (sync context, not in event loop yet)
        import time as _time
        _time.sleep(1)
        try:
            setup_em_profile()
            if not em_profile:
                raise Exception("Profile empty after retry")
        except Exception as e2:
            health["Profile"] = "❌ Failed — running on defaults"
            print(f"em_profile load failed after retry: {e2}")

    # iCloud — tested lazily on first use, mark unknown here
    health["iCloud"] = "⚠️ Not tested yet"

    print("✅ Infrastructure setup complete")
    return health

async def send_startup_message(app, health):
    """Send startup health check if this is a new Railway deployment."""
    try:
        settings_ws = spreadsheet.worksheet("Settings")
        records = settings_ws.get_all_records()
        last_deploy_id = ""
        last_deploy_row = None
        for i, r in enumerate(records, start=2):
            if r.get("Key") == "last_deployment_id":
                last_deploy_id = r.get("Value", "")
                last_deploy_row = i
                break

        # Only send if deployment ID changed
        if RAILWAY_DEPLOYMENT_ID and RAILWAY_DEPLOYMENT_ID == last_deploy_id:
            return  # Same deployment restart — silent

        # New deploy — update stored ID
        if RAILWAY_DEPLOYMENT_ID:
            if last_deploy_row:
                settings_ws.update_cell(last_deploy_row, 2, RAILWAY_DEPLOYMENT_ID)
            else:
                settings_ws.append_row(["last_deployment_id", RAILWAY_DEPLOYMENT_ID])

        # Build status message — match em status format exactly
        issues = []
        for k, v in health.items():
            if k == "iCloud":
                continue  # iCloud tested lazily, skip on startup
            if "❌" in v:
                issues.append(f"• {k}: {v.replace('❌ ', '')}")
            elif "⚠️" in v:
                issues.append(f"• {k}: {v.replace('⚠️ ', '')}")

        if "❌" in health.get("Profile", ""):
            issues.append("• Reply 'reload profile' to retry loading preferences.")

        if not issues:
            msg = "✅ Systems all green"
        else:
            msg = "⚠️ Issues detected on startup:\n\n" + "\n".join(issues)

        await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)

    except Exception as e:
        print(f"send_startup_message error: {e}")

# --- Sheet References (after setup) ---
def get_sheet(name):
    try:
        return spreadsheet.worksheet(name)
    except Exception as e:
        print(f"Warning: Could not get sheet '{name}': {e}")
        return None

# We reference these after setup runs
def crm_sheet():
    return get_sheet("CRM")

def todo_sheet():
    return get_sheet("Todos")

# --- Anthropic Setup ---
client = Anthropic(api_key=ANTHROPIC_API_KEY)
conversation_histories = {}
edit_sessions = {}

def track_anthropic_call(success):
    """Track Anthropic API call success/failure for downtime detection."""
    global _anthropic_failure_count, _anthropic_down_notified
    if success:
        if _anthropic_down_notified:
            _anthropic_down_notified = False
            # Recovery will be notified by the next scheduled check
        _anthropic_failure_count = 0
    else:
        _anthropic_failure_count += 1

async def notify_anthropic_down(app):
    """Send Anthropic API down notification if threshold reached."""
    global _anthropic_down_notified
    if _anthropic_failure_count >= ANTHROPIC_FAILURE_THRESHOLD and not _anthropic_down_notified:
        _anthropic_down_notified = True
        try:
            await app.bot.send_message(
                chat_id=YOUR_CHAT_ID,
                text=(
                    "⚠️ Anthropic API appears to be down — some features unavailable:\n"
                    "• Receipt vision parsing\n"
                    "• Expense text parsing\n"
                    "• Birthday greetings\n"
                    "• Reminder parsing\n"
                    "• Market narrative\n\n"
                    "I'll notify you when it recovers."
                )
            )
        except Exception as e:
            print(f"notify_anthropic_down error: {e}")

def sheets_call_with_retry(fn, *args, max_retries=2, **kwargs):
    """Wrap a Google Sheets API call with quota retry logic.
    Sleep is kept short (5s) — this is a sync function and cannot use asyncio.sleep.
    Quota errors are rare; long sleeps block the entire event loop."""
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            err_str = str(e).lower()
            if "quota" in err_str or "rate" in err_str or "429" in err_str:
                print(f"Sheets quota hit — retrying immediately (attempt {attempt+1})")
                # No sleep here: sheets_call_with_retry is sync and called from async context.
                # Sleeping here blocks the event loop. Quota hits are rare; immediate retry is fine.
            else:
                raise
    raise Exception("Google Sheets rate limit exceeded after retries.")


def get_calendar(name=None):
    global _icloud_down
    import time as _time

    def _attempt():
        caldav_client = caldav.DAVClient(
            url="https://caldav.icloud.com",
            username=ICLOUD_USERNAME,
            password=ICLOUD_PASSWORD
        )
        principal = caldav_client.principal()
        calendars = principal.calendars()
        if name:
            for cal in calendars:
                if name.lower() in cal.name.lower():
                    return cal
            return None
        return calendars[0] if calendars else None

    # Attempt 1
    try:
        result = _attempt()
        _icloud_down = False
        return result
    except Exception:
        pass

    # Retry immediately (removed blocking sleep — was freezing event loop)
    try:
        result = _attempt()
        _icloud_down = False
        return result
    except Exception as e:
        _icloud_down = True
        print(f"iCloud Calendar unavailable after retry: {e}")
        return None

async def check_icloud_daily(app):
    """Daily check — notify if iCloud Calendar still down."""
    try:
        global _icloud_down, _icloud_last_notified
        today = date.today()
        if _icloud_down:
            if _icloud_last_notified != today:
                _icloud_last_notified = today
                await app.bot.send_message(
                    chat_id=YOUR_CHAT_ID,
                    text="⚠️ iCloud Calendar still unavailable — calendar features won't work.\nCheck your credentials in Railway."
                )
        else:
            # Test connection
            cal = get_calendar()
            if cal is None and _icloud_last_notified != today:
                _icloud_last_notified = today
                await app.bot.send_message(
                    chat_id=YOUR_CHAT_ID,
                    text="⚠️ iCloud Calendar unavailable — calendar features won't work.\nCheck your credentials in Railway."
                )
    except Exception as e:
        print(f"check_icloud_daily error: {e}")



# --- Helpers ---
def format_date(date_str):
    try:
        d = datetime.strptime(date_str, "%d/%m/%Y")
        return d.strftime("%d %b %Y")
    except (ValueError, TypeError):
        return date_str

def calculate_age(birthday_str):
    try:
        bday = datetime.strptime(birthday_str, "%d/%m/%Y").date()
        today = date.today()
        age = today.year - bday.year - ((today.month, today.day) < (bday.month, bday.day))
        return str(age)
    except (ValueError, TypeError):
        return ""

def format_contact(r, show_private=False):
    name = r.get("Name", "")
    alias = r.get("Alias", "")
    birthday = r.get("Birthday", "")
    age = calculate_age(birthday) if birthday else ""
    relationship = r.get("Relationship", "")
    context = r.get("Context", "")
    notes_raw = r.get("Notes", "")
    followup_date = r.get("Follow Up Date", "")
    followup_notes = r.get("Follow Up Notes", "")
    last_updated = r.get("Last Updated", "")
    referred_by = r.get("Referred By", "")
    referral_date = r.get("Referral Date", "")
    email = r.get("Email", "")
    address = r.get("Address", "")

    lines = [f"*{name}*"]
    if alias:
        lines.append(f"- Also known as: {alias}")
    if relationship:
        lines.append(f"- Relationship: {relationship}")
    if context:
        lines.append(f"- Context: {context}")
    if birthday and age:
        lines.append(f"- Birthday: {format_date(birthday)} (age {age})")
    elif birthday:
        lines.append(f"- Birthday: {format_date(birthday)}")

    if notes_raw:
        note_items = [n.strip() for n in notes_raw.split(";") if n.strip()]
        lines.append("- Notes:")
        for n in note_items:
            lines.append(f"  - {n}")
    else:
        lines.append("- Notes:\n  - None")

    if followup_date:
        lines.append(f"- Follow Up: {format_date(followup_date)}")
        if followup_notes:
            lines.append(f"  - {followup_notes}")

    if referred_by:
        ref_line = f"- Referred by: {referred_by}"
        if referral_date:
            ref_line += f" on {format_date(referral_date)}"
        lines.append(ref_line)

    # Private fields — only shown when explicitly asked
    if show_private:
        if email:
            lines.append(f"- Email: {email}")
        if address:
            lines.append(f"- Address: {address}")

    if last_updated:
        lines.append(f"\n_Last updated: {format_date(last_updated)}_")

    return "\n".join(lines)

# CRM record cache — invalidated on any write; TTL: 5 minutes
_crm_cache = None
_crm_cache_ts = None
_CRM_CACHE_TTL = 300  # seconds

def _invalidate_crm_cache():
    global _crm_cache, _crm_cache_ts
    _crm_cache = None
    _crm_cache_ts = None

def _get_crm_records():
    """Return CRM records from cache, fetching from sheet if stale or TTL expired."""
    global _crm_cache, _crm_cache_ts
    import time as _time
    now = _time.monotonic()
    if _crm_cache is None or _crm_cache_ts is None or (now - _crm_cache_ts) > _CRM_CACHE_TTL:
        sheet = crm_sheet()
        _crm_cache = sheet.get_all_records() if sheet else []
        _crm_cache_ts = now
    return _crm_cache

def find_row(name):
    """Single-pass CRM lookup across Name, Alias, and first name.
    Returns (row_num, record) for single match, or ('disambig', message) for multiple.
    Uses in-memory cache — invalidated on any CRM write."""
    records = _get_crm_records()
    name_lower = name.strip().lower()

    exact, alias_exact, sub_name, sub_alias, first = [], [], [], [], []

    for i, r in enumerate(records):
        full = r.get("Name", "")
        alias = r.get("Alias", "")
        full_l = full.lower()
        alias_l = alias.lower()
        first_name = full_l.split()[0] if full_l else ""
        alias_first = alias_l.split()[0] if alias_l else ""

        if full_l == name_lower:
            return i + 2, r                          # exact name — return immediately
        if alias_l == name_lower:
            alias_exact.append((i + 2, r))
        elif name_lower in full_l:
            sub_name.append((i + 2, r))
        elif name_lower in alias_l:
            sub_alias.append((i + 2, r))
        elif name_lower in (first_name, alias_first):
            first.append((i + 2, r))

    for tier in (alias_exact, sub_name, sub_alias, first):
        if len(tier) == 1:
            return tier[0]
        if len(tier) > 1:
            return "disambig", disambiguate_contacts(tier)

    return None, None

def find_all_rows(name):
    """Like find_row but returns all matches for disambiguation."""
    records = _get_crm_records()
    name_lower = name.strip().lower()
    results = []

    for i, r in enumerate(records):
        full_name = r.get("Name", "").lower()
        alias = r.get("Alias", "").lower()
        first_name = full_name.split()[0] if full_name else ""
        alias_first = alias.split()[0] if alias else ""
        if (name_lower in full_name or name_lower in alias or
                name_lower == first_name or name_lower == alias_first):
            results.append((i + 2, r))
    return results

def disambiguate_contacts(matches):
    """Return a disambiguation prompt if multiple contacts match."""
    names = [r.get("Name", "?") for _, r in matches]
    options = " or ".join(f"*{n}*" for n in names)
    return f"Did you mean {options}?"

# --- CRM Functions ---
# Pending duplicate contact saves: { user_id: { "data": str, "existing_name": str } }
pending_contact_saves = {}
pending_restaurant_saves = {}  # { user_id: { "name": str, "location": str, ... } }

def save_contact(data, force_new=False):
    try:
        sheet = crm_sheet()
        parts = [p.strip() for p in data.split(",")]
        while len(parts) < 13:
            parts.append("")
        name = parts[0]
        alias = parts[1] if len(parts) > 1 else ""
        birthday = parts[2] if len(parts) > 2 else ""
        relationship = parts[3] if len(parts) > 3 else ""
        context = parts[4] if len(parts) > 4 else ""
        notes = parts[5] if len(parts) > 5 else ""
        followup_date = parts[6] if len(parts) > 6 else ""
        followup_notes = parts[7] if len(parts) > 7 else ""
        last_updated = date.today().strftime("%d/%m/%Y")
        if not name:
            return "❌ Name is required"

        # Duplicate check unless force_new
        if not force_new:
            existing_row, existing_record = find_row(name)
            if existing_record and existing_row != "disambig":
                return f"_DUPLICATE_:{name}"  # Signal to caller to handle

        sheet.append_row([
            name, alias, birthday, relationship, context, notes,
            followup_date, followup_notes, last_updated, "", "", "", "", ""
        ])
        _invalidate_crm_cache()
        return f"✅ Contact saved!\n\n" + format_contact({
            "Name": name, "Alias": alias, "Birthday": birthday,
            "Relationship": relationship, "Context": context,
            "Notes": notes, "Follow Up Date": followup_date,
            "Follow Up Notes": followup_notes, "Last Updated": last_updated
        })
    except Exception as e:
        return f"❌ Error saving contact: {str(e)}"

def find_contact(name, show_private=False):
    try:
        results = find_all_rows(name)
        if not results:
            return f"❌ No contact found for '{name}'"
        if len(results) > 1:
            return disambiguate_contacts(results)
        return format_contact(results[0][1], show_private=show_private)
    except Exception as e:
        return f"❌ Error finding contact: {str(e)}"

def add_note(data):
    try:
        sheet = crm_sheet()
        parts = data.split("-", 1)
        if len(parts) < 2:
            return "❌ Format: note Name - your note here"
        name = parts[0].strip()
        note = parts[1].strip()
        row_num, record = find_row(name)
        if not record:
            return f"❌ No contact found for '{name}'"
        existing = record.get("Notes", "")
        new_note = f"{existing}; {note}" if existing else note
        # Col 6 = Notes in new schema
        sheet.update_cell(row_num, 6, new_note)
        sheet.update_cell(row_num, 9, date.today().strftime("%d/%m/%Y"))  # Last Updated
        _invalidate_crm_cache()
        return f"✅ Note added to *{record.get('Name')}*"
    except Exception as e:
        return f"❌ Error adding note: {str(e)}"

def set_followup(data):
    try:
        sheet = crm_sheet()
        parts = [p.strip() for p in data.split(",", 2)]
        if len(parts) < 2:
            return "❌ Format: followup Name, DD/MM/YYYY, notes"
        name = parts[0]
        followup_date = parts[1]
        followup_notes = parts[2] if len(parts) > 2 else ""
        row_num, record = find_row(name)
        if not record:
            return f"❌ No contact found for '{name}'"
        sheet.update_cell(row_num, 7, followup_date)   # Follow Up Date
        sheet.update_cell(row_num, 8, followup_notes)   # Follow Up Notes
        sheet.update_cell(row_num, 9, date.today().strftime("%d/%m/%Y"))  # Last Updated
        return f"✅ Follow up set for *{record.get('Name')}* on {format_date(followup_date)}"
    except Exception as e:
        return f"❌ Error setting follow up: {str(e)}"

def update_field(data):
    try:
        sheet = crm_sheet()
        parts = [p.strip() for p in data.split(",", 2)]
        if len(parts) < 3:
            return "❌ Format: update Name, field, new value"
        name, field, value = parts
        # Col indices: Name=1, Alias=2, Birthday=3, Relationship=4, Context=5,
        # Notes=6, Follow Up Date=7, Follow Up Notes=8, Last Updated=9,
        # Birthday Greeted=10, Referred By=11, Referral Date=12, Email=13, Address=14
        field_map = {
            "alias": 2, "birthday": 3, "relationship": 4, "context": 5,
            "notes": 6, "follow up date": 7, "follow up notes": 8,
            "referred by": 11, "referral date": 12, "email": 13, "address": 14
        }
        col = field_map.get(field.lower())
        if not col:
            valid = ", ".join(field_map.keys())
            return f"❌ Unknown field '{field}'. Options: {valid}"
        row_num, record = find_row(name)
        if not record:
            return f"❌ No contact found for '{name}'"
        sheet.update_cell(row_num, col, value)
        sheet.update_cell(row_num, 9, date.today().strftime("%d/%m/%Y"))  # Last Updated
        _invalidate_crm_cache()
        return f"✅ {field.title()} updated for *{record.get('Name')}*"
    except Exception as e:
        return f"❌ Error updating contact: {str(e)}"

def update_contact_field_natural(name, field, value):
    """Update a single field by natural language field name. Returns reply string."""
    return update_field(f"{name}, {field}, {value}")

def delete_contact(name):
    try:
        sheet = crm_sheet()
        row_num, record = find_row(name)
        if not record:
            # Fallback: raw scan for stub entries (e.g. malformed rows with no proper Name)
            all_values = sheet.get_all_values()
            name_lower = name.strip().lower()
            for i, row in enumerate(all_values[1:], start=2):  # skip header
                if any(name_lower == str(cell).strip().lower() for cell in row if cell):
                    display = next((str(c).strip() for c in row if c), name)
                    sheet.delete_rows(i)
                    _invalidate_crm_cache()
                    return f"✅ Entry *{display}* deleted"
            return f"❌ No contact found for '{name}'"
        sheet.delete_rows(row_num)
        _invalidate_crm_cache()
        return f"✅ Contact *{record.get('Name')}* deleted"
    except Exception as e:
        return f"❌ Error deleting contact: {str(e)}"

def search_contacts(keyword):
    try:
        records = _get_crm_records()
        results = []
        for r in records:
            if any(keyword.lower() in str(v).lower() for v in r.values()):
                results.append(r)
        if not results:
            return f"No contacts found matching '{keyword}' in the CRM."
        return f"🔍 *{len(results)} result(s) for '{keyword}':*\n\n" + "\n\n".join(format_contact(r) for r in results)
    except Exception as e:
        return f"❌ Error searching: {str(e)}"

def list_contacts():
    try:
        records = _get_crm_records()
        if not records:
            return "❌ No contacts found"
        response = f"📋 *{len(records)} contact(s):*\n\n"
        for r in records:
            rel = r.get("Relationship", "") or r.get("Context", "") or "Unknown"
            response += f"👤 {r.get('Name', '')} — {rel}\n"
        return response
    except Exception as e:
        return f"❌ Error listing contacts: {str(e)}"

def get_stats():
    try:
        records = _get_crm_records()
        today = date.today()
        total = len(records)
        followups_due = 0
        birthdays_month = 0
        for r in records:
            fu = r.get("Follow Up Date", "")
            if fu:
                try:
                    fu_date = datetime.strptime(fu, "%d/%m/%Y").date()
                    if fu_date >= today:
                        followups_due += 1
                except Exception as e:
                    print(f"get_stats: bad follow up date in row: {e}")
            bday = r.get("Birthday", "")
            if bday:
                try:
                    b = datetime.strptime(bday, "%d/%m/%Y").date()
                    this_year = b.replace(year=today.year)
                    if this_year < today:
                        this_year = b.replace(year=today.year + 1)
                    if (this_year - today).days <= 30:
                        birthdays_month += 1
                except Exception as e:
                    print(f"get_stats: bad birthday in row: {e}")
        return (
            f"📊 *CRM Stats*\n\n"
            f"👥 Total contacts: {total}\n"
            f"📅 Upcoming follow ups: {followups_due}\n"
            f"🎂 Birthdays in next 30 days: {birthdays_month}"
        )
    except Exception as e:
        return f"❌ Error getting stats: {str(e)}"

def upcoming_followups():
    try:
        records = _get_crm_records()
        today = date.today()
        upcoming = []
        for r in records:
            fu_date = r.get("Follow Up Date", "")
            if fu_date:
                try:
                    fu = datetime.strptime(fu_date, "%d/%m/%Y").date()
                    if fu >= today:
                        upcoming.append((fu, r))
                except Exception as e:
                    print(f"upcoming_followups: bad date in row: {e}")
        if not upcoming:
            return "✅ No upcoming follow ups!"
        upcoming.sort(key=lambda x: x[0])
        response = "📅 *Upcoming Follow Ups:*\n\n"
        for fu, r in upcoming:
            response += f"👤 *{r.get('Name')}* — {format_date(r.get('Follow Up Date'))}\n"
            if r.get('Follow Up Notes'):
                response += f"  - {r.get('Follow Up Notes')}\n"
            response += "\n"
        return response
    except Exception as e:
        return f"❌ Error fetching follow ups: {str(e)}"

def overdue_followups():
    try:
        records = _get_crm_records()
        today = date.today()
        overdue = []
        for r in records:
            fu_date = r.get("Follow Up Date", "")
            if fu_date:
                try:
                    fu = datetime.strptime(fu_date, "%d/%m/%Y").date()
                    if fu < today:
                        overdue.append((fu, r))
                except Exception as e:
                    print(f"overdue_followups: bad date in row: {e}")
        if not overdue:
            return "✅ No overdue follow ups!"
        overdue.sort(key=lambda x: x[0])
        response = "⚠️ *Overdue Follow Ups:*\n\n"
        for fu, r in overdue:
            days_ago = (today - fu).days
            response += f"👤 *{r.get('Name')}* — {format_date(r.get('Follow Up Date'))} ({days_ago} days ago)\n"
            if r.get('Follow Up Notes'):
                response += f"  - {r.get('Follow Up Notes')}\n"
            response += "\n"
        return response
    except Exception as e:
        return f"❌ Error fetching overdue: {str(e)}"

def upcoming_birthdays(days=30):
    try:
        records = _get_crm_records()
        today = date.today()
        upcoming = []
        for r in records:
            bday_str = r.get("Birthday", "")
            if bday_str:
                try:
                    bday = datetime.strptime(bday_str, "%d/%m/%Y").date()
                    this_year = bday.replace(year=today.year)
                    if this_year < today:
                        this_year = bday.replace(year=today.year + 1)
                    days_away = (this_year - today).days
                    if days_away <= days:
                        upcoming.append((days_away, r))
                except Exception as e:
                    print(f"upcoming_birthdays: bad date in row: {e}")
        if not upcoming:
            return f"🎂 No birthdays in the next {days} days!"
        upcoming.sort(key=lambda x: x[0])
        label = "next 7 days" if days == 7 else "next 30 days"
        response = f"🎂 *Birthdays in the {label}:*\n\n"
        for days_away, r in upcoming:
            response += f"👤 *{r.get('Name')}* — {format_date(r.get('Birthday'))} ({'today! 🎉' if days_away == 0 else f'in {days_away} days'})\n"
        return response
    except Exception as e:
        return f"❌ Error fetching birthdays: {str(e)}"

def last_contact(name):
    try:
        row_num, record = find_row(name)
        if not record:
            return f"❌ No contact found for '{name}'"
        last = record.get("Last Updated", "")
        if last:
            return f"👤 *{record.get('Name')}*\n📅 Last updated: {format_date(last)}"
        return f"👤 *{record.get('Name')}*\n📅 No updates recorded yet"
    except Exception as e:
        return f"❌ Error: {str(e)}"

# --- Referral Tracking ---
def set_referral(referrer_name, referred_name):
    """Record that referrer_name referred referred_name."""
    try:
        sheet = crm_sheet()
        row_num, record = find_row(referred_name)
        if not record:
            return f"❌ No contact found for '{referred_name}'. Add them first."
        today = date.today().strftime("%d/%m/%Y")
        sheet.update_cell(row_num, 11, referrer_name)   # Referred By
        sheet.update_cell(row_num, 12, today)            # Referral Date
        sheet.update_cell(row_num, 9, today)             # Last Updated
        # Set relationship to Prospect if blank
        if not record.get("Relationship", ""):
            sheet.update_cell(row_num, 4, "Prospect")
        return f"✅ Got it — {referred_name} was referred by {referrer_name} (Referral Date: {format_date(today)})"
    except Exception as e:
        return f"❌ Error recording referral: {str(e)}"

def get_referrals_by(referrer_name):
    """List all contacts referred by a given person."""
    try:
        records = _get_crm_records()
        results = [r for r in records if r.get("Referred By", "").lower() == referrer_name.lower()]
        if not results:
            return f"No referrals found from {referrer_name}."
        lines = [f"*Referrals from {referrer_name}:*\n"]
        for r in results:
            ref_date = format_date(r.get("Referral Date", "")) if r.get("Referral Date") else "unknown date"
            lines.append(f"👤 {r.get('Name')} — {r.get('Relationship', 'Prospect')} (referred {ref_date})")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error: {str(e)}"

def get_all_referrals():
    """List all contacts with a referral, grouped by referrer."""
    try:
        records = _get_crm_records()
        referrals = [r for r in records if r.get("Referred By", "")]
        if not referrals:
            return "No referrals recorded yet."
        grouped = {}
        for r in referrals:
            ref_by = r.get("Referred By", "Unknown")
            grouped.setdefault(ref_by, []).append(r)
        lines = ["*All Referrals:*\n"]
        for referrer, contacts in sorted(grouped.items(), key=lambda x: -len(x[1])):
            lines.append(f"*{referrer}* ({len(contacts)} referral{'s' if len(contacts) != 1 else ''})")
            for c in contacts:
                lines.append(f"  → {c.get('Name')} ({c.get('Relationship', 'Prospect')})")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error: {str(e)}"

def get_top_referrers():
    """Rank contacts by number of referrals made."""
    try:
        records = _get_crm_records()
        counts = {}
        for r in records:
            ref_by = r.get("Referred By", "")
            if ref_by:
                counts[ref_by] = counts.get(ref_by, 0) + 1
        if not counts:
            return "No referrals recorded yet."
        ranked = sorted(counts.items(), key=lambda x: -x[1])
        lines = ["*Top Referrers:*\n"]
        for i, (name, count) in enumerate(ranked, 1):
            lines.append(f"{i}. {name} — {count} referral{'s' if count != 1 else ''}")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error: {str(e)}"

# --- Excel Import ---
# excel_import_sessions tracks state: { user_id: { "step": "awaiting_columns"|"awaiting_file", "column_order": [...] } }
excel_import_sessions = {}

def parse_excel_column_order(text):
    """Parse user's declared column order from a message like 'Name, Email, Date of Birth, Alias'."""
    # Strip numbering like "1. Name 2. Email" or "Name, Email, DOB"
    text = re.sub(r'\d+[\.\)]\s*', '', text)
    parts = [p.strip() for p in re.split(r'[,\n]', text) if p.strip()]
    return parts

async def handle_excel_import(file_bytes, column_order, update):
    """Import contacts from Excel bytes using declared column order."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws_xl = wb.active
        rows = list(ws_xl.iter_rows(values_only=True))

        # Skip header row if first row looks like headers
        data_rows = rows
        if rows:
            first = [str(c).strip() if c else "" for c in rows[0]]
            if any(h.lower() in ["name", "email", "alias", "date of birth", "dob", "birthday",
                                  "relationship", "address", "context", "referred by"] for h in first):
                data_rows = rows[1:]

        sheet = crm_sheet()
        existing_records = sheet.get_all_records()
        existing_names_set = set()
        for r in existing_records:
            n = r.get("Name", "").strip().lower()
            if n:
                existing_names_set.add(n)

        # Build column index map from declared order
        col_map = {}
        for i, col_name in enumerate(column_order):
            col_map[col_name.strip().lower()] = i

        def get_col(row, *keys):
            for k in keys:
                idx = col_map.get(k.lower())
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        return str(val).strip().replace('\xa0', '').strip()
            return ""

        def normalise_birthday(val):
            """Convert various birthday formats to DD/MM/YYYY."""
            if val is None:
                return ""
            # Already a datetime object (openpyxl reads Excel dates natively)
            if isinstance(val, (datetime,)):
                return val.strftime("%d/%m/%Y")
            s = str(val).strip()
            if not s or s.lower() == "none":
                return ""
            for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y",
                        "%d %b %Y", "%d %B %Y", "%Y-%m-%d %H:%M:%S"]:
                try:
                    return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
                except (ValueError, TypeError):
                    pass
            # Try Excel serial
            try:
                serial = int(float(s))
                epoch = datetime(1899, 12, 30)
                return (epoch + timedelta(days=serial)).strftime("%d/%m/%Y")
            except (ValueError, TypeError):
                pass
            return s  # return raw if nothing parsed

        def normalise_name(raw):
            """Strip whitespace/NBSP, Title Case if ALL CAPS."""
            if not raw:
                return ""
            cleaned = str(raw).replace('\xa0', '').strip()
            if cleaned.isupper():
                cleaned = cleaned.title()
            return cleaned

        def normalise_email(raw):
            return str(raw).strip().lower() if raw else ""

        imported = 0
        skipped = 0
        today = date.today().strftime("%d/%m/%Y")

        rows_to_append = []
        for row in data_rows:
            if not any(c for c in row if c is not None):
                continue

            # Get birthday raw value directly from index (handles datetime objects)
            bday_idx = col_map.get("birthday")
            bday_raw = row[bday_idx] if bday_idx is not None and bday_idx < len(row) else None
            birthday = normalise_birthday(bday_raw)

            name = normalise_name(get_col(row, "Name", "name"))
            if not name:
                continue

            if name.lower() in existing_names_set:
                skipped += 1
                continue

            alias = get_col(row, "Alias", "alias")
            relationship = get_col(row, "Relationship", "relationship")
            context = get_col(row, "Context", "context")
            notes = get_col(row, "Notes", "notes")
            followup_date = get_col(row, "Follow Up Date", "follow up date")
            followup_notes = get_col(row, "Follow Up Notes", "follow up notes")
            referred_by = get_col(row, "Referred By", "referred by")
            referral_date = get_col(row, "Referral Date", "referral date")
            email_raw = get_col(row, "Email", "email")
            email = normalise_email(email_raw)
            address = get_col(row, "Address", "address")

            rows_to_append.append([
                name, alias, birthday, relationship, context, notes,
                followup_date, followup_notes, today, "", referred_by, referral_date, email, address
            ])
            existing_names_set.add(name.lower())
            imported += 1

        # Batch append for speed
        if rows_to_append:
            sheet.append_rows(rows_to_append)

        msg = f"✅ Import done — {imported} contact(s) added"
        if skipped:
            msg += f", {skipped} skipped (already exist)"
        await update.message.reply_text(msg)

    except ImportError:
        await update.message.reply_text("openpyxl isn't installed. Run `pip install openpyxl` and redeploy.")
    except Exception as e:
        await update.message.reply_text(f"❌ Import failed: {str(e)}")

# --- Todo Functions ---
def add_todo(task):
    try:
        sheet = todo_sheet()
        sheet.append_row([task, "Pending", date.today().strftime("%d/%m/%Y")])
        return f"✅ Added to your to-do list: _{task}_"
    except Exception as e:
        return f"❌ Error adding task: {str(e)}"

def complete_todo(task):
    try:
        sheet = todo_sheet()
        records = sheet.get_all_records()
        matches = [
            (i + 2, r) for i, r in enumerate(records)
            if task.lower() in r.get("Task", "").lower() and r.get("Status") == "Pending"
        ]
        if not matches:
            return f"❌ No pending task found matching '{task}'"
        if len(matches) > 1:
            lines = [f"Found {len(matches)} matching tasks — which one?"]
            for i, (_, r) in enumerate(matches, 1):
                lines.append(f"{i}. {r.get('Task')}")
            lines.append("\nReply with the number.")
            return "_DISAMBIG_TODO_COMPLETE_:" + "|".join(r.get("Task", "") for _, r in matches)
        row_idx, r = matches[0]
        sheet.update_cell(row_idx, 2, "Done")
        return f"✅ Marked as done: _{r.get('Task')}_"
    except Exception as e:
        return f"❌ Error completing task: {str(e)}"

def delete_todo(task):
    """Delete a todo by task name, with disambiguation if multiple match."""
    try:
        sheet = todo_sheet()
        records = sheet.get_all_records()
        matches = [
            (i + 2, r) for i, r in enumerate(records)
            if task.lower() in r.get("Task", "").lower()
        ]
        if not matches:
            return f"❌ No task found matching '{task}'"
        if len(matches) > 1:
            return "_DISAMBIG_TODO_DELETE_:" + "|".join(r.get("Task", "") for _, r in matches)
        row_idx, r = matches[0]
        sheet.delete_rows(row_idx)
        return f"Deleted — {r.get('Task')} ✅"
    except Exception as e:
        return f"❌ Error deleting task: {str(e)}"



def list_todos():
    try:
        sheet = todo_sheet()
        records = sheet.get_all_records()
        pending = [r for r in records if r.get("Status") == "Pending"]
        if not pending:
            return "✅ No pending tasks!"
        response = f"📝 *{len(pending)} pending task(s):*\n\n"
        for r in pending:
            response += f"• {r.get('Task')} _(added {format_date(r.get('Added', ''))})_\n"
        return response
    except Exception as e:
        return f"❌ Error listing tasks: {str(e)}"

# --- Calendar Functions ---
def get_events(days=1):
    try:
        calendar = get_calendar("Personal")
        if not calendar:
            return "⚠️ Couldn't connect to iCloud Calendar — check that ICLOUD_USERNAME and ICLOUD_PASSWORD are set correctly in Railway."
        start = datetime.now()
        end = start + timedelta(days=days)
        events = calendar.date_search(start=start, end=end)
        if not events:
            label = "today" if days == 1 else "this week"
            return f"📅 No events {label}"
        label = "Today's events" if days == 1 else "This week's events"
        response = f"📅 *{label}:*\n\n"
        for event in events:
            e = event.vobject_instance.vevent
            summary = str(e.summary.value)
            dtstart = e.dtstart.value
            if hasattr(dtstart, 'strftime'):
                response += f"• *{summary}* — {dtstart.strftime('%d %b, %H:%M')}\n"
            else:
                response += f"• *{summary}* — {dtstart}\n"
        return response
    except Exception as e:
        return f"⚠️ Calendar error: {str(e)} — if this keeps happening, check your iCloud credentials in Railway."

def delete_calendar_event(title):
    try:
        calendar = get_calendar()
        if not calendar:
            return "❌ Could not connect to iCloud Calendar"
        start = datetime.now()
        end = start + timedelta(days=30)
        events = calendar.date_search(start=start, end=end)
        for event in events:
            e = event.vobject_instance.vevent
            if title.lower() in str(e.summary.value).lower():
                event.delete()
                return f"✅ Event *{str(e.summary.value)}* deleted"
        return f"❌ No event found matching '{title}' in the next 30 days"
    except Exception as e:
        return f"❌ Error deleting event: {str(e)}"

def is_calendar_request(text):
    lower = text.lower().strip()

    # Tier 1: Hard keyword triggers — zero-cost instant match
    CALENDAR_TRIGGERS = [
        "schedule ", "add event", "create event", "new event",
        "book ", "set up a meeting", "set up meeting",
        "pencil in", "block out", "block off",
        "put in my calendar", "add to calendar", "add to my calendar",
        "add a meeting", "add meeting", "meeting at ", "meeting on ",
        "dinner at ", "dinner on ", "lunch at ", "lunch on ",
        "drinks at ", "drinks on ", "breakfast at ",
        "call at ", "call on ", "appointment at ", "appointment on ",
        "catch up at ", "catch up on ", "event at ", "event on ",
        "remind me on calendar", "add reminder to calendar",
    ]
    if any(t in lower for t in CALENDAR_TRIGGERS):
        return True

    # Tier 2: Hard exclusions — zero-cost instant reject
    EXCLUSIONS = [
        "remind me", "reminder", "spent", "paid", "bought", "cost",
        "expense", "bill", "price", "how much", "what time", "what's the time",
        "stock", "portfolio", "restaurant", "save ", "note ", "find ",
        "search ", "todo", "weather",
    ]
    if any(e in lower for e in EXCLUSIONS):
        return False

    # Tier 3: Require a time anchor — calendar events always have one.
    # Without a time anchor there is nothing to schedule, so skip the API call entirely.
    TIME_WORDS = [
        "tomorrow", "tonight", "monday", "tuesday", "wednesday",
        "thursday", "friday", "saturday", "sunday", "next week",
        "this evening", "this afternoon", "this morning", "next month",
        "on the ", "this friday", "this saturday", "this sunday",
    ]
    AT_PATTERN = bool(re.search(r'\b(at|@)\s*\d{1,2}(:\d{2})?\s*(am|pm)?\b', lower))
    DATE_PATTERN = bool(re.search(r'\b\d{1,2}[\/\-]\d{1,2}\b', lower))
    has_time_anchor = AT_PATTERN or DATE_PATTERN or any(w in lower for w in TIME_WORDS)

    if not has_time_anchor:
        return False  # No time anchor → not a calendar request, no API call

    # Tier 4: Has time anchor + event word → call Claude to confirm
    # This is the only path that hits the API, and only when genuinely ambiguous
    AMBIGUOUS_SIGNALS = [
        "meet", "meeting", "catch up", "dinner", "lunch", "drinks",
        "coffee", "call", "appointment", "event", "session",
        "interview", "presentation", "visit",
    ]
    if any(s in lower for s in AMBIGUOUS_SIGNALS) or AT_PATTERN:
        try:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=10,
                messages=[{"role": "user", "content":
                    f'Is this asking to add a calendar event? Reply YES or NO only.\n\n"{text}"'}]
            )
            return response.content[0].text.strip().upper() == "YES"
        except Exception as e:
            print(f"is_calendar_request API fallback error: {e}")
            return False

    return False

# --- Smart Calendar ---
def smart_add_event(text, user_id):
    try:
        today = date.today()
        tomorrow = today + timedelta(days=1)
        calendar_names = [
            "Estate Planning", "Work Meeting", "Personal",
            "Closing", "Urgent", "MinimaList", "Appointment"
        ]
        parse_prompt = f"""Today is {today.strftime('%d %b %Y')} ({today.strftime('%A')}).
Tomorrow is {tomorrow.strftime('%d %b %Y')}.

Available calendars: {', '.join(calendar_names)}

The user wants to add a calendar event. Extract the details from this message:
"{text}"

Respond ONLY with a JSON object in this exact format, nothing else:
{{
  "title": "event title",
  "start": "DD/MM/YYYY HH:MM",
  "end": "DD/MM/YYYY HH:MM",
  "notes": "any notes or empty string",
  "calendar": "exact calendar name from the list above or Personal if unclear"
}}

Rules:
- If no end time given, assume 1 hour after start
- If AM/PM not specified, use context to determine (6pm not 6am for dinner etc)
- Match calendar name exactly from the available list
- If no calendar specified, use Personal"""

        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{"role": "user", "content": parse_prompt}]
        )

        raw = response.content[0].text.strip()
        clean = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(clean)

        title = parsed.get("title", "")
        start_str = parsed.get("start", "")
        end_str = parsed.get("end", "")
        notes = parsed.get("notes", "")
        cal_name = parsed.get("calendar", "Personal")

        if not title or not start_str or not end_str:
            return "❌ Couldn't parse the event details. Try something like: schedule dinner with James tomorrow 7pm"

        start = datetime.strptime(start_str, "%d/%m/%Y %H:%M")
        end = datetime.strptime(end_str, "%d/%m/%Y %H:%M")

        calendar = get_calendar(cal_name)
        if not calendar:
            calendar = get_calendar("Personal")

        ics = (
            "BEGIN:VCALENDAR\nVERSION:2.0\nBEGIN:VEVENT\n"
            f"SUMMARY:{title}\n"
            f"DTSTART:{start.strftime('%Y%m%dT%H%M%S')}\n"
            f"DTEND:{end.strftime('%Y%m%dT%H%M%S')}\n"
            f"DESCRIPTION:{notes}\n"
            "END:VEVENT\nEND:VCALENDAR"
        )
        calendar.add_event(ics)

        return (
            f"✅ Done!\n\n"
            f"📅 *{title}*\n"
            f"🕐 {start.strftime('%d %b %Y, %H:%M')} → {end.strftime('%H:%M')}\n"
            f"📁 {calendar.name}\n"
            + (f"📝 {notes}" if notes else "")
        )

    except json.JSONDecodeError:
        return "⚠️ Couldn't parse that — what are the event details?\n(e.g. Team Offsite, 30 Apr 2026, 2pm)"
    except Exception as e:
        if "iCloud" in str(e) or "caldav" in str(e).lower() or "calendar" in str(e).lower():
            return "⚠️ Couldn't connect to iCloud Calendar — check credentials in Railway."
        return "⚠️ Couldn't parse that — what are the event details?\n(e.g. Team Offsite, 30 Apr 2026, 2pm)"

# --- Edit Session Handler ---
async def handle_edit_session(user_id, text, update):
    session = edit_sessions[user_id]
    step = session["step"]
    fields = ["alias", "birthday", "relationship", "context", "notes",
              "follow up date", "follow up notes", "email", "address"]

    if step == "choose_field":
        field = text.lower().strip()
        if field == "cancel":
            del edit_sessions[user_id]
            await update.message.reply_text("Cancelled.")
            return
        if field not in fields:
            await update.message.reply_text(
                f"Pick a field to edit:\n1. Alias\n2. Birthday\n3. Relationship\n4. Context\n"
                f"5. Notes\n6. Follow up date\n7. Follow up notes\n8. Email\n9. Address\n\n"
                f"Or type *cancel* to exit.",
                parse_mode="Markdown"
            )
            return
        session["field"] = field
        session["step"] = "enter_value"
        await update.message.reply_text(f"Enter the new value for *{field.title()}*:", parse_mode="Markdown")

    elif step == "enter_value":
        field = session["field"]
        name = session["name"]
        result = update_field(f"{name}, {field}, {text}")
        del edit_sessions[user_id]
        await update.message.reply_text(result, parse_mode="Markdown")

# --- Scheduled Reminders ---
DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"]

def parse_date_flexible(date_str):
    """Try multiple date formats. Returns date object or None."""
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            continue
    return None

async def send_followup_reminders(app):
    try:
        records = _get_crm_records()
        today = date.today()
        for i, r in enumerate(records, start=2):
            fu_date_str = r.get("Follow Up Date", "")
            if not fu_date_str:
                continue
            fu_date = parse_date_flexible(fu_date_str)
            if fu_date is None:
                # Unrecognised format — ask for correct date
                await app.bot.send_message(
                    chat_id=YOUR_CHAT_ID,
                    text=(
                        f"⚠️ Couldn't read the follow-up date for *{r.get('Name', '?')}*.\n"
                        f"What's the correct date? (e.g. 25 Apr 2026)\n"
                        f"Reply: followup date {r.get('Name', '')} [date]"
                    ),
                    parse_mode="Markdown"
                )
                continue
            if fu_date == today:
                message = (
                    f"🔔 *Follow up reminder!*\n\n"
                    f"👤 *{r.get('Name')}*\n"
                    f"📝 {r.get('Follow Up Notes') or 'No notes'}\n\n"
                    f"Don't forget to reach out today!"
                )
                await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=message, parse_mode="Markdown")
    except Exception as e:
        print(f"Error sending follow up reminders: {e}")

# --- Birthday Greeting State ---
# Tracks who got a 12pm prompt today, pending acknowledgement
birthday_pending = {}

def persist_birthday_pending():
    """Persist birthday_pending to Settings sheet. Fire-and-forget — never raises."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        today_str = date.today().strftime("%d/%m/%Y")
        data = json.dumps({"date": today_str, "pending": birthday_pending})
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if r.get("Key") == "birthday_pending":
                sheet.update_cell(i + 2, 2, data)
                return
        sheet.append_row(["birthday_pending", data])
    except Exception as e:
        print(f"persist_birthday_pending error: {e}")

def load_birthday_pending_from_sheet():
    """Restore birthday_pending from Settings sheet on startup. Only restore if date matches today."""
    global birthday_pending
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "birthday_pending":
                raw = r.get("Value", "")
                if raw:
                    loaded = json.loads(raw)
                    if loaded.get("date") == date.today().strftime("%d/%m/%Y"):
                        birthday_pending = loaded.get("pending", {})
                        if birthday_pending:
                            print(f"Restored {len(birthday_pending)} birthday pending(s) from sheet")
                return
    except Exception as e:
        print(f"load_birthday_pending_from_sheet error: {e}")

def ensure_birthday_greeted_column():
    """Birthday Greeted is col 10 in new CRM schema — nothing to add."""
    pass  # Column already defined in new header structure

def get_birthday_greeted_col():
    """Return the column index (1-based) of Birthday Greeted — always 10 in new schema."""
    return 10

def generate_birthday_greeting(name, age, relationship, context, notes):
    """Generate a personalised birthday greeting via Claude. Runs sync — call via asyncio.to_thread."""
    context_str = f"Relationship: {relationship}. Context: {context}. Notes: {notes or 'none'}."
    greeting_prompt = (
        f"Write a warm, casual birthday greeting that someone could copy and paste to send to {name}, "
        f"who is turning {age} today.\n\n"
        f"Context: {context_str}\n\n"
        f"Rules:\n"
        f"- Must open with Happy birthday\n"
        f"- Casual and warm throughout, never stiff or corporate\n"
        f"- Add one personal touch based on their notes if relevant\n"
        f"- End with a warm closing line\n"
        f"- No dashes anywhere\n"
        f"- Do NOT use: Hope you had a great one, Hope its been a good one\n"
        f"- 2 to 4 sentences max\n"
        f"- Write it in first person, ready to copy and send"
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": greeting_prompt}]
        )
        return resp.content[0].text.strip()
    except Exception as e:
        print(f"generate_birthday_greeting Claude error for {name}: {e}")
        return f"Happy birthday {name}! Hope you have a wonderful day ahead."

async def send_birthday_reminders(app):
    """12pm job: find today birthdays, generate greeting, send prompt."""
    global birthday_pending
    try:
        ensure_birthday_greeted_column()
        records = _get_crm_records()
        today = date.today()

        for i, r in enumerate(records):
            bday_str = r.get("Birthday", "")
            if not bday_str:
                continue
            try:
                bday = datetime.strptime(bday_str, "%d/%m/%Y").date()
                if bday.day != today.day or bday.month != today.month:
                    continue

                name = r.get("Name", "")
                age = calculate_age(bday_str)
                notes = r.get("Notes", "")
                relationship = r.get("Relationship", "")
                context = r.get("Context", "")

                # Skip if already greeted today
                already_greeted = r.get("Birthday Greeted", "")
                if already_greeted == today.strftime("%d/%m/%Y"):
                    continue

                greeting = await asyncio.to_thread(
                    generate_birthday_greeting, name, age, relationship, context, notes
                )

                msg = (
                    f"It's {name}'s birthday today! Turning {age}. 🎂\n\n"
                    f"Here's a message you can send:\n\n"
                    f"{greeting}\n\n"
                    f"Reply 'sent' when you've wished them, or 'skip' to dismiss."
                )
                await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)

                birthday_pending[name] = {
                    "row": i + 2,
                    "greeted": False,
                    "greeting": greeting
                }
                persist_birthday_pending()
                print(f"Birthday prompt sent for {name}")

            except Exception as e:
                print(f"Birthday reminder error for {r.get('Name', '?')}: {e}")

    except Exception as e:
        print(f"Error in send_birthday_reminders: {e}")

async def send_birthday_followups(app):
    """2pm job: follow up on unacknowledged birthdays, then drop."""
    global birthday_pending
    try:
        still_pending = {k: v for k, v in birthday_pending.items() if not v.get("greeted")}

        for name, data in still_pending.items():
            try:
                msg = (
                    f"Did you get a chance to wish {name} happy birthday? 🎂\n\n"
                    f"Here's that message again:\n\n"
                    f"{data['greeting']}"
                )
                await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
                print(f"2pm follow-up sent for {name}")
            except Exception as e:
                print(f"Error sending 2pm follow-up for {name}: {e}")

        birthday_pending = {}
        persist_birthday_pending()

    except Exception as e:
        print(f"Error in send_birthday_followups: {e}")

def mark_birthday_greeted(name):
    """Mark contact as greeted today in CRM sheet."""
    try:
        col = get_birthday_greeted_col()
        if not col:
            return
        row_num, record = find_row(name)
        if record and row_num != "disambig":
            sheet = crm_sheet()
            sheet.update_cell(row_num, col, date.today().strftime("%d/%m/%Y"))
    except Exception as e:
        print(f"Error marking birthday greeted for {name}: {e}")

def check_birthday_acknowledgement(text):
    """
    Check if user explicitly said 'sent' or 'skip' for a pending birthday.
    Must be called with the user's message text.
    Returns (True, reply_message) if handled, (False, None) otherwise.
    """
    global birthday_pending
    if not birthday_pending:
        return False, None

    lower = text.strip().lower()
    if lower not in ["sent", "done", "skip", "skipped", "sent it", "sent!",
                     "yeah sent it", "already sent", "ok done", "yep sent",
                     "yup sent", "sent already", "done already", "ok skip",
                     "yeah skip", "just sent", "just sent it"]:
        return False, None

    acknowledged = []
    for name, data in list(birthday_pending.items()):
        if not data.get("greeted"):
            if lower in ["skip", "skipped"]:
                data["greeted"] = True
                acknowledged.append(f"Skipped {name}")
            else:
                data["greeted"] = True
                mark_birthday_greeted(name)
                acknowledged.append(f"Marked {name} as greeted ✅")

    if acknowledged:
        persist_birthday_pending()
        return True, "\n".join(acknowledged)
    return False, None





# --- Meeting Recap ---
# Session state: { user_id: { "event_name": str, "notes": [str], "step": "collecting"|"confirming", "pending_recap": dict } }
meeting_sessions = {}

MEETING_START_PHRASES = [
    "meeting recap", "taking notes", "log this meeting", "meeting notes",
    "recap for", "notes for", "log meeting", "start recap", "new recap",
    "networking recap", "presentation recap"
]

MEETING_DONE_PHRASES = [
    "done", "that's it", "thats it", "save that", "save it",
    "finish", "finished", "end recap", "process this", "that's all", "thats all"
]

def is_meeting_start(text):
    lower = text.lower()
    return any(p in lower for p in MEETING_START_PHRASES)

def is_meeting_done(text):
    lower = text.lower().strip()
    return any(lower == p or lower.startswith(p) for p in MEETING_DONE_PHRASES)

def extract_event_name(text):
    """Try to pull event name from the start message."""
    lower = text.lower()
    for phrase in ["recap for", "notes for", "meeting notes for", "taking notes for",
                   "meeting recap for", "log meeting for"]:
        if phrase in lower:
            idx = lower.index(phrase) + len(phrase)
            return text[idx:].strip().strip(".,!?")
    return ""

def tag_crm_contacts(text):
    """Find any CRM contacts mentioned in the text."""
    try:
        records = _get_crm_records()
        tagged = []
        for r in records:
            name = r.get("Name", "")
            if name and name.lower() in text.lower():
                tagged.append(name)
        return tagged
    except Exception as e:
        print(f"tag_crm_contacts error: {e}")
        return []

def format_recap_confirmation(recap):
    """Format recap for user confirmation — no emoji on header line."""
    lines = []
    lines.append(f"Meeting Recap — {recap.get('event_name', 'Untitled')}")
    lines.append("")
    lines.append(f"Topic: {recap.get('topic', '')}")
    lines.append("")
    lines.append(f"Summary:\n{recap.get('summary', '')}")

    action_items = recap.get("action_items", [])
    if action_items:
        lines.append("")
        lines.append("Action Items:")
        for item in action_items:
            lines.append(f"• {item}")

    foreign = recap.get("foreign_phrases", {})
    if foreign:
        lines.append("")
        lines.append("Phrases:")
        for orig, trans in foreign.items():
            lines.append(f"• {orig} [{trans}]")

    contacts = recap.get("contacts_mentioned", [])
    if contacts:
        lines.append("")
        lines.append(f"Contacts tagged: {', '.join(contacts)}")

    lines.append("")
    lines.append("Reply Y to save or E to edit.")
    return "\n".join(lines)

def process_meeting_notes(event_name, notes_list):
    """Send all buffered notes to Claude and get a structured recap back as JSON.
    Falls back to saving raw notes if Claude call fails."""
    combined = "\n".join(notes_list)
    prompt = (
        f"You are processing meeting notes for an event called: {event_name or 'unknown event'}\n\n"
        f"Here are the raw notes:\n{combined}\n\n"
        f"Extract and return a JSON object with these fields:\n"
        f"- event_name: string (use the provided event name, or infer if not given)\n"
        f"- topic: string (1 line summary of what the meeting/event was about)\n"
        f"- summary: string (2-4 sentences capturing key points, insights, context)\n"
        f"- action_items: list of strings (only include if there are clear action items, otherwise empty list)\n"
        f"- contacts_mentioned: list of strings (names of people mentioned)\n"
        f"- foreign_phrases: dict (any non-English phrases found, key=original, value=English translation)\n\n"
        f"Return ONLY the JSON object, no markdown, no preamble."
    )
    try:
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
        return json.loads(raw), False  # (recap, is_raw)
    except Exception as e:
        print(f"process_meeting_notes Claude error: {e}")
        # Return raw notes as fallback
        return {
            "event_name": event_name or "Untitled",
            "topic": "Raw notes — processing failed",
            "summary": combined,
            "action_items": [],
            "contacts_mentioned": [],
            "foreign_phrases": {}
        }, True  # (recap, is_raw)

def save_meeting_recap(recap):
    """Save the confirmed recap to the Meeting Notes sheet. Returns (success, error_msg)."""
    try:
        sheet = spreadsheet.worksheet("Meeting Notes")
        today = date.today().strftime("%d/%m/%Y")
        action_items_str = "; ".join(recap.get("action_items", []))
        sheet.append_row([
            recap.get("event_name", ""),
            recap.get("topic", ""),
            recap.get("summary", ""),
            action_items_str,
            today
        ])
        return True, None
    except Exception as e:
        print(f"Error saving meeting recap: {e}")
        return False, str(e)

def search_meeting_notes(query):
    """Search meeting notes by keyword or date range."""
    try:
        sheet = spreadsheet.worksheet("Meeting Notes")
        try:
            records = sheet.get_all_records()
        except Exception as e:
            return f"⚠️ Couldn't search meeting notes right now — try again in a moment."

        if not records:
            return "No meeting notes saved yet."

        query_lower = query.lower().strip()
        results = []
        for r in records:
            searchable = " ".join(str(v).lower() for v in r.values())
            if query_lower in searchable:
                results.append(r)

        if not results:
            return f"No meeting notes found matching '{query}'."

        lines = [f"Found {len(results)} recap(s):\n"]
        for r in results:
            lines.append(f"📋 {r.get('Event Name', '')} — {r.get('Date', '')}")
            lines.append(f"   {r.get('Topic', '')}")
            lines.append("")
        return "\n".join(lines)
    except Exception as e:
        return "⚠️ Couldn't search meeting notes right now — try again in a moment."

async def handle_meeting_session(user_id, text, update):
    """Handle messages when user is in an active meeting recap session."""
    session = meeting_sessions[user_id]
    step = session.get("step")

    # Confirming step — waiting for Y or E
    if step == "confirming":
        if text.strip().upper() == "Y":
            saved, err = save_meeting_recap(session["pending_recap"])
            contacts = session["pending_recap"].get("contacts_mentioned", [])
            for name in contacts:
                row_num, record = find_row(name)
                if record and row_num != "disambig":
                    pass  # silently tagged via presence in recap
            del meeting_sessions[user_id]
            session_timestamps.pop(user_id, None)
            if saved:
                await update.message.reply_text("Saved! 📋")
            else:
                await update.message.reply_text(
                    f"⚠️ Couldn't save recap to sheet — keeping it in memory.\n"
                    f"Reply 'retry save' to try again or 'send recap' to have me send it as a message."
                )
                # Keep session alive with saved recap for retry
                meeting_sessions[user_id] = {**session, "step": "save_failed"}
                touch_session(user_id)

        elif text.strip().upper() == "E":
            session["step"] = "collecting"
            session["notes"] = []
            await update.message.reply_text(
                "No worries, let's redo it. Send your notes again and say done when you're finished."
            )
        else:
            await update.message.reply_text("Reply Y to save or E to start over.")
        return

    if step == "save_failed":
        lower_t = text.strip().lower()
        if lower_t == "retry save":
            saved, err = save_meeting_recap(session["pending_recap"])
            if saved:
                del meeting_sessions[user_id]
                session_timestamps.pop(user_id, None)
                await update.message.reply_text("Saved! 📋")
            else:
                await update.message.reply_text("Still couldn't save — try again in a moment.")
        elif lower_t == "send recap":
            recap = session["pending_recap"]
            msg = format_recap_confirmation(recap).replace("\nReply Y to save or E to edit.", "")
            del meeting_sessions[user_id]
            session_timestamps.pop(user_id, None)
            await update.message.reply_text(msg)
        return

    # Get event name step
    if step == "get_name":
        session["event_name"] = text.strip()
        session["step"] = "collecting"
        await update.message.reply_text(
            f"Got it. Send your notes for {text.strip()} and say done when you're finished."
        )
        return

    # Collecting step — buffering notes
    if is_meeting_done(text):
        if not session.get("notes"):
            await update.message.reply_text("You haven't sent any notes yet. Send your notes then say done.")
            return

        await update.message.reply_text("On it, give me a sec...")

        try:
            recap, is_raw = process_meeting_notes(session.get("event_name", ""), session["notes"])
            session["pending_recap"] = recap
            session["step"] = "confirming"
            touch_session(user_id)
            if is_raw:
                await update.message.reply_text(
                    "⚠️ Took too long to process — saved your raw notes.\n"
                    "Reply 'retry recap' to process them when ready, or Y to save as-is."
                )
            confirmation = format_recap_confirmation(recap)
            await update.message.reply_text(confirmation)
        except Exception as e:
            await update.message.reply_text(f"Couldn't process the notes: {str(e)}. Try again.")
        return

    # Still collecting — buffer the note
    session["notes"].append(text)
    # Acknowledge silently (no reply) to keep flow natural



# --- Custom Reminders ---
# Reminders sheet columns: ID, Message, Scheduled Time, Recurrence, Status, Attempts, Contact
# Status: pending | sent | cancelled
# Recurrence: once | daily | weekly | monthly | or natural string like "every Monday"

TIMEZONE = pytz.timezone("Asia/Kuala_Lumpur")

# Tracks the last reminder fired per user so they can say "remind me again in 2 hours"
last_fired_reminder = {}
_pending_reminders_cache = None  # None = stale, list = valid cache

def reminders_sheet():
    try:
        return spreadsheet.worksheet("Reminders")
    except Exception:
        print("Reminders sheet not found — creating it")
        ws = spreadsheet.add_worksheet(title="Reminders", rows=500, cols=8)
        ws.append_row(["ID", "Message", "Scheduled Time", "Recurrence", "Status", "Attempts", "Contact"])
        return ws

def generate_reminder_id():
    """Simple unique ID based on timestamp."""
    return datetime.now().strftime("%Y%m%d%H%M%S")

def parse_reminder_request(text):
    """Use Claude to parse a natural language reminder request. Falls back to manual input prompt."""
    now = datetime.now(TIMEZONE)
    prompt = (
        f"Today is {now.strftime('%A, %d %b %Y')} and the time is {now.strftime('%H:%M')} (Asia/Kuala_Lumpur).\n\n"
        f"Parse this reminder request and return a JSON object:\n"
        f"Request: {text}\n\n"
        f"Return ONLY a JSON object with these fields:\n"
        f"- message: string (what to remind about, concise)\n"
        f"- scheduled_time: string in format YYYY-MM-DD HH:MM (exact datetime to fire)\n"
        f"- recurrence: string — one of: once, daily, weekly, monthly, or a description like 'every Monday' (use 'once' if not recurring)\n"
        f"- contact: string (name of person mentioned, or empty string)\n\n"
        f"Rules:\n"
        f"- If no time specified, default to 09:00\n"
        f"- 'tomorrow' means {(now + timedelta(days=1)).strftime('%Y-%m-%d')}\n"
        f"- 'next week' means {(now + timedelta(days=7)).strftime('%Y-%m-%d')}\n"
        f"- For recurring reminders, scheduled_time is the FIRST occurrence\n"
        f"- Return ONLY the JSON, no markdown, no explanation"
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
        return json.loads(raw)
    except Exception as e:
        print(f"parse_reminder_request error: {e}")
        return None  # Caller handles None by prompting manual input

def parse_reschedule_request(text, original_message):
    """Parse a follow-up reschedule like 'remind me again in 2 hours'."""
    now = datetime.now(TIMEZONE)
    prompt = (
        f"Today is {now.strftime('%A, %d %b %Y')} and the time is {now.strftime('%H:%M')}.\n\n"
        f"The user wants to reschedule a reminder. Original reminder: '{original_message}'\n"
        f"Reschedule request: '{text}'\n\n"
        f"Return ONLY a JSON object:\n"
        f"- scheduled_time: string in format YYYY-MM-DD HH:MM\n"
        f"- recurrence: string (once, daily, weekly, monthly — use 'once' if unclear)\n\n"
        f"Return ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=100,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, ValueError) as e:
        print(f"parse_reschedule_request JSON error: {e} | raw: {raw[:100]}")
        return None

def add_reminder(message, scheduled_time_str, recurrence="once", contact=""):
    """Add a reminder to the Reminders sheet."""
    global _pending_reminders_cache
    sheet = reminders_sheet()
    reminder_id = generate_reminder_id()
    sheet.append_row([
        reminder_id,
        message,
        scheduled_time_str,
        recurrence,
        "pending",
        "0",
        contact
    ])
    _pending_reminders_cache = None  # invalidate cache
    return reminder_id

def cancel_reminder_by_keyword(keyword):
    """Cancel reminders matching a keyword. Returns list of cancelled, or '_DISAMBIG_:' prefix if multiple."""
    global _pending_reminders_cache
    sheet = reminders_sheet()
    records = sheet.get_all_records()
    matches = [
        (i + 2, r) for i, r in enumerate(records)
        if keyword.lower() in r.get("Message", "").lower() and r.get("Status") == "pending"
    ]
    if not matches:
        return []
    if len(matches) > 1:
        # Return signal for disambiguation
        return ["_DISAMBIG_:" + "|".join(f"{row}:{r.get('Message','')} — {r.get('Scheduled Time','')}"
                                          for row, r in matches)]
    row_idx, r = matches[0]
    sheet.update_cell(row_idx, 5, "cancelled")
    _pending_reminders_cache = None  # invalidate cache
    return [r.get("Message", "")]

def list_reminders():
    """List all pending reminders."""
    sheet = reminders_sheet()
    records = sheet.get_all_records()
    pending = [r for r in records if r.get("Status") == "pending"]
    if not pending:
        return "No upcoming reminders."
    lines = [f"You have {len(pending)} upcoming reminder(s):\n"]
    for r in pending:
        t = r.get("Scheduled Time", "")
        msg = r.get("Message", "")
        rec = r.get("Recurrence", "once")
        rec_str = f" ({rec})" if rec != "once" else ""
        lines.append(f"• {msg} — {t}{rec_str}")
    return "\n".join(lines)

def get_next_recurrence(scheduled_time_str, recurrence):
    """Calculate the next fire time for a recurring reminder."""
    try:
        dt = datetime.strptime(scheduled_time_str, "%Y-%m-%d %H:%M")
        if recurrence == "daily":
            return (dt + timedelta(days=1)).strftime("%Y-%m-%d %H:%M")
        elif recurrence == "weekly" or "monday" in recurrence.lower() or "every" in recurrence.lower():
            return (dt + timedelta(weeks=1)).strftime("%Y-%m-%d %H:%M")
        elif recurrence == "monthly":
            # Add roughly a month
            if dt.month == 12:
                next_dt = dt.replace(year=dt.year + 1, month=1)
            else:
                next_dt = dt.replace(month=dt.month + 1)
            return next_dt.strftime("%Y-%m-%d %H:%M")
        return None
    except Exception as e:
        print(f"get_next_recurrence error: {e}")
        return None

async def check_and_fire_reminders(app):
    """Check pending reminders every minute. Uses in-memory cache — only re-reads sheet if cache is stale."""
    global _pending_reminders_cache
    try:
        now = datetime.now(TIMEZONE).replace(second=0, microsecond=0)

        # Populate cache if stale
        if _pending_reminders_cache is None:
            sheet = reminders_sheet()
            records = sheet.get_all_records()
            _pending_reminders_cache = [(i, r) for i, r in enumerate(records) if r.get("Status") == "pending"]

        pending_records = _pending_reminders_cache

        for i, r in pending_records:
            scheduled_str = r.get("Scheduled Time", "")
            if not scheduled_str:
                continue

            try:
                scheduled = datetime.strptime(scheduled_str, "%Y-%m-%d %H:%M")
                scheduled = TIMEZONE.localize(scheduled)
            except Exception as e:
                print(f"check_and_fire_reminders: bad scheduled time '{scheduled_str}': {e}")
                continue

            # Fire if within the current minute
            if scheduled <= now:
                try:
                    attempts = int(r.get("Attempts", 0))
                    message = r.get("Message", "")
                    recurrence = r.get("Recurrence", "once")
                    contact = r.get("Contact", "")
                    row = i + 2

                    # Build reminder message (contact notes looked up from cache, not a fresh sheet read)
                    reminder_msg = f"🔔 Reminder: {message}"
                    if contact:
                        crm_records = _get_crm_records()
                        contact_l = contact.lower()
                        matched = next((r for r in crm_records if r.get("Name", "").lower() == contact_l), None)
                        if matched:
                            notes = matched.get("Notes", "")
                            if notes:
                                reminder_msg += f"\n\n({contact}: {notes.split(';')[0].strip()})"

                    await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=reminder_msg)

                    # Store as last fired for potential reschedule
                    last_fired_reminder[YOUR_CHAT_ID] = {
                        "id": r.get("ID"),
                        "message": message,
                        "row": row
                    }

                    if recurrence != "once":
                        # Schedule next occurrence
                        next_time = get_next_recurrence(scheduled_str, recurrence)
                        sheet = reminders_sheet()
                        if next_time:
                            sheet.update_cell(row, 3, next_time)
                            sheet.update_cell(row, 6, "0")  # reset attempts
                        else:
                            sheet.update_cell(row, 5, "sent")
                        _pending_reminders_cache = None  # invalidate cache after write
                    else:
                        # One-off: mark attempts, retry once after 2 hours if first attempt
                        sheet = reminders_sheet()
                        if attempts == 0:
                            retry_time = (now + timedelta(hours=2)).strftime("%Y-%m-%d %H:%M")
                            sheet.update_cell(row, 3, retry_time)
                            sheet.update_cell(row, 6, "1")
                        else:
                            # Second attempt done — mark sent and drop
                            sheet.update_cell(row, 5, "sent")
                        _pending_reminders_cache = None  # invalidate cache after write
                except Exception as e:
                    print(f"check_and_fire_reminders: failed to fire reminder id={r.get('ID', '?')}: {e}")

    except Exception as e:
        print(f"Error in check_and_fire_reminders: {e}")

def is_reminder_request(text):
    """Detect if a message is asking to set a reminder.
    Bare 'alert me', 'notify me', 'ping me' removed — too broad, collide with stock.
    'alert me if' is stock. 'alert me when/to' is reminder."""
    lower = text.lower()
    triggers = [
        "remind me", "remind", "set a reminder", "set me a reminder",
        "reminder for", "reminder at", "reminder to",
        "don't let me forget", "dont let me forget",
        "i need to remember to", "i need to remember",
        "don't let me forget to", "dont let me forget to",
        "alert me when", "alert me to",
        "notify me when", "notify me about", "notify me to",
        "ping me at", "ping me when", "ping me to",
        "drop me a reminder", "drop a reminder", "send me a reminder"
    ]
    return any(t in lower for t in triggers)

def is_reschedule_request(text):
    """Detect if message is rescheduling the last reminder."""
    lower = text.lower()
    triggers = ["remind me again", "snooze", "again in", "remind again",
                "push it to", "reschedule"]
    return any(t in lower for t in triggers)

def is_cancel_reminder_request(text):
    """Detect cancel reminder intent."""
    lower = text.lower()
    return ("cancel" in lower or "delete" in lower or "remove" in lower) and "reminder" in lower

def handle_new_reminder(text):
    """Parse and save a new reminder. Returns confirmation string."""
    try:
        parsed = parse_reminder_request(text)
        if parsed is None:
            return "⚠️ Couldn't parse that — what are the reminder details?\n(e.g. Call John, 30 Apr 2026, 9am)"

        message = parsed.get("message", text)
        scheduled_time = parsed.get("scheduled_time", "")
        recurrence = parsed.get("recurrence", "once")
        contact = parsed.get("contact", "")

        if not scheduled_time:
            return "Couldn't figure out when to remind you. Try: 'remind me to call James tomorrow at 3pm'."

        # Validate contact against CRM
        if contact:
            row, record = find_row(contact)
            if not record or row == "disambig":
                contact = ""

        add_reminder(message, scheduled_time, recurrence, contact)

        try:
            dt = datetime.strptime(scheduled_time, "%Y-%m-%d %H:%M")
            time_str = dt.strftime("%d %b %Y at %I:%M %p")
        except Exception as e:
            print(f"handle_new_reminder: time format error: {e}")
            time_str = scheduled_time

        rec_str = f" ({recurrence})" if recurrence != "once" else ""
        return f"Done, I'll remind you to {message} on {time_str}{rec_str}."

    except Exception as e:
        return f"⚠️ Couldn't parse that — what are the reminder details?\n(e.g. Call John, 30 Apr 2026, 9am)"

def handle_reschedule(text, user_id):
    """Reschedule the last fired reminder."""
    try:
        last = last_fired_reminder.get(user_id)
        if not last:
            return "Not sure which reminder you mean. Can you be more specific?"

        parsed = parse_reschedule_request(text, last["message"])
        if not parsed:
            return "Couldn't parse the reschedule — try 'remind me again in 2 hours'."
        new_time = parsed.get("scheduled_time", "")
        recurrence = parsed.get("recurrence", "once")

        if not new_time:
            return "Couldn't figure out the new time. Try 'remind me again in 2 hours'."

        sheet = reminders_sheet()
        row = last.get("row")
        if row:
            sheet.update_cell(row, 3, new_time)
            sheet.update_cell(row, 5, "pending")
            sheet.update_cell(row, 6, "0")

        try:
            dt = datetime.strptime(new_time, "%Y-%m-%d %H:%M")
            time_str = dt.strftime("%d %b at %I:%M %p")
        except Exception as e:
            print(f"handle_reschedule: time format error: {e}")
            time_str = new_time

        return f"Got it, I'll remind you about {last['message']} again on {time_str}."

    except Exception as e:
        return f"❌ Couldn't reschedule: {str(e)}"



# =============================================================================
# EXPENSE TRACKER
# =============================================================================

EXPENSE_CATEGORIES = ["FnB", "Entertainment", "Personal", "Family", "Work", "Transport", "Shopping", "Travel"]
EXPENSE_CARDS = ["Citi", "Maybank", "Amex", "UOB"]  # fallback only — live list read from Cards sheet

# In-memory caches to avoid repeated Sheets API calls
_merchant_cache = None          # list of dicts from Merchant Map sheet
_merchant_cache_ts = None       # monotonic timestamp of last load
_MERCHANT_CACHE_TTL = 300       # 5 minutes
_card_names_cache = None        # list of card name strings
_system_prompt_cache = None     # cached system prompt string
_system_prompt_overseas_key = None  # tracks when to invalidate system prompt cache

# Cards sheet new schema: Card Name | Last 4 | Default Category | Notes
CARDS_SCHEMA = ["Card Name", "Last 4", "Default Category", "Notes"]
INITIAL_CARDS = [
    ["Maybank", "4002", "FnB", ""],
    ["Citi", "1176", "General", ""],
    ["UOB", "5372", "", ""],
    ["Amex", "1008", "", ""],
]

def cards_sheet():
    return spreadsheet.worksheet("Cards")

def get_cards_live():
    """Read card list from Cards sheet, cached in memory."""
    global _card_names_cache
    if _card_names_cache is not None:
        return _card_names_cache
    try:
        ws = cards_sheet()
        _card_names_cache = ws.get_all_records()
        return _card_names_cache
    except Exception as e:
        print(f"get_cards_live error: {e}")
        return []

def get_card_names_live():
    """Return list of card names from Cards sheet."""
    cards = get_cards_live()
    names = [c.get("Card Name", "") for c in cards if c.get("Card Name")]
    return names if names else EXPENSE_CARDS

def get_card_default_for_category(category):
    """Return default card name for a given category from Cards sheet."""
    try:
        cards = get_cards_live()
        for c in cards:
            default_cat = c.get("Default Category", "").strip().lower()
            if default_cat and default_cat == category.strip().lower():
                return c.get("Card Name", "")
        # Fallback defaults
        cat_lower = category.strip().lower()
        if cat_lower in ["fnb", "food", "dining", "f&b"]:
            return "Maybank"
        if cat_lower in ["grab", "transport"]:
            return "Amex"
        return "Citi"
    except Exception as e:
        print(f"get_card_default_for_category error: {e}")
        return "Citi"

def get_card_by_last4(last4):
    """Match a card by last 4 digits. Returns card name or None."""
    try:
        cards = get_cards_live()
        for c in cards:
            if str(c.get("Last 4", "")).strip() == str(last4).strip():
                return c.get("Card Name", "")
    except Exception as e:
        print(f"get_card_by_last4 error: {e}")
    return None

def set_card_default_category(card_name, category):
    """Update default category for a card in Cards sheet."""
    try:
        ws = cards_sheet()
        records = ws.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("Card Name", "").lower() == card_name.lower():
                # Default Category is col 3
                ws.update_cell(i, 3, category)
                return f"Updated — {card_name} will now default to {category} ✅"
        return f"Card '{card_name}' not found. Your cards: {', '.join(get_card_names_live())}"
    except Exception as e:
        return f"❌ Couldn't update card default: {str(e)}"

def fuzzy_match_card(text):
    """Fuzzy match text against card names. Returns (matched_name, exact) or (None, False)."""
    cards = get_card_names_live()
    text_lower = text.strip().lower()
    # Exact match
    for c in cards:
        if c.lower() == text_lower:
            return c, True
    # Substring match
    for c in cards:
        if text_lower in c.lower() or c.lower() in text_lower:
            return c, True
    # First 3 chars match
    for c in cards:
        if len(text_lower) >= 3 and c.lower().startswith(text_lower[:3]):
            return c, False
    return None, False

def fuzzy_match_category(text):
    """Fuzzy match text against expense categories. Returns (matched, exact) or (None, False)."""
    text_lower = text.strip().lower()
    # Exact match
    for c in EXPENSE_CATEGORIES:
        if c.lower() == text_lower:
            return c, True
    # Synonym map
    synonyms = {
        "food": "FnB", "dining": "FnB", "restaurant": "FnB", "eat": "FnB", "f&b": "FnB",
        "fun": "Entertainment", "movie": "Entertainment", "movies": "Entertainment",
        "gym": "Personal", "health": "Personal",
        "taxi": "Transport", "grab": "Transport", "uber": "Transport", "mrt": "Transport", "bus": "Transport",
        "clothes": "Shopping", "shop": "Shopping",
        "trip": "Travel", "holiday": "Travel", "flight": "Travel",
        "office": "Work", "business": "Work",
    }
    if text_lower in synonyms:
        return synonyms[text_lower], True
    # Substring match
    for c in EXPENSE_CATEGORIES:
        if text_lower in c.lower() or c.lower() in text_lower:
            return c, False
    return None, False

# Foreign transaction fee estimates per card (%)
# CARD_FX_FEES removed — was defined but never referenced

# Overseas mode state
overseas_state = {
    "active": False,
    "destination": "",
    "currency": "SGD",
    "currencies": [],          # all currencies used this trip (multi-currency support)
    "return_date": "",
    "dep_job_id": None,        # scheduler job ID for departure activation
    "return_job_id": None,     # scheduler job ID for return deactivation
    "trip_start": None,        # date string DD/MM/YYYY when overseas mode activated
    "trip_destinations": [],   # list of destinations visited this trip
}

def persist_trip_setup():
    """Persist _trip_setup state to Settings sheet. Fire-and-forget — never raises."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        ts = overseas_state.get("_trip_setup")
        data = json.dumps(ts) if ts else ""
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if r.get("Key") == "trip_setup_state":
                sheet.update_cell(i + 2, 2, data)
                return
        sheet.append_row(["trip_setup_state", data])
    except Exception as e:
        print(f"persist_trip_setup error: {e}")

def load_trip_setup_from_sheet():
    """Restore _trip_setup state from Settings sheet on startup."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "trip_setup_state":
                raw = r.get("Value", "")
                if raw:
                    ts = json.loads(raw)
                    overseas_state["_trip_setup"] = ts
                    print(f"Restored _trip_setup from sheet: step={ts.get('step')}")
                return
    except Exception as e:
        print(f"load_trip_setup_from_sheet error: {e}")

# Global scheduler reference (set in post_init)
_scheduler = None
_app_ref = None

# Pending new merchant — waiting for user to confirm category + card
# { user_id: { "merchant": str, "amount": float, "currency": str, "step": "category"|"card" } }
expense_sessions = {}
delete_sessions = {}
portfolio_delete_sessions = {}  # { user_id: { "step": "pick", "rows": [...] } }
confirm_sessions = {}  # { user_id: { "action": str, "args": list, "target": str } }
def persist_sessions_to_sheet():
    """Persist receipt_confirm_sessions to Settings sheet. Fire-and-forget — never raises."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        data = json.dumps({str(k): v for k, v in receipt_confirm_sessions.items()})
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if r.get("Key") == "receipt_confirm_sessions":
                sheet.update_cell(i + 2, 2, data)
                return
        sheet.append_row(["receipt_confirm_sessions", data])
    except Exception as e:
        print(f"persist_sessions_to_sheet error: {e}")

def load_sessions_from_sheet():
    """Restore receipt_confirm_sessions from Settings sheet on startup."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "receipt_confirm_sessions":
                raw = r.get("Value", "")
                if raw:
                    loaded = json.loads(raw)
                    receipt_confirm_sessions.update({int(k): v for k, v in loaded.items()})
                    if loaded:
                        print(f"Restored {len(loaded)} receipt confirm session(s) from sheet")
                return
    except Exception as e:
        print(f"load_sessions_from_sheet error: {e}")

class _AutoPersistDict(dict):
    """Dict subclass that auto-persists to Settings sheet on write/delete.
    Debounced: coalesces rapid successive writes into a single sheet call after 2s.
    persist_sessions_to_sheet must be defined before this class is instantiated."""
    _timer = None

    def _schedule_persist(self):
        import threading
        if _AutoPersistDict._timer is not None:
            _AutoPersistDict._timer.cancel()
        _AutoPersistDict._timer = threading.Timer(2.0, persist_sessions_to_sheet)
        _AutoPersistDict._timer.daemon = True
        _AutoPersistDict._timer.start()

    def __setitem__(self, key, value):
        super().__setitem__(key, value)
        self._schedule_persist()

    def __delitem__(self, key):
        super().__delitem__(key)
        self._schedule_persist()

receipt_confirm_sessions = _AutoPersistDict()  # { user_id: { "merchant": str, "amount": float, ... } }
todo_disambig_sessions = {}  # { user_id: { "tasks": list, "action": str } }
market_summary_pending = {}  # { user_id: True }
interrupted_sessions = {}  # { user_id: { "label": str, "pending_text": str } }

# Reconciliation sessions — { user_id: { "step": str, "unmatched": [...], "index": int } }
recon_sessions = {}

SESSION_TIMEOUT_MINUTES = 5
SESSION_TIMEOUT_MESSAGES = {
    "expense": "Session timed out — nothing was logged. Start again when ready.",
    "delete": "Session timed out — nothing was deleted.",
    "portfolio_delete": "Session timed out — nothing was removed.",
    "confirm": "Session timed out — nothing was changed.",
    "receipt_confirm": "Session timed out — nothing was logged. Send the receipt again when ready.",
    "edit": "Session timed out — nothing was changed.",
    "meeting": "Meeting recap session timed out.",
}

def touch_session(user_id):
    """Record current time for session timeout tracking."""
    session_timestamps[user_id] = datetime.now(TIMEZONE)

def is_session_expired(user_id):
    """Return True if the session for user_id has exceeded SESSION_TIMEOUT_MINUTES."""
    ts = session_timestamps.get(user_id)
    if not ts:
        return True
    elapsed = (datetime.now(TIMEZONE) - ts).total_seconds() / 60
    return elapsed > SESSION_TIMEOUT_MINUTES

def clear_all_sessions(user_id):
    """Clear all active sessions for a user."""
    for d in [expense_sessions, delete_sessions, portfolio_delete_sessions,
              confirm_sessions, receipt_confirm_sessions, recon_sessions,
              edit_sessions, meeting_sessions, session_timestamps]:
        d.pop(user_id, None)

async def check_session_timeouts(user_id, update):
    """Check all active sessions for timeout. Returns True if any session was expired and cleared."""
    if not is_session_expired(user_id):
        return False

    expired = False
    if user_id in expense_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["expense"])
        del expense_sessions[user_id]
        expired = True
    if user_id in delete_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["delete"])
        del delete_sessions[user_id]
        expired = True
    if user_id in portfolio_delete_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["portfolio_delete"])
        del portfolio_delete_sessions[user_id]
        expired = True
    if user_id in confirm_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["confirm"])
        del confirm_sessions[user_id]
        expired = True
    if user_id in receipt_confirm_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["receipt_confirm"])
        del receipt_confirm_sessions[user_id]
        expired = True
    if user_id in edit_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["edit"])
        del edit_sessions[user_id]
        expired = True
    if user_id in meeting_sessions:
        await update.message.reply_text(SESSION_TIMEOUT_MESSAGES["meeting"])
        del meeting_sessions[user_id]
        expired = True

    session_timestamps.pop(user_id, None)
    return expired



# Cached FX rates — { "JPY_SGD": { "rate": 0.0093, "fetched_at": datetime } }
cached_fx_rates = {}

def get_fx_rate(currency):
    """Return SGD rate for given currency. Uses cache if under 12hrs, else fetches fresh.
    Falls back to manual rate if API unavailable. No Claude estimate."""
    if currency == "SGD":
        return 1.0
    cache_key = f"{currency}_SGD"

    # Check live cache
    cached = cached_fx_rates.get(cache_key)
    if cached:
        age = datetime.now(pytz.utc) - cached["fetched_at"]
        if age.total_seconds() < 43200:  # 12 hours
            return cached["rate"]

    # Fetch fresh from API
    if EXCHANGE_RATE_API_KEY:
        try:
            url = f"https://v6.exchangerate-api.com/v6/{EXCHANGE_RATE_API_KEY}/pair/{currency}/SGD"
            with httpx.Client(timeout=8) as hx:
                resp = hx.get(url)
            data = resp.json()
            if data.get("result") == "success":
                rate = float(data["conversion_rate"])
                cached_fx_rates[cache_key] = {"rate": rate, "fetched_at": datetime.now(pytz.utc)}
                print(f"FX cache updated: 1 {currency} = {rate} SGD")
                # Clear manual rate if API is back
                if currency in manual_fx_rates:
                    del manual_fx_rates[currency]
                return rate
            else:
                print(f"FX API non-success for {currency}: {data.get('error-type', data)}")
        except Exception as e:
            print(f"FX fetch error for {currency}: {e}")

    # Check manual rate cache
    manual = manual_fx_rates.get(currency)
    if manual:
        today_str = date.today().strftime("%Y-%m-%d")
        if manual.get("date") == today_str:
            return manual["rate"]
        # Stale manual rate from previous day — reuse but notify
        print(f"Using previous manual rate for {currency}: {manual['rate']}")
        return manual["rate"]

    # No rate available — return None to trigger manual input prompt
    return None

def save_manual_fx_rate(currency, rate, sgd_per_unit=True):
    """Save a manually entered FX rate. sgd_per_unit=True means rate is how much SGD per 1 unit of currency."""
    if not sgd_per_unit:
        # User entered SGD to currency rate, e.g. $1 SGD = 110 JPY → 1 JPY = 1/110 SGD
        rate = 1.0 / rate if rate else 0
    manual_fx_rates[currency] = {
        "rate": rate,
        "date": date.today().strftime("%Y-%m-%d"),
        "sgd_per_unit": sgd_per_unit
    }
    # Also update live cache
    cached_fx_rates[f"{currency}_SGD"] = {"rate": rate, "fetched_at": datetime.now(pytz.utc)}
    print(f"Manual FX rate saved: 1 {currency} = {rate} SGD")
    persist_fx_rates_to_sheet()

def persist_fx_rates_to_sheet():
    """Save cached manual FX rates to Settings sheet so they survive restarts."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        fx_data = json.dumps({
            k: {"rate": v["rate"], "fetched_at": v["fetched_at"].isoformat()}
            for k, v in cached_fx_rates.items()
            if isinstance(v.get("fetched_at"), datetime)
        })
        for i, r in enumerate(records):
            if r.get("Key") == "cached_fx_rates":
                sheet.update_cell(i + 2, 2, fx_data)
                return
        sheet.append_row(["cached_fx_rates", fx_data])
    except Exception as e:
        print(f"persist_fx_rates_to_sheet error: {e}")

def load_fx_rates_from_sheet():
    """Load cached FX rates from Settings sheet on startup."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "cached_fx_rates":
                raw = r.get("Value", "")
                if not raw:
                    return
                data = json.loads(raw)
                cutoff = datetime.now(pytz.utc) - timedelta(hours=12)
                for k, v in data.items():
                    try:
                        fetched_at = datetime.fromisoformat(v["fetched_at"])
                        if fetched_at.tzinfo is None:
                            fetched_at = pytz.utc.localize(fetched_at)
                        if fetched_at > cutoff:
                            cached_fx_rates[k] = {"rate": v["rate"], "fetched_at": fetched_at}
                            print(f"Restored FX rate: {k} = {v['rate']}")
                    except Exception as e:
                        print(f"load_fx_rates_from_sheet: bad entry {k}: {e}")
                return
    except Exception as e:
        print(f"load_fx_rates_from_sheet error: {e}")

def parse_manual_fx_input(text, currency):
    """Parse manual FX rate input. Supports multiple formats:
    - '110' or '1 SGD = 110 JPY' → SGD to foreign (sgd_per_unit=False)
    - '0.0093' or '1 JPY = 0.0093 SGD' → foreign to SGD (sgd_per_unit=True)
    - '1sgd to 3.11 myr' or '1sgd to 3.11 my' → SGD to foreign (partial currency codes accepted)
    Returns (rate_as_sgd_per_unit, display_str) or (None, None)
    """
    try:
        text = text.strip().replace(",", "")
        text_lower = text.lower()

        # Normalise partial currency codes to 3-letter (e.g. "my" → "myr", "rp" → "idr")
        partial_currency_map = {
            "my": "myr", "rm": "myr",
            "rp": "idr",
            "bt": "thb", "baht": "thb",
            "vnd": "vnd", "dong": "vnd",
            "php": "php", "peso": "php",
        }
        for partial, full in partial_currency_map.items():
            text_lower = re.sub(rf"\b{partial}\b", full, text_lower)

        # Extract all numbers from text
        nums = re.findall(r"[\d]+\.[\d]+|[\d]+", text_lower)
        if not nums:
            return None, None

        # Try to detect "1 SGD to X CURRENCY" pattern
        to_pattern = re.search(
            r"1\s*sgd\s+(?:to|=|:)\s*([\d.]+)\s*\w*", text_lower
        )
        if to_pattern:
            num = float(to_pattern.group(1))
            if num <= 0:
                return None, None
            sgd_per_unit = False
            display = f"$1 SGD = {num:.4f} {currency}"
            return (num, sgd_per_unit), display

        # Try to detect "1 CURRENCY to X SGD" pattern
        curr_to_sgd = re.search(
            rf"1\s*{re.escape(currency.lower())}\s+(?:to|=|:)\s*([\d.]+)\s*(?:sgd)?", text_lower
        )
        if curr_to_sgd:
            num = float(curr_to_sgd.group(1))
            if num <= 0:
                return None, None
            sgd_per_unit = True
            display = f"1 {currency} = ${num:.4f} SGD"
            return (num, sgd_per_unit), display

        # Fall back to extracting first number and inferring direction
        num_match = re.search(r"[\d.]+", text)
        if not num_match:
            return None, None
        num = float(num_match.group())
        if num <= 0:
            return None, None

        # Determine direction from keyword context
        if "sgd" in text_lower and currency.lower() in text_lower:
            sgd_pos = text_lower.index("sgd")
            curr_pos = text_lower.index(currency.lower()) if currency.lower() in text_lower else -1
            if curr_pos > sgd_pos:
                sgd_per_unit = False
                display = f"$1 SGD = {num:.4f} {currency}"
            else:
                sgd_per_unit = True
                display = f"1 {currency} = ${num:.4f} SGD"
        elif num > 10:
            # Large number → likely SGD to foreign (e.g. 110 JPY per SGD)
            sgd_per_unit = False
            display = f"$1 SGD = {num:.2f} {currency}"
        else:
            # Small number → likely foreign to SGD (e.g. 0.0093 SGD per JPY)
            sgd_per_unit = True
            display = f"1 {currency} = ${num:.4f} SGD"

        return (num, sgd_per_unit), display
    except Exception as e:
        print(f"parse_manual_fx_input error: {e}")
        return None, None



async def refresh_fx_rates(app=None):
    """Refresh cached FX rates for all active overseas currencies. Called twice daily."""
    try:
        currencies = [c for c in overseas_state.get("currencies", []) if c != "SGD"]
        if not currencies:
            return
        for currency in currencies:
            cache_key = f"{currency}_SGD"
            if EXCHANGE_RATE_API_KEY:
                try:
                    url = f"https://v6.exchangerate-api.com/v6/{EXCHANGE_RATE_API_KEY}/pair/{currency}/SGD"
                    with httpx.Client(timeout=8) as hx:
                        resp = hx.get(url)
                    data = resp.json()
                    if data.get("result") == "success":
                        rate = float(data["conversion_rate"])
                        cached_fx_rates[cache_key] = {"rate": rate, "fetched_at": datetime.now(pytz.utc)}
                        print(f"FX refresh: 1 {currency} = {rate} SGD")
                except Exception as e:
                    print(f"FX refresh error for {currency}: {e}")
    except Exception as e:
        print(f"refresh_fx_rates error: {e}")

EXPENSE_CATEGORY_EMOJI = {
    "FnB": "🍽️",
    "Transport": "🚗",
    "Entertainment": "🎬",
    "Personal": "🪞",
    "Family": "👨‍👩‍👧",
    "Work": "💼",
    "Shopping": "🛍️",
    "Household": "🏠",
    "Travel": "✈️",
}

EXPENSE_MERCHANT_OVERRIDES = {
    "FnB": [
        (["coffee", "starbucks", "kopitiam", "ya kun", "toast box", "kopi", "cafe", "espresso", "latte"], "☕"),
        (["ramen", "ichiran", "ippudo", "japanese", "sushi", "yakitori", "donburi", "izakaya"], "🍜"),
        (["mcdonald", "burger king", "wendy", "burger", "kfc", "popeyes", "fast food"], "🍔"),
        (["bar", "beer", "wine", "drinks", "cocktail", "pub", "taproom", "brewery"], "🍺"),
    ],
    "Transport": [
        (["flight", "airline", "air asia", "scoot", "singapore airlines", "cathay", "emirates", "jetstar"], "✈️"),
        (["grab", "gojek", "taxi", "uber", "ryde", "tada"], "🚕"),
        (["mrt", "bus", "transit", "ez-link", "train"], "🚌"),
    ],
    "Entertainment": [
        (["netflix", "disney", "hbo", "prime video", "apple tv", "streaming", "hulu", "mewatch"], "📺"),
        (["spotify", "apple music", "tidal", "deezer", "music"], "🎵"),
        (["cinema", "cathay", "gv", "shaw", "golden village", "movie"], "🎥"),
        (["steam", "playstation", "xbox", "nintendo", "game"], "🎮"),
    ],
    "Shopping": [
        (["guardian", "watsons", "unity", "pharmacy", "watson"], "💊"),
        (["ntuc", "giant", "cold storage", "fairprice", "supermarket", "grocery", "market"], "🛒"),
    ],
}

_category_emoji_cache = {}

def get_merchant_emoji(category, merchant):
    """Return contextual emoji based on category + merchant name."""
    merchant_lower = merchant.lower() if merchant else ""
    overrides = EXPENSE_MERCHANT_OVERRIDES.get(category, [])
    for keywords, emoji in overrides:
        if any(kw in merchant_lower for kw in keywords):
            return emoji
    if category in EXPENSE_CATEGORY_EMOJI:
        return EXPENSE_CATEGORY_EMOJI[category]
    if category:
        if category in _category_emoji_cache:
            return _category_emoji_cache[category]
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=5,
                messages=[{"role": "user", "content": f"Return a single emoji that best represents the expense category '{category}'. Return ONLY the emoji, nothing else."}]
            )
            inferred = resp.content[0].text.strip()
            _category_emoji_cache[category] = inferred
            return inferred
        except Exception:
            pass
    return "💳"

def expenses_sheet():
    return spreadsheet.worksheet("Expenses")

def merchant_map_sheet():
    return spreadsheet.worksheet("Merchant Map")

def _get_merchant_records():
    """Return merchant map records from cache, loading from sheet if needed or TTL expired."""
    global _merchant_cache, _merchant_cache_ts
    import time as _time
    now = _time.monotonic()
    if _merchant_cache is None or _merchant_cache_ts is None or (now - _merchant_cache_ts) > _MERCHANT_CACHE_TTL:
        try:
            _merchant_cache = merchant_map_sheet().get_all_records()
            _merchant_cache_ts = now
        except Exception as e:
            print(f"_get_merchant_records error: {e}")
            return []
    return _merchant_cache

def _invalidate_merchant_cache():
    global _merchant_cache, _merchant_cache_ts
    _merchant_cache = None
    _merchant_cache_ts = None

def get_merchant_memory(merchant):
    """Look up known merchant in Merchant Map. Returns (category, card, canonical_name) or (None, None, None)."""
    try:
        records = _get_merchant_records()
        merchant_lower = merchant.lower().strip()
        # Exact match first
        for r in records:
            if r.get("Merchant", "").lower() == merchant_lower:
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
        # Fuzzy match — check if either is a substring of the other
        for r in records:
            known = r.get("Merchant", "").lower().strip()
            if known and (known in merchant_lower or merchant_lower in known):
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
        # Word-level match — first significant word (brand name)
        merchant_words = [w for w in merchant_lower.split() if len(w) > 2]
        for r in records:
            known = r.get("Merchant", "").lower().strip()
            known_words = [w for w in known.split() if len(w) > 2]
            if merchant_words and known_words and merchant_words[0] == known_words[0]:
                return r.get("Category", ""), r.get("Card", ""), r.get("Merchant", "")
    except Exception as e:
        print(f"get_merchant_memory error for '{merchant}': {e}")
    return None, None, None

def save_merchant_memory(merchant, category, card):
    """Save new merchant to Merchant Map. Returns True on success, False on failure."""
    try:
        sheet = merchant_map_sheet()
        sheet.append_row([merchant, category, card])
        _invalidate_merchant_cache()
        return True
    except Exception as e:
        print(f"save_merchant_memory failed for '{merchant}': {e}")
        log_error_to_em_log("save_merchant_memory", f"{merchant} — {e}")
        return False

def delete_merchant(merchant_name):
    """Delete a merchant from the Merchant Map by name (fuzzy match)."""
    try:
        sheet = merchant_map_sheet()
        records = sheet.get_all_values()
        if len(records) <= 1:
            return "No merchants saved yet."
        matches = []
        for i, row in enumerate(records[1:], start=2):
            if row and merchant_name.lower() in row[0].lower():
                matches.append((i, row[0]))
        if not matches:
            return f"No merchant found matching \'{merchant_name}\'."
        if len(matches) == 1:
            row_idx, found_name = matches[0]
            sheet.delete_rows(row_idx)
            _invalidate_merchant_cache()
            return f"Deleted \'{found_name}\' from merchant memory ✅\nNext time you log there, I\'ll ask for category and card again."
        lines = [f"Found {len(matches)} merchants matching \'{merchant_name}\' — which one?"]
        for i, (_, name) in enumerate(matches, 1):
            lines.append(f"{i}. {name}")
        lines.append("\nReply \'delete merchant [exact name]\' to remove a specific one.")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error deleting merchant: {e}"

def parse_expense_text(text):
    """Redirect to v2 which uses live category/card lists."""
    return parse_expense_text_v2(text)
def parse_expense_text_v2(text):
    """Updated parse_expense_text using live category/card lists."""
    overseas_currency = overseas_state["currency"] if overseas_state["active"] else "SGD"
    live_cats = ", ".join(EXPENSE_CATEGORIES)
    live_cards = ", ".join(get_card_names_live())
    prompt = (
        f"Extract expense details from this message: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- merchant: string (brand name only, strip legal suffixes)\n"
        f"- amount: number (just the number, no currency symbol)\n"
        f"- currency: string (3-letter ISO code. If not mentioned, default to '{overseas_currency}')\n"
        f"- category: string (one of: {live_cats}) or empty if unclear\n"
        f"- card: string (one of: {live_cards}) or empty if not mentioned\n\n"
        f"Return ONLY the JSON."
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
        parsed = json.loads(raw)
        # Validate amount before returning — prevents unhandled float() cast downstream
        if "amount" in parsed:
            try:
                parsed["amount"] = float(parsed["amount"])
            except (ValueError, TypeError):
                parsed["amount"] = None
        return parsed
    except Exception as e:
        print(f"parse_expense_text Claude error: {e}")
        return None

def log_expense(merchant, amount, currency, sgd_amount, category, card, receipt_link="", reconciled="No", notes=""):
    """Append expense row to Expenses sheet. Returns True on success, False on failure."""
    try:
        today = date.today().strftime("%d/%m/%Y")
        sheet = expenses_sheet()
        sheet.append_row([
            today, merchant, amount, currency, sgd_amount,
            category, card, receipt_link, reconciled, notes
        ])
        _invalidate_expense_cache()
        return True
    except Exception as e:
        print(f"log_expense failed: {e}")
        log_error_to_em_log("log_expense", f"{merchant} {amount} {currency} — {e}")
        return False

def format_expense_confirmation(merchant, amount, currency, category, card, sgd_amount=None,
                                receipt_saved=False, last4=None, high_amount=False,
                                use_manual_rate=False, missing_fields=None):
    """Format expense confirmation in hybrid format with new spec."""
    emoji = get_merchant_emoji(category, merchant)
    lines = [f"{emoji} {merchant}"]

    # Amount line
    if currency != "SGD" and sgd_amount is not None:
        amount_str = f"${sgd_amount:.2f} ({amount:,.0f} {currency})"
    else:
        amount_str = f"${amount:.2f}" if amount else "⚠️ Amount?"

    # Category
    cat_str = f"⚠️ Category?" if not category else category

    # Card
    if not card:
        card_str = "⚠️ Card?"
    elif last4:
        card_str = f"{card} (*{last4})"
    else:
        card_str = card

    detail_line = f"{amount_str} | {cat_str} | {card_str}"
    lines.append(detail_line)

    # Flags
    if high_amount:
        lines.append("⚠️ Amount looks high — is this correct?")
    if use_manual_rate:
        lines.append("⚠️ Using manual rate — verify if needed")
    if receipt_saved:
        lines.append("🧾 Receipt saved to Drive")

    # Missing fields prompt
    if missing_fields:
        if len(missing_fields) >= 3:
            prompt = "yes / enter missing values / skip"
        elif len(missing_fields) == 2:
            prompt = f"yes / enter {' & '.join(f.title() for f in missing_fields)} / skip"
        else:
            prompt = f"yes / enter {missing_fields[0].title()} / skip"
        lines.append(f"\nLog this? ({prompt})")
    else:
        lines.append("\nLog this? (yes / edit / skip)")

    return "\n".join(lines)

def format_expense_logged(merchant, amount, currency, category, card, sgd_amount=None, last4=None):
    """Format the logged confirmation (after yes or after edit)."""
    emoji = get_merchant_emoji(category, merchant)
    if currency != "SGD" and sgd_amount is not None:
        amount_str = f"${sgd_amount:.2f} ({amount:,.0f} {currency})"
    else:
        amount_str = f"${amount:.2f}"
    card_str = f"{card} (*{last4})" if last4 else card
    return f"{emoji} {merchant}\n{amount_str} | {category} | {card_str}\n\nLogged ✅"



def get_monthly_summary(month=None, year=None):
    """Generate monthly expense summary."""
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No expenses logged yet."

        today = date.today()
        target_month = month or today.month
        target_year = year or today.year

        month_records = []
        for r in records:
            d = r.get("Date", "")
            if not d:
                continue
            try:
                dt = datetime.strptime(d, "%d/%m/%Y")
                if dt.month == target_month and dt.year == target_year:
                    month_records.append(r)
            except Exception as e:
                print(f"get_monthly_summary: bad date in row: {e}")

        if not month_records:
            return f"No expenses for {datetime(target_year, target_month, 1).strftime('%B %Y')}."

        total = sum(float(r.get("SGD Amount", 0) or r.get("Amount", 0)) for r in month_records)

        # Category breakdown
        cat_totals = {c: 0 for c in EXPENSE_CATEGORIES}
        for r in month_records:
            cat = r.get("Category", "")
            if cat in cat_totals:
                amt = float(r.get("SGD Amount", 0) or r.get("Amount", 0))
                cat_totals[cat] += amt

        # Card breakdown
        card_totals = {c: 0 for c in EXPENSE_CARDS}
        for r in month_records:
            card = r.get("Card", "")
            if card in card_totals:
                amt = float(r.get("SGD Amount", 0) or r.get("Amount", 0))
                card_totals[card] += amt

        month_label = datetime(target_year, target_month, 1).strftime("%B %Y")
        lines = [f"Monthly Summary — {month_label}\n"]

        for cat, amt in cat_totals.items():
            if amt > 0:
                lines.append(f"{cat}: ${amt:.2f}")

        lines.append(f"\nTotal: ${total:.2f}")
        lines.append("\nBy card:")
        for card, amt in card_totals.items():
            if amt > 0:
                lines.append(f"{card}: ${amt:.2f}")

        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error generating summary: {str(e)}"

def is_log_prefix_input(text):
    """Detect 'log [merchant] [amount]' — always routes to expense flow."""
    return text.lower().startswith("log ") and len(text) > 4

def is_bare_merchant_input(text):
    """Detect bare merchant name + amount — e.g. 'Starbucks $5.60'.
    Only triggers if: first word/phrase matches a known merchant AND a dollar amount is present.
    Safeguard: avoids triggering on sentences like 'Apple announced new products'.
    """
    lower = text.lower().strip()
    # Must contain a dollar amount
    if not re.search(r"\$[\d,.]+", text):
        return False
    # First word must match a known merchant (case-insensitive)
    first_word = text.strip().split()[0]
    known_cat, _, _ = get_merchant_memory(first_word)
    return bool(known_cat)

def get_expense_categories():
    """Pull live expense categories from the Expenses sheet data, falling back to defaults."""
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        cats = sorted(set(r.get("Category", "").strip() for r in records if r.get("Category", "").strip()))
        if not cats:
            cats = EXPENSE_CATEGORIES
    except Exception:
        cats = EXPENSE_CATEGORIES
    numbered = "\n".join(f"{i+1}. {c}" for i, c in enumerate(cats))
    return f"Your expense categories:\n\n{numbered}"


def get_merchant_list():
    """Pull all known merchants from Merchant Map, grouped by category."""
    try:
        sheet = merchant_map_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No merchants saved yet. They are added automatically when you log a new expense."
        by_cat = {}
        for r in records:
            merchant = r.get("Merchant", "").strip()
            cat = r.get("Category", "").strip() or "Uncategorised"
            card = r.get("Card", "").strip()
            if not merchant:
                continue
            if cat not in by_cat:
                by_cat[cat] = []
            entry = merchant + (f" ({card})" if card else "")
            by_cat[cat].append(entry)
        count = sum(len(v) for v in by_cat.values())
        lines = [f"Merchant Map ({count} merchants):\n"]
        for cat in sorted(by_cat):
            lines.append(f"*{cat}*")
            for m in sorted(by_cat[cat]):
                lines.append(f"  {m}")
            lines.append("")
        return "\n".join(lines).strip()
    except Exception as e:
        return f"❌ Error fetching merchants: {e}"

def is_expense_input(text):
    """Detect if message looks like an expense entry or expense question."""
    # Flight numbers are never expenses — guard before anything else
    if extract_flight_number(text):
        return False
    lower = text.lower()
    # Share purchases are stock, not expenses
    if re.search(r"\d+ shares|shares of |shares @|shares at \$", lower):
        return False
    # Exclude command-style messages that contain expense-related words but aren't entries
    exclusions = [
        "delete expense", "remove expense", "undo expense",
        "edit expense", "edit last expense",
        "rename category", "expense report", "monthly report",
        "trip summary", "last expense", "show last expense",
        "what expense categories", "list my categories", "what categories",
        "show categories", "what are my expense", "list categories",
        "expense categories", "my categories",
        "what merchants", "my merchants", "merchant map", "list merchants",
        "show merchants", "known merchants"
    ]
    if any(lower.startswith(e) or lower == e for e in exclusions):
        return False
    triggers = ["spent", "paid", "$", "sgd", "charged", "bought", "grabbed",
                "receipt", "bill was", "cost me", "picked up",
                "recorded in", "will it be", "logged in", "track", "how much have i"]
    return any(t in lower for t in triggers)

LOCATION_CONTEXT_WORDS = {
    "japan", "korea", "thailand", "malaysia", "indonesia", "vietnam", "philippines",
    "australia", "china", "hong kong", "taiwan", "india", "uk", "england", "france",
    "germany", "italy", "spain", "usa", "america", "canada", "dubai", "uae",
    "tokyo", "osaka", "seoul", "bangkok", "kuala lumpur", "kl", "jakarta", "bali",
    "sydney", "melbourne", "beijing", "shanghai", "guangzhou", "shenzhen",
    "taipei", "mumbai", "delhi", "london", "paris", "berlin", "rome", "barcelona",
    "new york", "los angeles", "chicago", "toronto", "vancouver",
    "hkg", "nrt", "icn", "bkk", "kul", "cgk", "sin", "syd", "pek", "pvg",
    "tpe", "bom", "del", "lhr", "cdg", "txl", "fco", "jfk", "lax", "yyz",
}

def is_overseas_mode_request(text):
    """Detect overseas mode toggle. Guards against misfires like 'i'm in a meeting'."""
    lower = text.lower()
    flights = extract_flight_number(text) if callable(extract_flight_number) else []
    has_flight = bool(flights)
    has_multiple_flights = len(re.findall(r'\b[A-Z]{1,3}\d{2,4}[A-Z]?\b', text.upper())) >= 2

    # Two flight numbers in one message = always a travel message
    if has_multiple_flights:
        return True

    # Single flight + travel intent words
    travel_words = [
        "flying", "flight", "boarding", "departure", "departing",
        "returning", "flying back", "headed to", "heading to",
        "on tr", "on sq", "on ak", "on mh", "on od", "on ek", "on cx",
    ]
    if has_flight and any(w in lower for w in travel_words):
        return True

    # Explicit travel phrases (no flight number needed)
    if any(phrase in lower for phrase in [
        "overseas", "travelling", "traveling", "flying to", "arrived in",
        "back home", "i'm back", "landed in", "just landed", "just arrived",
        "returned home"
    ]):
        return True

    # "i'm in [location]" — only if followed by a known place name
    if "i'm in" in lower or "im in" in lower:
        return any(loc in lower for loc in LOCATION_CONTEXT_WORDS)

    return False



def deactivate_overseas_mode():
    """Deactivate overseas mode and clear scheduled jobs."""
    global overseas_state, _scheduler
    try:
        close_trip()
    except Exception:
        pass
    overseas_state["active"] = False
    overseas_state["destination"] = ""
    overseas_state["currency"] = "SGD"
    overseas_state["currencies"] = []
    overseas_state["return_date"] = ""
    overseas_state["trip_start"] = None
    overseas_state["trip_destinations"] = []
    for job_key in ["dep_job_id", "return_job_id"]:
        job_id = overseas_state.get(job_key)
        if job_id and _scheduler:
            try:
                _scheduler.remove_job(job_id)
            except Exception:
                pass
        overseas_state[job_key] = None


async def activate_overseas_mode_scheduled(dest, curr, check_in, check_out):
    """Called by scheduler at departure time to activate overseas mode."""
    global overseas_state, _app_ref
    for d in [expense_sessions, receipt_confirm_sessions]:
        d.pop(YOUR_CHAT_ID, None)
    session_timestamps.pop(YOUR_CHAT_ID, None)
    overseas_state["active"] = True
    overseas_state["destination"] = dest
    overseas_state["currency"] = curr
    overseas_state["currencies"] = [curr] if curr != "SGD" else []
    overseas_state["trip_start"] = date.today().strftime("%d/%m/%Y")
    overseas_state["trip_destinations"] = [dest]
    save_trip(dest, curr, check_in=check_in, check_out=check_out)
    msg = f"Overseas mode on ✈️\nDestination: {dest}\nCurrency: {curr}\nI'll log expenses in {curr} with SGD equivalent."
    if _app_ref:
        try:
            await _app_ref.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
        except Exception as e:
            print(f"Failed to send overseas mode activation message: {e}")


async def deactivate_and_notify(app):
    """Called by scheduler at return arrival — deactivate overseas mode and notify."""
    import random
    dest = overseas_state.get("destination", "")
    deactivate_overseas_mode()
    greeting = random.choice(["Welcome back!", "Good to have you back!", "Hope the trip was great!"])
    msg = f"{greeting} Back in SG — switching to SGD. 🏠"
    try:
        await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
    except Exception as e:
        print(f"Failed to send return notification: {e}")


def extract_flight_number(text):
    """Extract first IATA-style flight number from text (e.g. TR450, SQ321)."""
    matches = re.findall(r'\b([A-Z]{1,3}\d{2,4}[A-Z]?)\b', text.upper())
    return matches[0] if matches else None


def _parse_trip_dates(text):
    """Extract one or two dates from free text. Returns list of date objects."""
    today = date.today()
    found = []
    seen = set()
    month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                 "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    yr = today.year

    def _add(d):
        if d not in seen:
            seen.add(d)
            found.append(d)

    for word, delta in [("today", 0), ("tomorrow", 1), ("tmr", 1), ("tmrw", 1)]:
        if word in text.lower():
            _add(today + timedelta(days=delta))

    for m in re.finditer(r'\b(\d{1,2})(?:st|nd|rd|th)?\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:\s+(\d{4}))?\b', text, re.IGNORECASE):
        try:
            _add(date(int(m.group(3)) if m.group(3) else yr, month_map[m.group(2).lower()], int(m.group(1))))
        except ValueError:
            pass

    for m in re.finditer(r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})(?:st|nd|rd|th)?(?:\s+(\d{4}))?\b', text, re.IGNORECASE):
        try:
            _add(date(int(m.group(3)) if m.group(3) else yr, month_map[m.group(1).lower()], int(m.group(2))))
        except ValueError:
            pass

    for m in re.finditer(r'\b(\d{4}-\d{2}-\d{2})\b', text):
        try:
            _add(datetime.strptime(m.group(1), "%Y-%m-%d").date())
        except ValueError:
            pass

    for m in re.finditer(r'\b(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?\b', text):
        try:
            y = int(m.group(3)) if m.group(3) else yr
            if y < 100:
                y += 2000
            _add(date(y, int(m.group(2)), int(m.group(1))))
        except ValueError:
            pass

    return found


def _get_currency_for_dest(destination):
    """Ask Haiku for the currency of a destination. Returns 3-letter ISO code."""
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=10,
            messages=[{"role": "user", "content":
                f"What is the primary currency ISO code for {destination}? Reply with ONLY the 3-letter code."}]
        )
        code = resp.content[0].text.strip().upper()
        if re.match(r'^[A-Z]{3}$', code):
            return code
    except Exception as e:
        print(f"_get_currency_for_dest error: {e}")
    return "SGD"


def handle_overseas_request(text):
    """Start a trip setup session or handle return-home."""
    global overseas_state
    lower = text.lower()

    # Returning home
    if any(p in lower for p in ["back home", "returned", "i'm back", "landed back", "home now"]):
        import random
        greeting = random.choice(["Welcome back!", "Good to have you back!", "Hope the trip was great!"])
        deactivate_overseas_mode()
        return f"{greeting} Switching back to SGD. 🏠"

    # Extract flight number if present — begin trip setup session with it
    flight_num = extract_flight_number(text)

    # Try to extract destination from text (e.g. "flying to Bangkok", "going to KL")
    dest_match = re.search(r'(?:to|in|flying to|going to|headed to|heading to|trip to)\s+([A-Za-z][A-Za-z\s]{2,25}?)(?:\s+on|\s+\d|$|[,.])', text, re.IGNORECASE)
    dest_hint = dest_match.group(1).strip().title() if dest_match else None

    overseas_state["_trip_setup"] = {
        "step": "destination",
        "flight_number": flight_num or "",
        "destination": dest_hint or "",
        "check_in": "",
        "check_out": "",
        "currency": "",
        "hotel_name": "",
        "hotel_local_name": "",
        "hotel_address": "",
        "notes": "",
    }

    if dest_hint:
        overseas_state["_trip_setup"]["step"] = "check_in"
        persist_trip_setup()
        return f"Got it — {dest_hint} 🌏\nCheck-in date? (or 'skip')"
    persist_trip_setup()
    return "Where are you headed?"


async def _send_trip_confirm(update, ts):
    """Send trip confirmation summary asking Y/N."""
    dest = ts.get("destination", "—")
    curr = ts.get("currency", "") or "auto-detect"
    check_in = ts.get("check_in", "") or "—"
    check_out = ts.get("check_out", "") or "—"
    flight = ts.get("flight_number", "")
    dep_display = ts.get("dep_time_display", "")
    hotel = ts.get("hotel_name", "") or "—"

    lines = [f"✈️ *{dest}* ({curr})"]
    lines.append(f"Check-in: {check_in} → Check-out: {check_out}")
    if flight:
        lines.append(f"Flight: {flight}" + (f" @ {dep_display}" if dep_display else ""))
    if hotel != "—":
        lines.append(f"Hotel: {hotel}")
    lines.append("\nConfirm? (Y / cancel)")

    ts["step"] = "confirm"
    overseas_state["_trip_setup"] = ts
    persist_trip_setup()
    try:
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception:
        await update.message.reply_text("\n".join(lines))


def parse_multi_field_edit(text, field_keywords=None):
    """Parse a multi-field edit message like 'merchant Tsukemen Fuunji category Dining'.
    Returns dict of {field: value}. Supports synonyms and quoted merchant names."""
    if field_keywords is None:
        field_keywords = {
            "merchant": "merchant", "shop": "merchant", "store": "merchant", "place": "merchant",
            "amount": "amount", "price": "amount", "total": "amount", "cost": "amount",
            "currency": "currency",
            "category": "category", "cat": "category",
            "card": "card", "payment": "card", "pay": "card",
        }
    result = {}
    # Handle quoted merchant names first
    text = re.sub(r'"([^"]+)"', lambda m: m.group(0).replace(" ", "_SPACE_"), text)
    tokens = text.strip().split()
    i = 0
    current_field = None
    current_value_parts = []

    def flush():
        if current_field and current_value_parts:
            val = " ".join(current_value_parts).replace("_SPACE_", " ").strip('"')
            result[current_field] = val

    while i < len(tokens):
        token_lower = tokens[i].lower()
        if token_lower in field_keywords:
            flush()
            current_field = field_keywords[token_lower]
            current_value_parts = []
        elif current_field:
            current_value_parts.append(tokens[i])
        i += 1
    flush()
    return result

async def handle_receipt_confirm_session(user_id, text, update):
    """Handle receipt/expense confirm session: yes / skip / field edits / card name override."""
    session = receipt_confirm_sessions.get(user_id)
    if not session:
        return False

    touch_session(user_id)
    lower = text.strip().lower()

    # Cancel / skip
    if lower in ["skip", "cancel", "no", "n", "nope", "nah"]:
        del receipt_confirm_sessions[user_id]
        session_timestamps.pop(user_id, None)
        receipt_link = session.get("receipt_link", "")
        if receipt_link:
            await update.message.reply_text("Entry not logged — receipt saved to Drive if you need it later.")
        else:
            await update.message.reply_text("Entry not logged.")
        return True

    # FX rate input
    if session.get("step") == "fx_rate":
        currency = session.get("currency", "")
        parsed_rate, display = parse_manual_fx_input(text, currency)
        if parsed_rate is None:
            await update.message.reply_text(
                f"Couldn't read that rate — try one of these formats:\n"
                f"• `3.11` (1 {currency} = 3.11 SGD)\n"
                f"• `1 SGD to 3.11 {currency}`\n"
                f"• `1 {currency} = 3.11 SGD`"
            )
            return True  # Stay in session, don't drop
        num, sgd_per_unit = parsed_rate
        save_manual_fx_rate(currency, num, sgd_per_unit)
        rate = get_fx_rate(currency)
        session["sgd_amount"] = round(session["amount"] * rate, 2)
        session["step"] = "receipt_confirm"
        session["use_manual_rate"] = True
        receipt_confirm_sessions[user_id] = session
        await update.message.reply_text(
            f"Got it — using {display} for today.\n\n" +
            format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                last4=session.get("last4"), high_amount=session.get("high_amount", False),
                use_manual_rate=True, missing_fields=session.get("missing_fields", [])
            )
        )
        return True

    # Duplicate confirm
    if session.get("step") == "duplicate_confirm":
        if lower in ["yes", "y", "yep", "yeah", "yup"]:
            receipt_confirm_sessions[user_id] = session
            await update.message.reply_text(
                format_expense_confirmation(
                    session["merchant"], session["amount"], session["currency"],
                    session.get("category", ""), session.get("card", ""),
                    session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                    last4=session.get("last4"), high_amount=session.get("high_amount", False),
                    use_manual_rate=session.get("use_manual_rate", False),
                    missing_fields=session.get("missing_fields", [])
                )
            )
        else:
            del receipt_confirm_sessions[user_id]
            await update.message.reply_text("Skipped duplicate.")
        return True

    if lower in ["yes", "y", "yep", "yeah", "yup"]:
        await _finalise_receipt_confirm(user_id, session, update)
        return True

    # Check if it's a single card name override
    card_match, _ = fuzzy_match_card(text.strip())
    if card_match and len(text.strip().split()) == 1:
        session["card"] = card_match
        # Log immediately after card override
        await _finalise_receipt_confirm(user_id, session, update)
        return True

    # Shorthand edit prompts — "edit name", "edit amount", etc.
    edit_field_map = {
        "edit name": "merchant", "edit merchant": "merchant",
        "edit amount": "amount", "edit price": "amount",
        "edit category": "category", "edit cat": "category",
        "edit card": "card", "edit payment": "card",
    }
    if lower in edit_field_map:
        field = edit_field_map[lower]
        session["_awaiting_edit"] = field
        receipt_confirm_sessions[user_id] = session
        field_label = {"merchant": "merchant name", "amount": "amount", "category": "category", "card": "card"}[field]
        await update.message.reply_text(f"Enter the new {field_label}:")
        return True

    # Handle awaiting edit value
    if session.get("_awaiting_edit"):
        field = session.pop("_awaiting_edit")
        receipt_confirm_sessions[user_id] = session
        # Reuse multi-field edit logic by constructing the edit string
        synthetic = f"{field} {text.strip()}"
        edits = parse_multi_field_edit(synthetic)
        if edits:
            for f, value in edits.items():
                if f == "merchant":
                    session["merchant"] = value
                elif f == "amount":
                    try:
                        session["amount"] = float(re.sub(r"[^\d.]", "", value))
                        if session["currency"] == "SGD":
                            session["sgd_amount"] = session["amount"]
                    except ValueError:
                        pass
                elif f == "category":
                    matched_cat, _ = fuzzy_match_category(value)
                    if matched_cat:
                        session["category"] = matched_cat
                        session["card"] = get_card_default_for_category(matched_cat)
                elif f == "card":
                    matched_card, _ = fuzzy_match_card(value)
                    if matched_card:
                        session["card"] = matched_card
        receipt_confirm_sessions[user_id] = session
        await update.message.reply_text(
            format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), receipt_saved=bool(session.get("receipt_link")),
                last4=session.get("last4"), high_amount=session.get("high_amount", False),
                use_manual_rate=session.get("use_manual_rate", False),
                missing_fields=session.get("missing_fields", [])
            )
        )
        return True

    # Try multi-field edit
    edits = parse_multi_field_edit(text)
    if edits:
        # Check for keyword-in-merchant-name ambiguity
        merchant_val = edits.get("merchant", "")
        field_keys = {"merchant", "amount", "currency", "category", "card", "price", "total", "cost", "shop", "store", "place", "payment", "pay", "cat"}
        if merchant_val:
            words = merchant_val.lower().split()
            collision = [w for w in words if w in field_keys]
            if collision and '"' not in text:
                # Ambiguity detected — confirm with user using Option A
                # Build what we parsed
                lines = ["Just to confirm:"]
                for f, v in edits.items():
                    lines.append(f"{f.title()}: {v}")
                # Add unchanged fields
                for f in ["amount", "category", "card"]:
                    if f not in edits:
                        val = session.get(f, "")
                        if val:
                            lines.append(f"{f.title()}: {val}")
                lines.append("\nyes / no / edit")
                session["_pending_edits"] = edits
                receipt_confirm_sessions[user_id] = session
                await update.message.reply_text("\n".join(lines))
                return True

        # Apply edits
        for field, value in edits.items():
            if field == "merchant":
                session["merchant"] = value
                # Re-run merchant memory
                known_cat, known_card, canonical = get_merchant_memory(value)
                if canonical:
                    session["merchant"] = canonical
                if known_cat and not edits.get("category"):
                    session["category"] = known_cat
                # Always re-derive card from category, not merchant memory
                if session.get("category") and not edits.get("card"):
                    session["card"] = get_card_default_for_category(session["category"])
            elif field == "amount":
                try:
                    session["amount"] = float(re.sub(r"[^\d.]", "", value))
                    if session["currency"] == "SGD":
                        session["sgd_amount"] = session["amount"]
                except ValueError:
                    pass
            elif field == "currency":
                session["currency"] = value.upper()
            elif field == "category":
                matched_cat, _ = fuzzy_match_category(value)
                if matched_cat:
                    session["category"] = matched_cat
                    # Update card default if not explicitly set
                    if not edits.get("card"):
                        session["card"] = get_card_default_for_category(matched_cat)
            elif field == "card":
                matched_card, _ = fuzzy_match_card(value)
                if matched_card:
                    session["card"] = matched_card

        # Clear missing fields that are now filled
        session["missing_fields"] = [f for f in session.get("missing_fields", [])
                                      if not session.get(f)]
        session["high_amount"] = session.get("sgd_amount", 0) > SGD_HIGH_AMOUNT_THRESHOLD

        # Log immediately after edit — no re-confirm
        await _finalise_receipt_confirm(user_id, session, update)
        return True

    # Handle pending ambiguity confirm
    if session.get("_pending_edits"):
        if lower in ["yes", "y"]:
            edits = session.pop("_pending_edits")
            for field, value in edits.items():
                session[field] = value
            await _finalise_receipt_confirm(user_id, session, update)
        elif lower in ["no", "n"]:
            session.pop("_pending_edits", None)
            receipt_confirm_sessions[user_id] = session
            await update.message.reply_text(
                'Re-enter with quotes for merchant names containing field keywords.\n'
                'e.g. merchant "Gift Card Store" category Shopping'
            )
        return True

    # Unrecognised input
    await update.message.reply_text("Didn't catch that — reply yes to log, skip to cancel, or edit fields (e.g. category FnB card Citi)")
    return True

async def _finalise_receipt_confirm(user_id, session, update):
    """Log expense from receipt confirm session and send logged confirmation."""
    merchant = session["merchant"]
    amount = session["amount"]
    currency = session["currency"]
    category = session.get("category", "")
    card = session.get("card", "Citi")
    sgd_amount = session.get("sgd_amount", amount)
    receipt_link = session.get("receipt_link", "")
    last4 = session.get("last4")
    is_new_merchant = session.get("is_new_merchant", False)

    success = log_expense(merchant, amount, currency, sgd_amount, category, card, receipt_link=receipt_link)
    if not success:
        await update.message.reply_text("⚠️ Failed to save expense to sheet — please try again or log manually.")
        return

    # Only delete session and save merchant memory after confirmed write
    del receipt_confirm_sessions[user_id]
    session_timestamps.pop(user_id, None)

    if is_new_merchant and category and card:
        if not save_merchant_memory(merchant, category, card):
            print(f"Warning: merchant memory not saved for '{merchant}' — will re-ask next time")

    reply = format_expense_logged(merchant, amount, currency, category, card, sgd_amount, last4)
    try:
        await update.message.reply_text(reply, parse_mode="Markdown")
    except Exception:
        await update.message.reply_text(reply)

async def handle_expense_session(user_id, text, update):
    """Handle multi-step expense onboarding — fx_rate only. Confirm flow now handled by receipt_confirm_sessions."""
    session = expense_sessions.get(user_id)
    if not session:
        return

    touch_session(user_id)
    step = session.get("step")

    if step == "fx_rate":
        currency = session.get("currency", "")
        parsed_rate, display = parse_manual_fx_input(text, currency)
        if parsed_rate is None:
            await update.message.reply_text(
                f"Couldn't read that rate — try one of these formats:\n"
                f"• `3.11` (1 {currency} = 3.11 SGD)\n"
                f"• `1 SGD to 3.11 {currency}`\n"
                f"• `1 {currency} = 3.11 SGD`"
            )
            return  # Stay in session, don't drop
        num, sgd_per_unit = parsed_rate
        save_manual_fx_rate(currency, num, sgd_per_unit)
        rate = get_fx_rate(currency)
        session["sgd_amount"] = round(session["amount"] * rate, 2)
        session["use_manual_rate"] = True

        # Move to receipt confirm flow
        del expense_sessions[user_id]
        receipt_confirm_sessions[user_id] = {**session, "step": "receipt_confirm",
                                              "missing_fields": [], "is_new_merchant": True}
        touch_session(user_id)
        await update.message.reply_text(
            f"Got it — using {display} for today.\n\n" +
            format_expense_confirmation(
                session["merchant"], session["amount"], currency,
                session.get("category", ""), session.get("card", ""),
                session["sgd_amount"], use_manual_rate=True,
                missing_fields=[]
            )
        )
        return

    if step == "duplicate_confirm":
        lower = text.strip().lower()
        if lower in ["yes", "y", "yep", "yeah", "yup"]:
            receipt_confirm_sessions[user_id] = {**session, "step": "receipt_confirm"}
            touch_session(user_id)
            confirmation = format_expense_confirmation(
                session["merchant"], session["amount"], session["currency"],
                session.get("category", ""), session.get("card", ""),
                session.get("sgd_amount"), missing_fields=[]
            )
            await update.message.reply_text(confirmation)
        else:
            del expense_sessions[user_id]
            await update.message.reply_text("Skipped.")
        return

# _finalise_expense_session removed — was legacy dead code, use _finalise_receipt_confirm directly

# In-memory cache for same-day duplicate detection: set of (date, merchant_lower, amount_str, currency)
_same_day_expense_cache: set = set()
_same_day_expense_cache_date: str = ""

def _invalidate_expense_cache():
    global _same_day_expense_cache, _same_day_expense_cache_date
    _same_day_expense_cache = set()
    _same_day_expense_cache_date = ""

def check_same_day_duplicate(merchant, amount, currency):
    """Return True if same merchant+amount+currency already logged today.
    Uses in-memory cache populated on first call each day; invalidated on write."""
    global _same_day_expense_cache, _same_day_expense_cache_date
    today = date.today().strftime("%d/%m/%Y")
    try:
        if _same_day_expense_cache_date != today:
            sheet = expenses_sheet()
            records = sheet.get_all_records()
            _same_day_expense_cache = {
                (r.get("Date", ""), r.get("Merchant", "").lower(),
                 str(r.get("Amount", "")), r.get("Currency", ""))
                for r in records if r.get("Date") == today
            }
            _same_day_expense_cache_date = today
        return (today, merchant.lower(), str(amount), currency) in _same_day_expense_cache
    except Exception as e:
        print(f"Duplicate check error: {e}")
    return False

QUESTION_WORDS = {"what", "how", "why", "when", "which", "who", "where", "list", "show", "tell", "is", "are", "do", "does", "can", "could", "would", "should"}
COMMAND_PREFIXES = ["delete", "remove", "undo", "edit", "rename", "show", "list", "what", "how"]

def _looks_like_command(text):
    """Return True if text looks like a command or question rather than an expense entry."""
    lower = text.lower().strip()
    first_word = lower.split()[0] if lower.split() else ""
    if first_word in QUESTION_WORDS:
        return True
    if any(lower.startswith(p) for p in COMMAND_PREFIXES):
        return True
    return False

SGD_HIGH_AMOUNT_THRESHOLD = 5000.0

def handle_expense_text(text, user_id, receipt_link="", last4=None):
    """
    Process a text expense entry. Returns (reply_str, needs_session, session_data).
    Option B: auto-log if fully known, confirm screen if anything missing or uncertain.
    """
    try:
        parsed = parse_expense_text(text)
        if not parsed:
            return (
                "Couldn't read that as an expense — try: 'Starbucks $5.60' or 'spent $12 on lunch'."
            ), False, None
        merchant = parsed.get("merchant", "Unknown")
        try:
            amount = float(parsed.get("amount", 0) or 0)
        except (ValueError, TypeError):
            print(f"handle_expense_text: invalid amount from parse: {parsed.get('amount')!r}")
            amount = 0
        currency = parsed.get("currency", "SGD")
        category = parsed.get("category", "")
        card = parsed.get("card", "")

        # Blank expense guard
        if not merchant or merchant.lower() in ("unknown", "") or amount == 0:
            return (
                "Wasn't sure what to do with that — did you mean to log an expense, "
                "or were you asking something else?"
            ), False, None

        # Sanity check: zero or negative
        if amount <= 0:
            return "Amount can't be zero or negative — what's the correct amount?", False, None

        # Fuzzy merchant lookup — use canonical name if found
        known_cat, known_card, canonical = get_merchant_memory(merchant)
        if canonical:
            merchant = canonical
        if known_cat and not category:
            category = known_cat
        # Card priority: (1) explicit user input, (2) category default, (3) global fallback
        # Never let merchant memory or category default override an explicit card from the user
        explicit_card = card  # preserve what the parser extracted from user text

        if not explicit_card:
            # No explicit card — derive from category
            if category:
                card = get_card_default_for_category(category)
            else:
                card = "Citi"  # global fallback
        # else: keep explicit_card as-is

        # Card override from last4 (receipt scan) — only if no explicit card
        if last4 and not explicit_card:
            matched_card = get_card_by_last4(last4)
            if matched_card:
                card = matched_card

        # Resolve FX rate if needed
        sgd_amount = amount
        use_manual_rate = False
        if currency != "SGD":
            rate = get_fx_rate(currency)
            if rate is None:
                # Check if we have a stale manual rate to use
                manual = manual_fx_rates.get(currency)
                if manual:
                    rate = manual["rate"]
                    use_manual_rate = True
                    await_msg = (
                        f"Live rate unavailable — using manual rate: 1 {currency} = ${rate:.4f} SGD\n"
                        f"I'll keep checking for the live rate."
                    )
                else:
                    # No rate at all — ask for manual input
                    session = {
                        "merchant": merchant, "amount": amount, "currency": currency,
                        "category": category, "card": card, "step": "fx_rate",
                        "receipt_link": receipt_link, "last4": last4
                    }
                    return (
                        f"Couldn't get the {currency}/SGD rate right now.\n"
                        f"Enter the exchange rate:\n"
                        f"• SGD to {currency}: e.g. 110 (meaning $1 SGD = {currency} 110)\n"
                        f"• {currency} to SGD: e.g. 0.0093 (meaning 1 {currency} = $0.0093 SGD)"
                    ), True, session
            sgd_amount = round(amount * rate, 2)
            if overseas_state["active"] and currency not in overseas_state["currencies"]:
                overseas_state["currencies"].append(currency)

        # High amount sanity check
        high_amount = sgd_amount > SGD_HIGH_AMOUNT_THRESHOLD

        # Same-day duplicate check
        if check_same_day_duplicate(merchant, amount, currency):
            session = {
                "merchant": merchant, "amount": amount, "currency": currency,
                "sgd_amount": sgd_amount, "category": category, "card": card,
                "step": "duplicate_confirm", "receipt_link": receipt_link, "last4": last4
            }
            return (
                f"Heads up — looks like you already logged {merchant} {currency} {amount:,.0f} today. Log it again? (yes / no)"
            ), True, session

        # Determine missing fields
        missing = []
        if not category:
            missing.append("category")
        if not amount:
            missing.append("amount")

        is_new_merchant = not bool(known_cat)

        # Option B: auto-log only if everything known, no flags, AND it's a known merchant
        # New merchants always go through confirm screen
        if not missing and not high_amount and not use_manual_rate and not is_new_merchant:
            success = log_expense(merchant, amount, currency, sgd_amount, category, card, receipt_link=receipt_link)
            if not success:
                return "⚠️ Failed to save expense — please try again.", False, None
            save_merchant_memory(merchant, category, card)
            return format_expense_logged(merchant, amount, currency, category, card, sgd_amount, last4), False, None

        # Otherwise show confirm screen
        session = {
            "merchant": merchant, "amount": amount, "currency": currency,
            "sgd_amount": sgd_amount, "category": category, "card": card,
            "step": "receipt_confirm", "missing_fields": missing,
            "receipt_link": receipt_link, "last4": last4,
            "high_amount": high_amount, "use_manual_rate": use_manual_rate,
            "is_new_merchant": is_new_merchant
        }
        receipt_sessions_key = user_id
        receipt_confirm_sessions[receipt_sessions_key] = session
        confirmation = format_expense_confirmation(
            merchant, amount, currency, category, card, sgd_amount,
            receipt_saved=bool(receipt_link), last4=last4,
            high_amount=high_amount, use_manual_rate=use_manual_rate,
            missing_fields=missing
        )
        return confirmation, False, None  # Session stored separately in receipt_confirm_sessions

    except (json.JSONDecodeError, ValueError, KeyError) as e:
        print(f"parse_expense_text error: {e} | input: {text}")
        return (
            f"❌ Couldn't parse that as an expense ({type(e).__name__}: {str(e)[:60]}).\n"
            "Try: 'log [merchant] [amount]' — e.g. 'log Starbucks $5.60'"
        ), False, None
    except Exception as e:
        print(f"handle_expense_text unexpected error: {e}")
        return f"❌ Something went wrong logging that ({type(e).__name__}: {str(e)[:80]}). Try again or use 'log [merchant] [amount]'.", False, None



def delete_last_expense():
    """Delete the most recently added expense row."""
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return "No expenses to delete."
        last_row = len(all_values)
        last = all_values[last_row - 1]
        sheet.delete_rows(last_row)
        return f"Deleted last expense: {last[1]} ${last[2]}"
    except Exception as e:
        return f"❌ Error deleting expense: {str(e)}"

def get_recent_expenses(n=5):
    """Return the last N expense rows as list of (sheet_row_index, dict)."""
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return []
        headers = all_values[0]
        rows = all_values[1:]
        recent = rows[-(n):]
        result = []
        for i, row in enumerate(recent):
            sheet_row = len(all_values) - len(recent) + i + 1  # 1-indexed sheet row
            d = {headers[j]: row[j] if j < len(row) else "" for j in range(len(headers))}
            result.append((sheet_row, d))
        return list(reversed(result))  # Most recent first
    except Exception as e:
        print(f"get_recent_expenses error: {e}")
        return []

def format_delete_list(expenses):
    """Format numbered list of expenses for delete selection."""
    lines = ["Which expense to delete?\n"]
    for i, (_, r) in enumerate(expenses, 1):
        merchant = r.get("Merchant", "?")
        amount = r.get("Amount", "")
        currency = r.get("Currency", "SGD")
        sgd = r.get("SGD Amount", "")
        date_str = r.get("Date", "")
        category = r.get("Category", "")
        if currency != "SGD" and sgd:
            amt_str = f"${float(sgd):.2f} ({currency} {float(amount):,.0f})"
        else:
            amt_str = f"${float(amount):.2f}" if amount else "$?"
        lines.append(f"{i}. {merchant} — {amt_str} | {category} | {date_str}")
    lines.append("\nReply with a number, or 'search [merchant]' to find another.")
    return "\n".join(lines)

def search_expenses_by_merchant(query):
    """Search expenses by merchant name, return list of (sheet_row, dict)."""
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return []
        headers = all_values[0]
        rows = all_values[1:]
        q = query.lower().strip()
        matches = []
        for i, row in enumerate(rows):
            d = {headers[j]: row[j] if j < len(row) else "" for j in range(len(headers))}
            if q in d.get("Merchant", "").lower():
                matches.append((i + 2, d))  # +2: 1-indexed + header row
        return list(reversed(matches))[:10]  # Most recent first, cap at 10
    except Exception as e:
        print(f"search_expenses_by_merchant error: {e}")
        return []

def delete_expense_by_row(sheet_row):
    """Delete an expense by its sheet row number. Returns confirmation string."""
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if sheet_row < 2 or sheet_row > len(all_values):
            return "Couldn't find that expense."
        row = all_values[sheet_row - 1]
        merchant = row[1] if len(row) > 1 else "?"
        amount = row[2] if len(row) > 2 else "?"
        sheet.delete_rows(sheet_row)
        return f"Deleted: {merchant} ${amount}"
    except Exception as e:
        return f"Error deleting expense: {str(e)}"

def get_last_expense():
    """Return the last logged expense row as a dict."""
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        if not records:
            return None
        return records[-1]
    except Exception as e:
        print(f"get_last_expense error: {e}")
        return None

EDIT_FIELD_SYNONYMS = {
    "merchant": "Merchant", "shop": "Merchant", "store": "Merchant", "place": "Merchant",
    "amount": "Amount", "price": "Amount", "total": "Amount", "cost": "Amount",
    "currency": "Currency",
    "category": "Category", "cat": "Category",
    "card": "Card", "payment": "Card",
    "notes": "Notes", "note": "Notes",
    "sgd": "SGD Amount",
}

def edit_last_expense(edit_text):
    """Edit fields in the last expense row. Supports multi-field edits in one message.
    Returns formatted logged message after editing."""
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if len(all_values) <= 1:
            return "No expenses to edit."
        headers = all_values[0]
        last_row_idx = len(all_values)
        last_row = all_values[last_row_idx - 1]

        # Parse multi-field edits
        edits = parse_multi_field_edit(edit_text, field_keywords=EDIT_FIELD_SYNONYMS)
        if not edits:
            return "Couldn't parse that edit. Try: 'edit merchant Starbucks category FnB'"

        applied = []
        merchant_changed = False

        for field_key, new_value in edits.items():
            col_name = EDIT_FIELD_SYNONYMS.get(field_key.lower())
            if not col_name or col_name not in headers:
                continue

            col_idx = headers.index(col_name) + 1

            # Validate and normalise per field
            if col_name == "Amount":
                try:
                    new_value = str(float(re.sub(r"[^\d.]", "", new_value)))
                except ValueError:
                    continue
            elif col_name == "Category":
                matched_cat, _ = fuzzy_match_category(new_value)
                if not matched_cat:
                    continue
                new_value = matched_cat
            elif col_name == "Card":
                matched_card, _ = fuzzy_match_card(new_value)
                if not matched_card:
                    continue
                new_value = matched_card
            elif col_name == "Merchant":
                merchant_changed = True

            sheet.update_cell(last_row_idx, col_idx, new_value)
            applied.append((col_name, new_value))

        if not applied:
            return "Couldn't match any valid fields to edit."

        # If merchant changed, re-run merchant memory for category/card
        if merchant_changed:
            new_merchant = next((v for k, v in edits.items() if EDIT_FIELD_SYNONYMS.get(k) == "Merchant"), "")
            if new_merchant:
                known_cat, known_card, canonical = get_merchant_memory(new_merchant)
                if canonical:
                    m_col = headers.index("Merchant") + 1
                    sheet.update_cell(last_row_idx, m_col, canonical)
                if known_cat and "Category" not in [c for c, _ in applied]:
                    c_col = headers.index("Category") + 1
                    sheet.update_cell(last_row_idx, c_col, known_cat)
                    applied.append(("Category", known_cat))
                if known_card and "Card" not in [c for c, _ in applied]:
                    cd_col = headers.index("Card") + 1
                    sheet.update_cell(last_row_idx, cd_col, known_card)
                    applied.append(("Card", known_card))

        # Re-read updated row to show logged confirmation
        updated_values = sheet.get_all_values()
        updated_row_data = updated_values[last_row_idx - 1] if len(updated_values) >= last_row_idx else last_row
        updated = {headers[i]: updated_row_data[i] if i < len(updated_row_data) else "" for i in range(len(headers))}

        merchant = updated.get("Merchant", "")
        amount = updated.get("Amount", "")
        currency = updated.get("Currency", "SGD")
        sgd_amount = updated.get("SGD Amount", amount)
        category = updated.get("Category", "")
        card = updated.get("Card", "")

        return format_expense_logged(merchant, float(amount) if amount else 0,
                                     currency, category, card,
                                     float(sgd_amount) if sgd_amount else None)

    except Exception as e:
        return f"❌ Error editing expense: {str(e)}"



def show_last_expense():
    """Return a formatted view of the last logged expense."""
    r = get_last_expense()
    if not r:
        return "No expenses logged yet."
    currency = r.get("Currency", "SGD")
    amount = r.get("Amount", "")
    sgd = r.get("SGD Amount", "")
    _merchant = r.get('Merchant', '')
    _cat = r.get('Category', '')
    _emoji = get_merchant_emoji(_cat, _merchant)
    lines = [
        f"Last expense:",
        f"{_emoji} {_merchant}",
    ]
    if currency != "SGD" and sgd:
        lines.append(f"${float(sgd):.2f} ({currency} {float(amount):,.0f})")
    else:
        lines.append(f"${float(amount):.2f}" if amount else "")
    lines.append(f"🗂 {r.get('Category', '')} | 💳 {r.get('Card', '')}")
    lines.append(f"📅 {r.get('Date', '')}")
    lines.append("\nTo edit: 'edit [field] to [value]' e.g. 'edit category to Transport'")
    return "\n".join(l for l in lines if l)

def get_trip_summary():
    """Return expense summary for the current or most recent trip."""
    trip_start = overseas_state.get("trip_start")
    destinations = overseas_state.get("trip_destinations", [])
    if not trip_start:
        return "No trip data — overseas mode hasn't been activated this session."
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        start_dt = datetime.strptime(trip_start, "%d/%m/%Y").date()
        trip_records = []
        for r in records:
            d = r.get("Date", "")
            if not d:
                continue
            try:
                dt = datetime.strptime(d, "%d/%m/%Y").date()
                if dt >= start_dt:
                    trip_records.append(r)
            except Exception:
                continue
        if not trip_records:
            return "No expenses logged for this trip yet."
        total_sgd = sum(float(r.get("SGD Amount") or r.get("Amount") or 0) for r in trip_records)
        cat_totals = {}
        currency_totals = {}
        for r in trip_records:
            cat = r.get("Category", "Other")
            amt = float(r.get("SGD Amount") or r.get("Amount") or 0)
            cat_totals[cat] = cat_totals.get(cat, 0) + amt
            curr = r.get("Currency", "SGD")
            orig = float(r.get("Amount") or 0)
            if curr != "SGD":
                currency_totals[curr] = currency_totals.get(curr, 0) + orig
        dest_str = " → ".join(destinations) if destinations else "trip"
        lines = [f"Trip summary — {dest_str}", f"From {trip_start}\n"]
        for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
            lines.append(f"{cat}: SGD ${amt:.2f}")
        lines.append(f"\nTotal: SGD ${total_sgd:.2f}")
        if currency_totals:
            lines.append("\nSpend by currency:")
            for curr, amt in currency_totals.items():
                lines.append(f"{curr}: {amt:,.0f}")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error generating trip summary: {str(e)}"

def rename_category(old_name, new_name):
    """Rename a category everywhere — category list, Merchant Map, Expenses sheet."""
    global EXPENSE_CATEGORIES
    old_title = old_name.strip().title() if old_name.lower() not in [c.lower() for c in EXPENSE_CATEGORIES] else next(c for c in EXPENSE_CATEGORIES if c.lower() == old_name.lower())
    matched = next((c for c in EXPENSE_CATEGORIES if c.lower() == old_name.lower()), None)
    if not matched:
        return f"Category '{old_name}' not found. Current categories: {', '.join(EXPENSE_CATEGORIES)}"
    new_clean = new_name.strip()
    EXPENSE_CATEGORIES = [new_clean if c == matched else c for c in EXPENSE_CATEGORIES]
    updated = 0
    # Update Expenses sheet
    try:
        sheet = expenses_sheet()
        all_values = sheet.get_all_values()
        if all_values:
            headers = all_values[0]
            if "Category" in headers:
                col_idx = headers.index("Category")
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) > col_idx and row[col_idx] == matched:
                        sheet.update_cell(i, col_idx + 1, new_clean)
                        updated += 1
    except Exception as e:
        print(f"rename_category expenses error: {e}")
    # Update Merchant Map
    try:
        sheet = merchant_map_sheet()
        all_values = sheet.get_all_values()
        if all_values:
            headers = all_values[0]
            if "Category" in headers:
                col_idx = headers.index("Category")
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) > col_idx and row[col_idx] == matched:
                        sheet.update_cell(i, col_idx + 1, new_clean)
    except Exception as e:
        print(f"rename_category merchant map error: {e}")
    # Update Cards sheet Default Category column
    try:
        ws = cards_sheet()
        records = ws.get_all_records()
        for i, r in enumerate(records, start=2):
            if r.get("Default Category", "").strip().lower() == matched.lower():
                ws.update_cell(i, 3, new_clean)  # Default Category is col 3
    except Exception as e:
        print(f"rename_category cards sheet error: {e}")

    return f"Renamed '{matched}' to '{new_clean}'. Updated {updated} expense(s)."

def get_expense_report(report_type="monthly"):
    """Generate expense report."""
    return get_monthly_summary()



# =============================================================================
# BILL REMINDERS (Step 6)
# =============================================================================

BILL_REMINDER_GREETINGS = [
    "Heads up", "Just a nudge", "Quick reminder", "Hey", "FYI"
]

def bills_sheet():
    return spreadsheet.worksheet("Bills")

def parse_bill_request(text):
    """Parse a natural language bill setup request."""
    prompt = (
        f"Extract bill details from: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- name: string (bill name, e.g. 'Citi Credit Card', 'Netflix', 'Electricity')\n"
        f"- bank: string (bank or provider name, or empty)\n"
        f"- due_day: number (day of month the bill is due, e.g. 15)\n"
        f"- estimated_amount: number (estimated amount in SGD, or 0 if not mentioned)\n"
        f"- notes: string (any extra notes, or empty)\n\n"
        f"Return ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, ValueError) as e:
        print(f"parse_bill_request JSON error: {e} | raw: {raw[:100]}")
        return None

def add_bill(name, bank, due_day, estimated_amount, notes=""):
    """Add a bill to the Bills sheet."""
    sheet = bills_sheet()
    sheet.append_row([name, bank, str(due_day), str(estimated_amount), notes])

def list_bills():
    """List all bills."""
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No bills set up yet."
        lines = ["Your bills:\n"]
        for r in records:
            amt = r.get("Estimated Amount", "")
            amt_str = f" — ~${amt}" if amt and str(amt) != "0" else ""
            lines.append(f"• {r.get('Name', '')} (due day {r.get('Due Date', '')}){amt_str}")
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error listing bills: {str(e)}"

def delete_bill(name):
    """Delete a bill by name."""
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if name.lower() in r.get("Name", "").lower():
                sheet.delete_rows(i + 2)
                return f"Deleted bill: {r.get('Name')}"
        return f"No bill found matching '{name}'."
    except Exception as e:
        return f"❌ Error deleting bill: {str(e)}"

def get_cycle_expenses(card_name, due_day):
    """Sum expenses for a card in the current billing cycle."""
    try:
        sheet = expenses_sheet()
        records = sheet.get_all_records()
        today = date.today()
        # Billing cycle: from last due date to today
        if today.day >= due_day:
            cycle_start = today.replace(day=due_day)
        else:
            if today.month == 1:
                cycle_start = today.replace(year=today.year - 1, month=12, day=due_day)
            else:
                cycle_start = today.replace(month=today.month - 1, day=due_day)

        total = 0
        for r in records:
            try:
                dt = datetime.strptime(r.get("Date", ""), "%d/%m/%Y").date()
                card = r.get("Card", "")
                if dt >= cycle_start and card_name.lower() in card.lower():
                    total += float(r.get("SGD Amount", 0) or r.get("Amount", 0))
            except Exception as e:
                print(f"get_cycle_expenses: bad row: {e}")
        return total
    except Exception as e:
        print(f"get_cycle_expenses error: {e}")
        return 0

async def send_bill_reminders(app):
    """Daily 9am job — check for bills due in 7 days."""
    import random
    try:
        sheet = bills_sheet()
        records = sheet.get_all_records()
        today = date.today()

        for r in records:
            try:
                due_day = int(r.get("Due Date", 0))
                if not due_day:
                    continue

                # Calculate next due date
                if today.day <= due_day:
                    next_due = today.replace(day=due_day)
                else:
                    if today.month == 12:
                        next_due = today.replace(year=today.year + 1, month=1, day=due_day)
                    else:
                        next_due = today.replace(month=today.month + 1, day=due_day)

                days_away = (next_due - today).days

                if days_away == 7:
                    name = r.get("Name", "")
                    estimated = r.get("Estimated Amount", "")

                    # Try to get actual cycle spend
                    cycle_total = get_cycle_expenses(name, due_day)
                    amount_str = f"${cycle_total:.2f} logged this cycle" if cycle_total > 0 else (f"~${estimated}" if estimated and str(estimated) != "0" else "amount unknown")

                    greeting = random.choice(BILL_REMINDER_GREETINGS)
                    msg = (
                        f"{greeting} — your {name} bill is due in 7 days ({next_due.strftime('%d %b')}).\n"
                        f"{amount_str}."
                    )
                    await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)

            except Exception as e:
                print(f"Bill reminder error for {r.get('Name', '?')}: {e}")

    except Exception as e:
        print(f"Error in send_bill_reminders: {e}")

def is_bill_request(text):
    """Detect bill setup intent.
    'remind me about my' removed — collides with is_reminder_request.
    Bill intent always has 'bill', 'due', or 'credit card' explicitly."""
    lower = text.lower()
    # "due on the" only counts if followed by a day number (e.g. "due on the 15th")
    has_due_on_the = bool(re.search(r"due on the \d{1,2}", lower))
    triggers = ["bill is due", "bill due", "set up a bill", "add a bill",
                "credit card bill", "due every", "my citi bill", "my maybank bill",
                "my amex bill", "my uob bill", "my credit card bill"]
    return has_due_on_the or any(t in lower for t in triggers)

def handle_new_bill(text):
    """Parse and save a new bill."""
    try:
        parsed = parse_bill_request(text)
        name = parsed.get("name", "")
        bank = parsed.get("bank", "")
        due_day = parsed.get("due_day", 0)
        estimated_amount = parsed.get("estimated_amount", 0)
        notes = parsed.get("notes", "")

        if not name or not due_day:
            return "Couldn't get the bill details. Try: 'my Citi bill is due on the 15th, usually around $800'."

        add_bill(name, bank, due_day, estimated_amount, notes)
        amt_str = f", estimated ~${estimated_amount}" if estimated_amount else ""
        return f"Got it, I'll remind you about your {name} bill 7 days before it's due (day {due_day} of each month{amt_str})."
    except Exception as e:
        return f"❌ Couldn't save that bill: {str(e)}"


# =============================================================================
# RESTAURANT TRACKER (Step 7)
# =============================================================================

def restaurants_sheet():
    return spreadsheet.worksheet("Restaurants")

def parse_restaurant_save(text):
    """Parse a restaurant save request — name, location, tags, notes."""
    prompt = (
        f"Extract restaurant details from: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- name: string (restaurant name)\n"
        f"- location: string (address or area, e.g. 'Teck Lim Road' or 'Shibuya, Tokyo')\n"
        f"- country: string (country, default 'Singapore' if not mentioned)\n"
        f"- tags: string (comma-separated tags like 'date night, japanese, omakase' — only if mentioned, else empty)\n"
        f"- notes: string (any notes like 'need reservation', 'cash only' — only if mentioned, else empty)\n\n"
        f"Return ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
    return json.loads(raw)

def lookup_restaurant_from_maps(url):
    """Extract restaurant name and address from a Google Maps URL via Claude."""
    prompt = (
        f"Given this Google Maps URL: {url}\n\n"
        f"Extract the restaurant/place name and address from the URL itself (don't browse it).\n"
        f"Return ONLY a JSON object with:\n"
        f"- name: string (place name if visible in URL, else empty)\n"
        f"- location: string (area or address if visible, else empty)\n"
        f"- country: string (country if determinable from URL, else 'Singapore')\n\n"
        f"Return ONLY the JSON."
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
        result = json.loads(raw)
        # Flag if location is blank — needs manual input
        if not result.get("location"):
            result["_needs_location"] = True
        return result
    except Exception as e:
        print(f"lookup_restaurant_from_maps error: {e}")
        return {"name": "", "location": "", "country": "Singapore", "_needs_location": True}

def save_restaurant(name, location, country="Singapore", tags="", notes="", force_new=False):
    """Save a restaurant to the Restaurants sheet. Checks for duplicates unless force_new."""
    try:
        if not force_new:
            try:
                sheet = restaurants_sheet()
                records = sheet.get_all_records()
                for r in records:
                    if r.get("Name", "").lower() == name.lower():
                        return "_DUPLICATE_:" + name
            except Exception as e:
                print(f"save_restaurant duplicate check error: {e}")
        sheet = restaurants_sheet()
        sheet.append_row([name, location, country, tags, notes])
        return None  # Success
    except Exception as e:
        print(f"save_restaurant error: {e}")
        return f"_ERROR_:{str(e)}"

def format_restaurant_saved(name, location, tags="", notes="", country="Singapore"):
    """Format the restaurant saved confirmation."""
    lines = ["Saved!"]
    lines.append(f"🏪 {name}")
    loc_str = f"{location}, {country}" if country and country != "Singapore" else location
    lines.append(f"📍 {loc_str}")
    if tags:
        lines.append(f"🏷 {tags}")
    if notes:
        lines.append(f"📝 {notes}")
    return "\n".join(lines)

def search_restaurants(query):
    """Search restaurants by cuisine, location, or tag."""
    try:
        sheet = restaurants_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No restaurants saved yet."

        query_lower = query.lower()
        results = []
        for r in records:
            searchable = " ".join(str(v).lower() for v in r.values())
            if query_lower in searchable:
                results.append(r)

        if not results:
            return f"No restaurants found matching '{query}'."

        lines = [f"{len(results)} restaurant(s) found:\n"]
        for r in results:
            line = f"🏪 {r.get('Name', '')} — 📍 {r.get('Location', '')}"
            if r.get("Tags"):
                line += f" ({r.get('Tags')})"
            lines.append(line)
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error searching restaurants: {str(e)}"

def list_restaurants(country_filter=None):
    """List all saved restaurants, optionally filtered by country."""
    try:
        sheet = restaurants_sheet()
        records = sheet.get_all_records()
        if not records:
            return "No restaurants saved yet."

        if country_filter:
            records = [r for r in records if country_filter.lower() in r.get("Country", "").lower()]

        if not records:
            return f"No restaurants saved for {country_filter}."

        lines = [f"{len(records)} saved restaurant(s):\n"]
        for r in records:
            line = f"🏪 {r.get('Name', '')} — 📍 {r.get('Location', '')}"
            if r.get("Tags"):
                line += f" ({r.get('Tags')})"
            lines.append(line)
        return "\n".join(lines)
    except Exception as e:
        return f"❌ Error listing restaurants: {str(e)}"

def delete_restaurant(name):
    """Delete a restaurant by name."""
    try:
        sheet = restaurants_sheet()
        records = sheet.get_all_records()
        for i, r in enumerate(records):
            if name.lower() in r.get("Name", "").lower():
                sheet.delete_rows(i + 2)
                return f"Removed {r.get('Name')} from your list."
        return f"No restaurant found matching '{name}'."
    except Exception as e:
        return f"❌ Error: {str(e)}"

def is_restaurant_review_request(text):
    """Detect restaurant review request intent. Avoids colliding with CRM contact lookup."""
    lower = text.lower()
    # Explicit review triggers — always match
    if any(t in lower for t in ["reviews for", "review of", "reviews of"]):
        return True
    # "how is X" / "how's X" — only match if X looks like a place, not a person
    # Person names are typically 1-2 words with capitals; place names often have
    # food/place context words or are followed by nothing (just the name)
    how_match = re.search(r"how(?:'s| is)\s+(.+?)(?:\?|$)", lower)
    if how_match:
        subject = how_match.group(1).strip()
        # Reject if it looks like a person (single capitalised name, common person triggers)
        person_signals = ["he", "she", "they", "him", "her", "doing", "feeling", "been"]
        if any(p in subject.split() for p in person_signals):
            return False
        # Accept if food/place words present
        food_signals = ["restaurant", "place", "cafe", "ramen", "sushi", "bbq", "bar",
                        "bistro", "hawker", "eatery", "kitchen", "grill"]
        if any(f in subject for f in food_signals):
            return True
        # Accept if followed by nothing (just "how is Ichiran" — short subject, no person context)
        if len(subject.split()) <= 3 and not any(p in subject for p in person_signals):
            return True
    # "is it good" / "worth going" with restaurant context
    restaurant_context = ["restaurant", "place", "cafe", "bar", "eatery"]
    if any(t in lower for t in ["is it good", "is it worth", "worth going", "worth visiting", "any good"]):
        if any(c in lower for c in restaurant_context):
            return True
    return False

def get_restaurant_emoji(name, cuisine_hint=""):
    """Pick a contextual emoji for a restaurant based on name/cuisine."""
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=10,
            messages=[{
                "role": "user",
                "content": (
                    f"Pick ONE emoji that best represents this restaurant: {name}. "
                    f"Context: {cuisine_hint or 'Singapore restaurant'}. "
                    "Consider cuisine, vibe, occasion. Reply with just the emoji, nothing else."
                )
            }]
        )
        emoji = resp.content[0].text.strip()
        return emoji if emoji else "🍽"
    except Exception:
        return "🍽"

def get_restaurant_review(name):
    """Fetch RSS headlines and generate formatted review with emoji, 2 bullets, and plain text summary."""
    try:
        import xml.etree.ElementTree as ET
        headlines = []
        queries = [f"{name} restaurant review Singapore", f"{name} Singapore food"]
        for q_text in queries:
            if len(headlines) >= 4:
                break
            q = q_text.replace(" ", "+")
            url = f"https://news.google.com/rss/search?q={q}&hl=en-SG&gl=SG&ceid=SG:en"
            try:
                with httpx.Client(timeout=5) as hx:
                    resp = hx.get(url)
                root = ET.fromstring(resp.content)
                for item in root.findall(".//item"):
                    if len(headlines) >= 4:
                        break
                    title = item.findtext("title", "").split(" - ")[0].strip()
                    if title and title not in headlines:
                        headlines.append(title)
            except Exception as e:
                print(f"Restaurant review RSS error for {name}: {e}")
                continue

        if not headlines:
            return f"Couldn't find recent reviews for {name} — try searching online for the latest."

        # Look up area from saved restaurants sheet, then normalise to neighbourhood
        area = ""
        try:
            ws = restaurants_sheet()
            records = ws.get_all_records()
            for r in records:
                if name.lower() in r.get("Name", "").lower():
                    area = r.get("Location", "").strip()
                    break
        except Exception as e:
            print(f"Restaurant area lookup error: {e}")

        # Always normalise raw location to neighbourhood via Claude
        try:
            location_hint = f"The restaurant's saved address is: {area}. " if area else ""
            area_resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=20,
                messages=[{
                    "role": "user",
                    "content": (
                        f"{location_hint}What neighbourhood or area is '{name}' in Singapore known to be in? "
                        "Reply with just the neighbourhood name (e.g. Dempsey Hill, Tanjong Pagar, Orchard). "
                        "If unknown, reply 'Singapore'."
                    )
                }]
            )
            area = area_resp.content[0].text.strip()
        except Exception:
            area = area or ""

        headline_text = "\n".join(f"- {h}" for h in headlines)
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=250,
                messages=[{
                    "role": "user",
                    "content": (
                        f"Here are headlines about the restaurant '{name}':\n{headline_text}\n\n"
                        "Write a review in this exact format:\n"
                        "EMOJI: [one contextual emoji for the restaurant vibe/cuisine — varied, creative, not always 🍽]\n"
                        "BULLET1: [one sentence — a highlight or strength, grounded in the headlines]\n"
                        "BULLET2: [one sentence — a caveat, limitation, or honest note]\n"
                        "SUMMARY: [exactly 2 short sentences — overall impression. Plain, honest, no hedging about sources.]\n"
                    )
                }]
            )
            raw = resp.content[0].text.strip()

            # Parse sections
            def get_field(text, key):
                if key not in text:
                    return ""
                line = text.split(key)[-1].split("\n")[0].strip().lstrip(":").strip()
                return line

            emoji = get_field(raw, "EMOJI:")
            bullet1 = get_field(raw, "BULLET1:")
            bullet2 = get_field(raw, "BULLET2:")
            summary = get_field(raw, "SUMMARY:")

            if not emoji:
                emoji = "🍽"

            area_str = f" ({area})" if area else ""
            result = f"{emoji} *{name}*{area_str}\n"
            if bullet1:
                result += f"\n• {bullet1}\n"
            if bullet2:
                result += f"\n• {bullet2}"
            if summary:
                result += f"\n\n{summary}"

            return result
        except Exception as e:
            print(f"Restaurant review Claude error: {e}")
            return f"Found some mentions of {name} but couldn't summarise them right now."
    except Exception as e:
        return f"Couldn't fetch reviews for {name} right now."

def is_restaurant_suggestion_request(text):
    """Detect similar restaurant suggestion request."""
    lower = text.lower()
    # Explicit suggest restaurant command
    if re.match(r"suggest restaurant", lower):
        return True
    if re.match(r"recommend (a )?restaurant", lower):
        return True
    triggers = ["similar to", "like ", "anything like", "places like",
                "restaurants like", "something like", "alternatives to", "similar places"]
    return any(t in lower for t in triggers)

def get_similar_restaurants(text):
    """Suggest similar restaurants based on cuisine/vibe — grounded in RSS results for real places."""
    try:
        import xml.etree.ElementTree as ET
        lower = text.lower()
        ref_name = None
        # Handle explicit "suggest restaurant [cuisine/area]" or "recommend restaurant [type]"
        explicit_match = re.match(r"(?:suggest|recommend)(?: a)? restaurant[s]?\s+(.*)", lower)
        if explicit_match:
            ref_name = explicit_match.group(1).strip().rstrip("?").strip()
        else:
            for trigger in ["similar to", "like ", "anything like", "places like",
                            "restaurants like", "something like", "alternatives to", "similar places to"]:
                if trigger in lower:
                    idx = lower.index(trigger) + len(trigger)
                    ref_name = text[idx:].strip().rstrip("?").strip()
                    break

        # Look up tags from saved sheet
        context = ""
        if ref_name:
            try:
                ws = restaurants_sheet()
                records = ws.get_all_records()
                for r in records:
                    if ref_name.lower() in r.get("Name", "").lower():
                        tags = r.get("Tags", "")
                        location = r.get("Location", "")
                        context = f"'{r['Name']}' is tagged as: {tags}." if tags else ""
                        break
            except Exception as e:
                print(f"Similar restaurant sheet lookup error: {e}")

        # Fetch real restaurant names from Google News RSS to ground suggestions
        real_names = []
        try:
            query = f"best restaurants similar to {ref_name or text} Singapore".replace(" ", "+")
            url = f"https://news.google.com/rss/search?q={query}&hl=en-SG&gl=SG&ceid=SG:en"
            with httpx.Client(timeout=3) as hx:
                resp = hx.get(url)
            root = ET.fromstring(resp.content)
            for item in root.findall(".//item")[:5]:
                title = item.findtext("title", "").split(" - ")[0].strip()
                if title:
                    real_names.append(title)
        except Exception as e:
            print(f"Similar restaurants RSS error: {e}")

        real_context = f"These headlines may help identify real options: {'; '.join(real_names[:3])}" if real_names else ""

        prompt = (
            f"Suggest up to 3 restaurants in Singapore similar to '{ref_name or text}'. "
            f"{context} {real_context} "
            "Only suggest real restaurants you are fully confident exist and can be found on Google Maps in Singapore right now. "
            "If you cannot confidently name any, reply with exactly: NONE "
            "For each suggestion: name, area, and one sentence on why it's similar. "
            "Format each as: 🍽 [Name] — [Area] — [Why similar]\n---"
        )
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        result = resp.content[0].text.strip()
        if result.upper().startswith("NONE"):
            return "Nothing comes to mind right now."
        return result
    except Exception as e:
        return "Couldn't generate suggestions right now — try again in a moment."

def is_restaurant_save(text):
    """Detect restaurant save intent.
    Added: 'save this restaurant', 'try this restaurant', 'this place' + save verb."""
    lower = text.lower()
    triggers = [
        "save restaurant", "add restaurant", "save this place", "add this place",
        "save this restaurant", "try this restaurant", "this restaurant to try",
        "want to try", "add to my list", "save to my list", "maps.google",
        "goo.gl/maps", "maps.app.goo", "restaurant to try", "place to try",
        "log this restaurant", "note this restaurant", "remember this place",
        "remember this restaurant"
    ]
    return any(t in lower for t in triggers)

def is_restaurant_search(text):
    """Detect restaurant search intent."""
    lower = text.lower()
    triggers = ["find a restaurant", "search restaurants", "any restaurants",
                "restaurant recommendations", "where to eat", "restaurants in",
                "show my restaurants", "my restaurant list", "saved restaurants"]
    return any(t in lower for t in triggers)

def infer_restaurant_location(name, country="Singapore"):
    """Use Claude to infer location and outlets for a restaurant name."""
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{
                "role": "user",
                "content": (
                    f"For the restaurant '{name}' in {country}, do the following:\n"
                    "1. Does it have multiple outlets? (yes/no)\n"
                    "2. If yes, list up to 4 outlets as area names only (e.g. Dempsey Hill, Jewel Changi).\n"
                    "3. If no, give the single area/neighbourhood (e.g. Dempsey Hill, Tanjong Pagar).\n\n"
                    "Return ONLY a JSON object with:\n"
                    "- multiple_outlets: boolean\n"
                    "- outlets: list of area strings (1 item if single, up to 4 if multiple)\n"
                    "Return ONLY the JSON."
                )
            }]
        )
        raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
        try:
            return json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            return {"multiple_outlets": False, "outlets": []}
    except Exception as e:
        print(f"infer_restaurant_location error: {e}")
        return {"multiple_outlets": False, "outlets": []}

def handle_save_restaurant(text, force_new=False):
    """Handle saving a restaurant from text or maps link."""
    try:
        # Check if it contains a Maps URL
        if "maps.google" in text or "goo.gl/maps" in text or "maps.app.goo" in text:
            words = text.split()
            url = next((w for w in words if "map" in w.lower() or "goo.gl" in w.lower()), "")
            parsed = lookup_restaurant_from_maps(url)
            name = parsed.get("name", "")
            location = parsed.get("location", "")
            country = parsed.get("country", "Singapore")
            tags = ""
            notes = ""

            if not name:
                return "I got the link but couldn't extract the name. Try: 'save Burnt Ends, Teck Lim Road, tag: date night'"

            if parsed.get("_needs_location") or not location:
                return f"_NEEDS_LOCATION_:{name}:{country}"

        else:
            parsed = parse_restaurant_save(text)
            name = parsed.get("name", "")
            location = parsed.get("location", "")
            country = parsed.get("country", "Singapore")
            tags = parsed.get("tags", "")
            notes = parsed.get("notes", "")

        if not name:
            return "What's the restaurant name? Try: 'save Burnt Ends'"

        # If location already provided in the message, skip inference
        if location:
            result = save_restaurant(name, location, country, tags, notes, force_new=force_new)
            if result and result.startswith("_DUPLICATE_:"):
                existing = result.split(":", 1)[1]
                return f"_DUPLICATE_RESTAURANT_:{existing}:{name}:{location}:{country}:{tags}:{notes}"
            return format_restaurant_saved(name, location, tags, notes)

        # Infer location — signal back to handler to start confirm flow
        location_data = infer_restaurant_location(name, country)
        outlets = location_data.get("outlets", [])
        multiple = location_data.get("multiple_outlets", False)
        return f"_INFER_LOCATION_:{name}:{country}:{tags}:{int(multiple)}:" + "|".join(outlets)

    except Exception as e:
        return f"❌ Couldn't save that restaurant: {str(e)}"

def handle_search_restaurants(text):
    """Handle a restaurant search request."""
    try:
        lower = text.lower()
        # Check if listing all
        if "my restaurant list" in lower or "show my restaurants" in lower or "saved restaurants" in lower:
            return list_restaurants()

        # Extract search term
        for trigger in ["restaurants in", "find a restaurant", "any restaurants",
                        "where to eat", "restaurant recommendations", "search restaurants"]:
            if trigger in lower:
                query = lower.replace(trigger, "").strip()
                if query:
                    return search_restaurants(query)
                else:
                    return list_restaurants()

        # Fallback — search by full text
        return search_restaurants(text)
    except Exception as e:
        return f"❌ Error: {str(e)}"



# =============================================================================
# STOCK MARKET ACCESS (Step 8)
# =============================================================================
# Uses Yahoo Finance via yfinance (free, 15min delay for US, near-realtime for others)
# Portfolio stored in Portfolio sheet
# Price alerts stored in memory + checked every 15 minutes

import urllib.request

# Price alerts: { "AAPL": { "condition": "below", "price": 180.0, "active": True } }
price_alerts = {}

def persist_price_alerts_to_sheet():
    """Save active price alerts to Settings sheet so they survive Railway restarts."""
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        data = json.dumps({k: v for k, v in price_alerts.items() if v.get("active")})
        for i, r in enumerate(records):
            if r.get("Key") == "price_alerts":
                sheet.update_cell(i + 2, 2, data)
                return
        sheet.append_row(["price_alerts", data])
    except Exception as e:
        print(f"persist_price_alerts_to_sheet error: {e}")

def load_price_alerts_from_sheet():
    """Load price alerts from Settings sheet on startup."""
    global price_alerts
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "price_alerts":
                raw = r.get("Value", "")
                if raw:
                    loaded = json.loads(raw)
                    price_alerts.update(loaded)
                    print(f"Restored {len(loaded)} price alert(s) from sheet")
                return
    except Exception as e:
        print(f"load_price_alerts_from_sheet error: {e}")

# Indices to track for weekly summary
MARKET_INDICES = {
    "US": {"^GSPC": "S&P 500"},
    "China": {"000001.SS": "Shanghai"},
    "India": {"^NSEI": "Nifty 50"},
}

MARKET_FLAGS_MAP = {
    "US": "🇺🇸",
    "China": "🇨🇳",
    "India": "🇮🇳",
}


# SGX and HK ticker normalisation — common names/codes mapped to Yahoo Finance format
SGX_TICKER_MAP = {
    # DBS
    "dbs": "D05.SI", "d05": "D05.SI",
    # OCBC
    "ocbc": "O39.SI", "o39": "O39.SI",
    # UOB
    "uob": "U11.SI", "u11": "U11.SI",
    # Singtel
    "singtel": "Z74.SI", "z74": "Z74.SI",
    # CapitaLand Investment
    "capitaland": "9CI.SI", "capitaland investment": "9CI.SI", "9ci": "9CI.SI", "cli": "9CI.SI",
    # Keppel
    "keppel": "BN4.SI", "bn4": "BN4.SI",
    # Wilmar
    "wilmar": "F34.SI", "f34": "F34.SI",
    # Singapore Airlines
    "sia": "C6L.SI", "singapore airlines": "C6L.SI", "c6l": "C6L.SI",
    # Jardine Matheson
    "jardine": "J36.SI",
    # Thai Bev
    "thai bev": "Y92.SI", "thaibev": "Y92.SI",
}

HK_TICKER_MAP = {
    # Tencent
    "tencent": "0700.HK", "0700": "0700.HK", "TENCENT": "0700.HK",
    # Alibaba HK
    "alibaba": "9988.HK", "9988": "9988.HK", "ALIBABA": "9988.HK",
    # Meituan
    "meituan": "3690.HK", "3690": "3690.HK", "MEITUAN": "3690.HK",
    # HSBC
    "hsbc": "0005.HK", "0005": "0005.HK", "HSBC": "0005.HK",
    # AIA
    "aia": "1299.HK", "1299": "1299.HK", "AIA": "1299.HK",
    # BYD
    "byd": "1211.HK", "1211": "1211.HK", "BYD": "1211.HK",
    # Xiaomi
    "xiaomi": "1810.HK", "1810": "1810.HK", "XIAOMI": "1810.HK",
    # JD
    "jd": "9618.HK", "9618": "9618.HK", "JD": "9618.HK",
    # NetEase
    "netease": "9999.HK", "9999": "9999.HK", "NETEASE": "9999.HK",
    # CNOOC
    "cnooc": "0883.HK", "0883": "0883.HK", "CNOOC": "0883.HK",
}

def normalise_ticker(ticker):
    """Normalise a ticker string to Yahoo Finance format.
    Handles SGX (.SI), HK (.HK), and common name lookups."""
    t = ticker.strip().lower()
    if t in SGX_TICKER_MAP:
        return SGX_TICKER_MAP[t]
    if t in HK_TICKER_MAP:
        return HK_TICKER_MAP[t]
    upper = ticker.strip().upper()
    # Check uppercase variants in maps (safety net for all-caps input)
    if upper in HK_TICKER_MAP:
        return HK_TICKER_MAP[upper]
    if upper in SGX_TICKER_MAP:
        return SGX_TICKER_MAP[upper]
    # Already has exchange suffix
    if "." in upper:
        return upper
    # Looks like an HK ticker (4-digit number)
    if re.match(r"^\d{4}$", upper):
        return f"{upper}.HK"
    return upper

def fetch_weekly_change(ticker):
    """Fetch weekly % change — previous completed calendar week (last Mon open → last Fri close)."""
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1wk&range=1mo"
        with httpx.Client(timeout=10, headers={"User-Agent": "Mozilla/5.0"}) as hx:
            resp = hx.get(url)
        data = resp.json()
        result = data["chart"]["result"][0]
        closes = result["indicators"]["quote"][0].get("close", [])
        opens = result["indicators"]["quote"][0].get("open", [])
        # Pair opens/closes, filter incomplete candles (None close = week still in progress)
        pairs = [(o, c) for o, c in zip(opens, closes) if o is not None and c is not None]
        # Take second-to-last = last fully completed week
        if len(pairs) >= 2:
            week_open, week_close = pairs[-2]
            if week_open:
                pct = (week_close - week_open) / week_open * 100
                return pct
    except Exception as e:
        print(f"fetch_weekly_change error for {ticker}: {e}")
    return None

def fetch_price(ticker):
    """Fetch current price — Yahoo Finance primary (full data), Alpha Vantage secondary."""
    ticker = normalise_ticker(ticker)

    # --- Yahoo Finance (primary) ---
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=1y"
        with httpx.Client(timeout=10, headers={"User-Agent": "Mozilla/5.0"}) as hx:
            resp = hx.get(url)
        data = resp.json()
        result = data["chart"]["result"][0]
        meta = result["meta"]
        price = meta.get("regularMarketPrice", 0)
        prev_close = meta.get("chartPreviousClose", 0)
        currency = meta.get("currency", "USD")
        name = meta.get("longName") or meta.get("shortName") or ticker
        change = price - prev_close
        change_pct = (change / prev_close * 100) if prev_close else 0
        market_state = meta.get("marketState", "")  # REGULAR, PRE, POST, CLOSED
        exchange = meta.get("exchange", "")
        full_exchange = meta.get("fullExchangeName", "")
        # 52-week range from historical closes
        closes = [c for c in result["indicators"]["quote"][0].get("close", []) if c is not None]
        week52_low = min(closes) if closes else None
        week52_high = max(closes) if closes else None
        # Exchange flag by ticker suffix
        suffix = ticker.split(".")[-1] if "." in ticker else ""
        exchange_flags = {
            "SI": "🇸🇬", "HK": "🇭🇰", "L": "🇬🇧", "AX": "🇦🇺",
            "T": "🇯🇵", "NS": "🇮🇳", "BO": "🇮🇳", "SS": "🇨🇳", "SZ": "🇨🇳",
        }
        flag = exchange_flags.get(suffix, "🇺🇸")
        return {
            "ticker": ticker,
            "name": name,
            "price": price,
            "prev_close": prev_close,
            "change": change,
            "change_pct": change_pct,
            "currency": currency,
            "market_state": market_state,
            "exchange": exchange,
            "fullExchangeName": full_exchange,
            "week52_low": week52_low,
            "week52_high": week52_high,
            "flag": flag,
        }
    except Exception as e:
        print(f"Yahoo Finance error for {ticker}: {e}")

    # --- Alpha Vantage (secondary fallback) ---
    if ALPHA_VANTAGE_API_KEY:
        try:
            url = (
                f"https://www.alphavantage.co/query"
                f"?function=GLOBAL_QUOTE&symbol={ticker}&apikey={ALPHA_VANTAGE_API_KEY}"
            )
            with httpx.Client(timeout=10) as hx:
                resp = hx.get(url)
            data = resp.json()
            quote = data.get("Global Quote", {})
            if quote.get("05. price"):
                price = float(quote["05. price"])
                prev_close = float(quote["08. previous close"])
                change = float(quote["09. change"])
                change_pct = float(quote["10. change percent"].replace("%", ""))
                av_suffix = ticker.split(".")[-1] if "." in ticker else ""
                av_exchange_flags = {
                    "SI": "🇸🇬", "HK": "🇭🇰", "L": "🇬🇧", "AX": "🇦🇺",
                    "T": "🇯🇵", "NS": "🇮🇳", "BO": "🇮🇳", "SS": "🇨🇳", "SZ": "🇨🇳",
                }
                return {
                    "ticker": ticker,
                    "name": ticker,
                    "price": price,
                    "prev_close": prev_close,
                    "change": change,
                    "change_pct": change_pct,
                    "currency": "USD",
                    "flag": av_exchange_flags.get(av_suffix, "🇺🇸"),
                }
        except Exception as e:
            print(f"Alpha Vantage fallback error for {ticker}: {e}")

    return None

# Readable source name lookup — maps domain/name fragments to display labels
SOURCE_LABELS = {
    "reuters": "Reuters", "bloomberg": "Bloomberg",
    "financial times": "FT", "ft.com": "FT",
    "cnbc": "CNBC", "straits times": "Straits Times",
    "business times": "Business Times", "nikkei": "Nikkei",
    "wall street journal": "WSJ", "wsj": "WSJ",
    "yahoo finance": "Yahoo Finance", "marketwatch": "MarketWatch",
    "seeking alpha": "Seeking Alpha", "channel news asia": "CNA",
    "cna": "CNA", "barrons": "Barron's", "fortune": "Fortune",
    "investopedia": "Investopedia", "motley fool": "Motley Fool",
    "benzinga": "Benzinga", "zacks": "Zacks",
}

# Sources too obscure to show — omit tag entirely
_OBSCURE_SOURCES = {"guruFocus", "forex.com", "indmoney", "gotrade", "traders union",
                    "cliftonlarsonallen", "simply wall st", "stockanalysis"}

def get_source_label(source_name):
    """Return a readable source label, or None if source is obscure/unknown."""
    lower = source_name.lower()
    for key, label in SOURCE_LABELS.items():
        if key in lower:
            return label
    # Check if obscure
    for obs in _OBSCURE_SOURCES:
        if obs in lower:
            return None
    return None  # Unknown — omit rather than show garbage

def _fetch_rss_headlines_for_stock(ticker, name):
    """Fetch up to 3 headlines from Google News + Yahoo Finance RSS in parallel."""
    import xml.etree.ElementTree as ET
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Best single query per source
    gn_query = f"{name} stock".replace(" ", "+")
    yf_query = f"{ticker} stock".replace(" ", "+")

    rss_sources = [
        f"https://news.google.com/rss/search?q={gn_query}&hl=en&gl=US&ceid=US:en",
        f"https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US",
    ]

    headlines = []
    sources = []
    seen = set()

    def fetch_one(url):
        try:
            with httpx.Client(timeout=3) as hx:
                resp = hx.get(url)
            root = ET.fromstring(resp.content)
            results = []
            for item in root.findall(".//item"):
                title = item.findtext("title", "")
                source = item.findtext("source", "") or title.split(" - ")[-1].strip()
                title = title.split(" - ")[0].strip()
                if title:
                    results.append((title, source))
            return results
        except Exception as e:
            print(f"RSS fetch error {url}: {e}")
            return []

    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = [ex.submit(fetch_one, url) for url in rss_sources]
        for f in as_completed(futures):
            for title, source in f.result():
                if title not in seen and len(headlines) < 4:
                    seen.add(title)
                    headlines.append(title)
                    sources.append(source)

    return headlines[:3], sources[:3]

def _generate_price_movement_summary(data):
    """Generate a factual price movement summary from price data alone — no Claude needed."""
    price = data.get("price", 0)
    change_pct = data.get("change_pct", 0)
    week52_low = data.get("week52_low")
    week52_high = data.get("week52_high")
    currency = data.get("currency", "")
    name = data.get("name", data.get("ticker", ""))

    direction = "up" if change_pct >= 0 else "down"
    sentences = []
    sentences.append(
        f"{name} is {direction} {abs(change_pct):.2f}% today, currently at {currency} {price:.2f}."
    )
    if week52_low and week52_high:
        position = (price - week52_low) / (week52_high - week52_low) * 100 if week52_high != week52_low else 50
        if position >= 75:
            range_desc = "trading near its 52-week high"
        elif position <= 25:
            range_desc = "trading near its 52-week low"
        else:
            range_desc = "trading in the middle of its 52-week range"
        sentences.append(
            f"It is {range_desc} ({currency} {week52_low:.2f} – {currency} {week52_high:.2f})."
        )
    return " ".join(sentences)

def fetch_stock_summary(ticker, name, price_data=None):
    """Fetch RSS headlines in parallel and generate a grounded summary. Falls back to price movement analysis."""
    from concurrent.futures import ThreadPoolExecutor

    headlines, sources = _fetch_rss_headlines_for_stock(ticker, name)

    # Build readable source labels — skip obscure ones
    labelled = []
    for h, s in zip(headlines, sources):
        label = get_source_label(s)
        labelled.append((h, label))  # label may be None

    usable = [(h, l) for h, l in labelled if l]  # only headlines with known sources

    if not usable and price_data:
        # No usable headlines — return pure price movement summary
        return _generate_price_movement_summary(price_data), []

    if not usable:
        return None, []

    headline_text = "\n".join(f"- {h}" for h, _ in usable[:3])
    source_context = "; ".join(f"{h[:60]} [{l}]" for h, l in usable[:3])

    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{
                "role": "user",
                "content": (
                    f"Headlines about {name} ({ticker}):\n{source_context}\n\n"
                    "Write 2-3 short factual sentences as one paragraph. "
                    "First sentence: where the stock sits in its 52-week range (near high, near low, or mid-range). "
                    "Remaining sentences: recent business developments or earnings from the headlines — factual only. "
                    "No timing advice, no buy/sell signals, no speculation, no phrases like 'investors should' or 'right time to buy'. "
                    "End with a single [SourceName] tag for the most relevant source. No source tag if no usable headlines."
                )
            }]
        )
        summary = resp.content[0].text.strip()
        return summary, [l for _, l in usable[:2]]
    except Exception as e:
        print(f"Stock summary Claude error: {e}")
        if price_data:
            return _generate_price_movement_summary(price_data), []
        return None, []

def format_price(data, summary=None):
    """Format stock price — flag, name, price, state, 52-week range, sourced summary."""
    if not data:
        return None
    flag = data.get("flag", "🌐")
    name = data.get("name", data["ticker"])
    ticker = data["ticker"]
    currency = data.get("currency", "")
    price = data.get("price", 0)
    change_pct = data.get("change_pct", 0)
    arrow = "▲" if change_pct >= 0 else "▼"
    market_state = data.get("market_state", "REGULAR")
    state_label = {"PRE": " (pre)", "POST": " (post)", "CLOSED": " (closed)", "REGULAR": ""}.get(market_state, "")

    week52_low = data.get("week52_low")
    week52_high = data.get("week52_high")
    range_line = f"52-week range: {currency} {week52_low:.2f} – {currency} {week52_high:.2f}" if week52_low and week52_high else ""

    if summary is None:
        summary, _ = fetch_stock_summary(ticker, name, price_data=data)

    lines = [f"{flag} {name} ({ticker})"]
    lines.append(f"{currency} {price:.2f} {arrow} {abs(change_pct):.2f}%{state_label}")
    if range_line:
        lines.append(range_line)
    lines.append("")
    lines.append(f"{summary if summary else _generate_price_movement_summary(data)}")

    return "\n".join(lines)

def portfolio_sheet():
    return spreadsheet.worksheet("Portfolio")

# --- Trips Sheet ---

def trips_sheet():
    return spreadsheet.worksheet("Trips")

def generate_trip_id():
    return date.today().strftime("TRIP-%Y%m%d")

def save_trip(destination, currency, check_in="", check_out="",
              hotel_name="", hotel_local_name="", hotel_address="", notes=""):
    """Write a new active trip row to Trips sheet (new schema)."""
    try:
        ws = trips_sheet()
        trip_id = generate_trip_id()
        ws.append_row([
            trip_id, destination, currency,
            check_in, check_out,
            hotel_name, hotel_local_name, hotel_address,
            notes, "active"
        ])
        return trip_id
    except Exception as e:
        print(f"save_trip error: {e}")
        return None

def close_trip(trip_id=None):
    """Mark the active trip as closed. Status is col 10."""
    try:
        ws = trips_sheet()
        records = ws.get_all_records()
        for i, row in enumerate(records, start=2):
            if row.get("Status") == "active" and (trip_id is None or row.get("Trip ID") == trip_id):
                ws.update_cell(i, 10, "closed")
                return True
    except Exception as e:
        print(f"close_trip error: {e}")
    return False

def get_active_trip():
    """Return the most recent active trip row, or None."""
    try:
        ws = trips_sheet()
        records = ws.get_all_records()
        for row in reversed(records):
            if row.get("Status") == "active":
                return row
    except Exception as e:
        print(f"get_active_trip error: {e}")
    return None

def restore_overseas_from_trips():
    """On startup — restore overseas_state from active trip if present."""
    try:
        trip = get_active_trip()
        if not trip:
            return False
        try:
            dest = trip.get("Destination", "") or ""
        except Exception:
            dest = ""
        try:
            curr = trip.get("Currency", "SGD") or "SGD"
        except Exception:
            curr = "SGD"
        if not dest or not curr or curr == "SGD":
            return False
        overseas_state["active"] = True
        overseas_state["destination"] = dest
        overseas_state["currency"] = curr
        overseas_state["currencies"] = [curr]
        overseas_state["trip_destinations"] = [dest]
        overseas_state["trip_start"] = trip.get("Check In", "")
        print(f"✅ Restored overseas mode: {dest} ({curr})")
        return True
    except Exception as e:
        print(f"restore_overseas_from_trips error: {e}")
    return False

def get_trip_history(n=5):
    """Return last n trips from Trips sheet."""
    try:
        ws = trips_sheet()
        records = ws.get_all_records()
        return list(reversed(records[-n:])) if records else []
    except Exception as e:
        print(f"get_trip_history error: {e}")
        return []

def format_trip_history():
    trips = get_trip_history(10)
    if not trips:
        return "No trips logged yet."
    lines = ["*Trip History*\n"]
    for t in trips:
        status = "✈️ Active" if t.get("Status") == "active" else "✅ Done"
        dest = t.get("Destination", "—")
        curr = t.get("Currency", "")
        check_in = t.get("Check In", "")
        check_out = t.get("Check Out", "")
        hotel = t.get("Hotel Name", "")
        line = f"{status} {dest} ({curr})"
        if check_in:
            line += f"\n{check_in}"
            if check_out:
                line += f" → {check_out}"
        if hotel:
            line += f" | {hotel}"
        lines.append(line)
    return "\n\n".join(lines)


def log_portfolio_buy(ticker, quantity, price, buy_date=None):
    """Log a stock purchase to Portfolio sheet."""
    sheet = portfolio_sheet()
    today = buy_date or date.today().strftime("%d/%m/%Y")
    sheet.append_row([ticker.upper(), str(quantity), str(price), today, ""])

def get_portfolio_holdings():
    """Get current holdings with average cost."""
    try:
        sheet = portfolio_sheet()
        records = sheet.get_all_records()
        holdings = {}
        for r in records:
            ticker = r.get("Stock", "").upper()
            if not ticker:
                continue
            try:
                qty = float(r.get("Quantity", 0))
                price = float(r.get("Buy Price", 0))
            except (ValueError, TypeError):
                print(f"get_portfolio_holdings: bad qty/price for {ticker}, skipping row")
                continue
            if ticker not in holdings:
                holdings[ticker] = {"total_qty": 0, "total_cost": 0}
            holdings[ticker]["total_qty"] += qty
            holdings[ticker]["total_cost"] += qty * price
        # Calculate averages
        result = {}
        for ticker, h in holdings.items():
            if h["total_qty"] > 0:
                result[ticker] = {
                    "qty": h["total_qty"],
                    "avg_cost": h["total_cost"] / h["total_qty"]
                }
        return result
    except Exception as e:
        print(f"Error getting portfolio: {e}")
        return {}

def get_portfolio_performance():
    """Get portfolio performance — current vs average cost."""
    holdings = get_portfolio_holdings()
    if not holdings:
        return "No holdings logged yet."

    lines = ["Portfolio:\n"]
    total_cost = 0
    total_value = 0

    for ticker, h in holdings.items():
        data = fetch_price(ticker)
        avg = h["avg_cost"]
        qty = h["qty"]
        cost = avg * qty

        if data:
            current = data["price"]
            value = current * qty
            pnl = value - cost
            pnl_pct = (pnl / cost * 100) if cost else 0
            sign = "+" if pnl >= 0 else ""
            flag = "✅" if pnl >= 0 else "⚠️"
            lines.append(
                f"{flag} {ticker}: {qty:.0f} shares @ avg {data['currency']} {avg:.2f} "
                f"| now {data['currency']} {current:.2f} | {sign}{pnl_pct:.1f}% ({sign}{data['currency']} {pnl:.2f})"
            )
            total_cost += cost
            total_value += value
        else:
            lines.append(f"• {ticker}: {qty:.0f} shares @ avg {avg:.2f} (price unavailable)")
            total_cost += cost

    if total_cost > 0:
        total_pnl = total_value - total_cost
        total_pct = (total_pnl / total_cost * 100)
        sign = "+" if total_pnl >= 0 else ""
        lines.append(f"\nTotal P&L: {sign}{total_pct:.1f}% ({sign}${total_pnl:.2f})")

    return "\n".join(lines)

def get_portfolio_rows():
    """Return list of (row_index, record) for all portfolio entries."""
    try:
        ws = portfolio_sheet()
        records = ws.get_all_records()
        return [(i + 2, r) for i, r in enumerate(records)]  # row 1 = header
    except Exception as e:
        print(f"get_portfolio_rows error: {e}")
        return []

def format_portfolio_delete_list(rows):
    """Format numbered list of portfolio holdings for delete selection."""
    if not rows:
        return "No holdings in portfolio."
    lines = ["Which holding do you want to remove?\n"]
    for i, (_, r) in enumerate(rows, 1):
        ticker = r.get("Stock", "?")
        qty = r.get("Quantity", "?")
        price = r.get("Buy Price", "?")
        buy_date = r.get("Buy Date", "")
        lines.append(f"{i}. {ticker} — {qty} shares @ ${price} ({buy_date})")
    lines.append("\nReply with a number, 'search [ticker]', or 'cancel'.")
    return "\n".join(lines)

def delete_portfolio_row(sheet_row):
    """Delete a portfolio row by sheet row index."""
    try:
        ws = portfolio_sheet()
        all_vals = ws.get_all_values()
        if sheet_row - 1 < len(all_vals):
            row_data = all_vals[sheet_row - 1]
            ticker = row_data[0] if row_data else "?"
        ws.delete_rows(sheet_row)
        return f"Removed {ticker} from portfolio."
    except Exception as e:
        return f"Couldn't remove that holding: {str(e)}"

def search_portfolio_by_ticker(query):
    """Search portfolio holdings by ticker substring."""
    rows = get_portfolio_rows()
    q = query.upper()
    return [(row_idx, r) for row_idx, r in rows if q in r.get("Stock", "").upper()]


def set_price_alert(ticker, condition, price):
    """Set a price alert for a ticker."""
    ticker = ticker.upper()
    price_alerts[ticker] = {"condition": condition, "price": price, "active": True}
    persist_price_alerts_to_sheet()

def parse_stock_request(text):
    """Use Claude to parse a stock-related request."""
    prompt = (
        f"Parse this stock market request: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- intent: string (one of: price_check, set_alert, portfolio_add, portfolio_view, "
        f"stock_suggest, market_summary, price_alert_check)\n"
        f"- ticker: string (stock ticker symbol, uppercase, or empty)\n"
        f"- quantity: number (shares, or 0)\n"
        f"- price: number (price per share, or 0)\n"
        f"- alert_condition: string (above or below, or empty)\n"
        f"- alert_price: number (alert trigger price, or 0)\n"
        f"- criteria: string (for stock suggestions — describe what user wants)\n\n"
        f"Return ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json","").replace("```","").strip()
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, ValueError) as e:
        print(f"parse_stock_request JSON error: {e} | raw: {raw[:100]}")
        return None
    if parsed.get("ticker"):
        parsed["ticker"] = normalise_ticker(parsed["ticker"])
    return parsed

def suggest_stocks(criteria):
    """Suggest 3 stocks based on described criteria."""
    prompt = (
        f"Suggest 3 stocks based on this criteria: '{criteria}'\n\n"
        f"For each stock provide a qualitative summary. Flag concerns with a warning, all clear with a checkmark.\n\n"
        f"Format your response exactly like this for each stock (with a divider line between each):\n\n"
        f"TICKER — Company Name\n"
        f"[checkmark or warning] [2-3 sentence qualitative summary. No numbers unless asked.]\n\n"
        f"---\n\n"
        f"Keep it concise and honest. Flag anything concerning."
    )
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text.strip()

async def check_price_alerts(app):
    """Check price alerts every 15 minutes."""
    try:
        for ticker, alert in list(price_alerts.items()):
            if not alert.get("active"):
                continue
            data = fetch_price(ticker)
            if not data:
                continue
            current = data["price"]
            condition = alert["condition"]
            trigger = alert["price"]

            triggered = (condition == "below" and current <= trigger) or                        (condition == "above" and current >= trigger)

            if triggered:
                msg = (
                    f"🔔 Price alert: {ticker} is now {data['currency']} {current:.2f} "
                    f"({condition} your target of {data['currency']} {trigger:.2f})"
                )
                await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
                price_alerts[ticker]["active"] = False  # deactivate after firing
                persist_price_alerts_to_sheet()

    except Exception as e:
        print(f"Error checking price alerts: {e}")

async def send_weekly_market_summary(app):
    """Monday 8am — send US, China, India market summary using agreed qualitative format."""
    try:
        msg = get_market_summary_now()
        await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg, parse_mode="Markdown")
        # Persist sent date to Settings so restart doesn't re-prompt
        try:
            sheet = get_sheet("Settings")
            if sheet:
                records = sheet.get_all_records()
                today_str = date.today().isoformat()
                for i, r in enumerate(records):
                    if r.get("Key") == "market_summary_last_sent":
                        sheet.update_cell(i + 2, 2, today_str)
                        break
                else:
                    sheet.append_row(["market_summary_last_sent", today_str])
        except Exception as e:
            print(f"market_summary_last_sent persist error: {e}")
    except Exception as e:
        print(f"Error sending weekly market summary: {e}")

def is_stock_request(text):
    """Detect stock market related requests using explicit trigger phrases only.
    Exclusions guard against collisions with reminder ('check my reminders') and expense ('bought coffee').
    Share-purchase patterns ('bought 10 shares of X') beat expense routing."""
    lower = text.lower()

    # Hard exclusions — these are never stock requests
    exclusions = ["reminders", "reminder", "my bill", "credit card bill"]
    if any(e in lower for e in exclusions):
        return False

    # Share purchase/sale — always stock, even if 'bought'/'sold' triggers expense
    share_patterns = [
        r"bought \d+ shares", r"sold \d+ shares", r"shares of [a-z]",
        r"shares @ ", r"shares at \$", r"\d+ shares"
    ]
    if any(re.search(p, lower) for p in share_patterns):
        return True

    explicit_triggers = [
        "pull up ", "look into ", "price of ",
        "alert me if", "alert if ", "add to portfolio",
        "suggest stocks", "stock ideas",
        "stock ", "ticker ", "p&l", "holdings",
        "how is the market", "market today",
        "weekly market", "how are markets", "portfolio performance"
    ]
    if any(t in lower for t in explicit_triggers):
        return True
    # "market summary" only fires if not a reminder request (e.g. "notify me about market summary")
    reminder_prefixes = ["remind me", "notify me", "ping me", "alert me when", "alert me to", "send me"]
    if "market summary" in lower and not any(lower.startswith(p) or lower[:25].startswith(p) for p in reminder_prefixes):
        return True

    # "check" only fires if not about reminders/bills AND not a reminder-style request
    reminder_prefix = any(lower.startswith(p) or p + " " in lower[:20]
                          for p in ["ping me", "notify me", "remind me", "alert me when", "alert me to"])
    if "check " in lower and not any(e in lower for e in ["reminders", "reminder", "bill"]) and not reminder_prefix:
        return True

    # "bought"/"sold" only fire when combined with known stock context words
    if ("bought " in lower or "sold " in lower) and any(
        w in lower for w in ["shares", "stock", "equity", "position", "portfolio", "aapl", "tsla"]
    ):
        return True

    # "what's X at" / "what is X at" — natural price queries
    if re.search(r"what'?s\s+\S+\s+at\b", lower):
        return True
    if re.search(r"what is\s+\S+\s+at\b", lower):
        return True
    if re.search(r"how much is\s+\S+\s+(trading|at|worth)\b", lower):
        return True

    ticker_match = re.search(r'\b[A-Z]{2,5}\b', text)
    if ticker_match and any(w in lower for w in ["doing", "worth", "performing", "price", "target", "outlook"]):
        return True
    return False

def handle_stock_request(text):
    """Route a stock request to the right handler."""
    try:
        parsed = parse_stock_request(text)
        if not parsed:
            return "Couldn't parse that stock request — try again with a ticker or clearer intent."
        intent = parsed.get("intent", "")
        ticker = parsed.get("ticker", "").upper()

        if intent == "price_check" and ticker:
            from concurrent.futures import ThreadPoolExecutor
            # Run price fetch and RSS fetch in parallel
            with ThreadPoolExecutor(max_workers=2) as ex:
                price_future = ex.submit(fetch_price, ticker)
                rss_future = ex.submit(_fetch_rss_headlines_for_stock, ticker, ticker)
                data = price_future.result()
                rss_headlines, rss_sources = rss_future.result()

            if data:
                # OTC filter — reject no-suffix tickers resolving to OTC markets
                exchange = data.get("exchange", "") or data.get("fullExchangeName", "")
                if "OTC" in exchange.upper() and "." not in normalise_ticker(ticker):
                    return (
                        f"⚠️ {ticker} appears to be an OTC/pink sheet listing — data may be unreliable.\n"
                        f"Try the primary exchange listing instead (e.g. 0700.HK for Tencent, 9988.HK for Alibaba)."
                    )
                # Pass pre-fetched headlines into summary to avoid second RSS call
                name = data.get("name", ticker)
                labelled = [(h, get_source_label(s)) for h, s in zip(rss_headlines, rss_sources)]
                usable = [(h, l) for h, l in labelled if l]
                if usable:
                    source_context = "; ".join(f"{h[:60]} [{l}]" for h, l in usable[:3])
                    try:
                        resp = client.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=120,
                            messages=[{
                                "role": "user",
                                "content": (
                                    f"Headlines about {name} ({ticker}):\n{source_context}\n\n"
                                    "Write 2-3 short factual sentences as one paragraph. "
                                    "First sentence: where the stock sits in its 52-week range (near high, near low, or mid-range). "
                                    "Remaining sentences: recent business developments or earnings from the headlines — factual only. "
                                    "No timing advice, no buy/sell signals, no speculation, no phrases like 'investors should' or 'right time to buy'. "
                                    "End with a single [SourceName] tag for the most relevant source. No source tag if no usable headlines."
                                )
                            }]
                        )
                        summary = resp.content[0].text.strip()
                    except Exception:
                        summary = _generate_price_movement_summary(data)
                else:
                    summary = _generate_price_movement_summary(data)

                # Pass pre-fetched summary into format_price to avoid duplicate output
                return format_price(data, summary=summary)

            if ALPHA_VANTAGE_API_KEY:
                return f"Couldn't fetch {ticker} — Alpha Vantage and Yahoo Finance both failed. The ticker may be wrong, or try again in a moment."
            return f"Couldn't fetch {ticker}. Check the ticker and try again."

        elif intent == "set_alert" and ticker:
            condition = parsed.get("alert_condition", "below")
            alert_price = parsed.get("alert_price", 0)
            if not alert_price:
                return "What price should I alert you at? Try: 'alert me if AAPL drops below $180'"
            set_price_alert(ticker, condition, alert_price)
            return f"Alert set — I'll let you know when {ticker} goes {condition} ${alert_price:.2f}."

        elif intent == "portfolio_add" and ticker:
            qty = parsed.get("quantity", 0)
            price = parsed.get("price", 0)
            if not qty or not price:
                return "I need the quantity and price. Try: 'bought 100 AAPL at $180'"
            log_portfolio_buy(ticker, qty, price)
            return f"Logged — {qty:.0f} {ticker} @ ${price:.2f}."

        elif intent == "portfolio_view":
            return get_portfolio_performance()

        elif intent == "stock_suggest":
            criteria = parsed.get("criteria", text)
            return suggest_stocks(criteria)

        elif intent == "market_summary":
            # Trigger on-demand summary
            return "Pulling the latest market data, give me a sec..."

        else:
            # Fallback — try price check if ticker found
            if ticker:
                data = fetch_price(ticker)
                if data:
                    return format_price(data)
                return f"Couldn't find data for {ticker} — check the ticker and try again."
            return "Couldn't work out what stock you're asking about. Try 'price of AAPL' or 'what's DBS at'."

    except Exception as e:
        print(f"handle_stock_request error: {e}")
        return f"Something went wrong with that stock request — try again in a moment."


# --- Natural Language CRM Update Detection ---

def detect_crm_natural_update(text):
    """
    Detect natural language CRM field updates like:
    "Sarah's email is sarah@gmail.com"
    "update James's address to 123 Orchard Road"
    "James referred Sarah"
    Returns (action, name, field, value) or None.
    """
    lower = text.lower()

    # Referral: "X referred Y"
    ref_match = re.search(
        r"([a-z][a-z\s']+?)\s+referred\s+([a-z][a-z\s']+?)(?:\s+to\s+me|\s+to\s+us)?\.?$",
        lower
    )
    if ref_match:
        referrer = ref_match.group(1).strip().title()
        referred = ref_match.group(2).strip().title()
        return ("referral", referrer, referred, None)

    # Pattern: "[Name]'s [field] is [value]"
    is_match = re.search(
        r"([a-z][a-z\s']+?)'s\s+(email|address|alias|birthday|relationship|context|notes)\s+is\s+(.+)",
        lower
    )
    if is_match:
        name = is_match.group(1).strip().title()
        field = is_match.group(2).strip()
        value = is_match.group(3).strip().rstrip(".")
        return ("update", name, field, value)

    # Pattern: "update [Name]'s [field] to [value]"
    update_match = re.search(
        r"update\s+([a-z][a-z\s']+?)'s\s+(email|address|alias|birthday|relationship|context|notes)\s+to\s+(.+)",
        lower
    )
    if update_match:
        name = update_match.group(1).strip().title()
        field = update_match.group(2).strip()
        value = update_match.group(3).strip().rstrip(".")
        return ("update", name, field, value)

    # Pattern: "[Name]'s email" / "what's [Name]'s email/address"
    ask_private = re.search(
        r"(?:what'?s?\s+)?([a-z][a-z\s']+?)'s\s+(email|address)",
        lower
    )
    if ask_private and any(w in lower for w in ["what", "show", "tell", "give"]):
        name = ask_private.group(1).strip().title()
        field = ask_private.group(2).strip()
        return ("show_private", name, field, None)

    return None


# Clickbait/speculative headline patterns to filter out
_HEADLINE_REJECT = [
    "will it", "will they", "should you", "best stocks", "stocks to watch",
    "to buy and watch", "to watch:", "opening:", "opening bell", "preview:",
    "what to expect", "top picks", "analyst picks", "should i", "is it time",
    "here's what", "what you need", "everything you need",
]

def fetch_market_rss_headlines(market_name):
    """Fetch top 3 informative market headlines from Google News RSS. Filters clickbait/speculative titles."""
    try:
        queries = {
            "US": ["US stock market today", "Wall Street S&P 500", "Federal Reserve economy"],
            "China": ["China stock market", "Shanghai economy", "China trade economy"],
            "India": ["India Nifty stock market", "RBI India economy", "BSE Sensex"],
        }
        key = None
        for k in queries:
            if k in market_name:
                key = k
                break
        if not key:
            return []

        import xml.etree.ElementTree as ET
        headlines = []
        for q_text in queries[key]:
            if len(headlines) >= 3:
                break
            q = q_text.replace(" ", "+")
            url = f"https://news.google.com/rss/search?q={q}&hl=en-SG&gl=SG&ceid=SG:en"
            try:
                with httpx.Client(timeout=5) as hx:
                    resp = hx.get(url)
                root = ET.fromstring(resp.content)
                for item in root.findall(".//item"):
                    if len(headlines) >= 3:
                        break
                    title = item.findtext("title", "").split(" - ")[0].strip()
                    if not title or title in headlines:
                        continue
                    # Filter clickbait/speculative headlines
                    lower_title = title.lower()
                    if any(pattern in lower_title for pattern in _HEADLINE_REJECT):
                        continue
                    headlines.append(title)
            except Exception as e:
                print(f"RSS fetch error for {market_name} query '{q_text}': {e}")
                continue
        return headlines[:3]
    except Exception as e:
        print(f"fetch_market_rss_headlines error for {market_name}: {e}")
        return []

def get_market_summary_now():
    """Generate market summary — one index per market, weekly %, filtered bullets, 2-3 sentence overall."""
    try:
        market_data_blocks = []
        all_headlines = []

        for market, indices in MARKET_INDICES.items():
            flag = MARKET_FLAGS_MAP.get(market, "🌐")
            ticker, index_name = list(indices.items())[0]

            # Weekly % change
            weekly_pct = fetch_weekly_change(ticker)
            if weekly_pct is None:
                # Fallback to daily
                price_data = fetch_price(ticker)
                weekly_pct = price_data["change_pct"] if price_data else 0
                arrow = "▲" if weekly_pct >= 0 else "▼"
                pct_str = f"Day {arrow} {abs(weekly_pct):.1f}%"
            else:
                arrow = "▲" if weekly_pct >= 0 else "▼"
                pct_str = f"Week {arrow} {abs(weekly_pct):.1f}%"

            headlines = fetch_market_rss_headlines(market)
            all_headlines.extend(headlines)

            block = {
                "market": market,
                "flag": flag,
                "index_name": index_name,
                "pct_str": pct_str,
                "weekly_pct": weekly_pct,
                "headlines": headlines,
            }
            market_data_blocks.append(block)

        # Build data context for Claude — bullet points + overall summary
        data_summary = ""
        for b in market_data_blocks:
            data_summary += f"\n{b['flag']} {b['market']} — {b['index_name']} {b['pct_str']}\n"
            if b["headlines"]:
                data_summary += "Headlines: " + "; ".join(b["headlines"]) + "\n"

        # Claude — generate bullet points (1 sentence each) + 2-3 sentence overall summary
        try:
            claude_resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=500,
                messages=[{
                    "role": "user",
                    "content": (
                        f"Here is this week's market data:\n{data_summary}\n\n"
                        "For each market (US, China, India), write exactly 3 bullet points. "
                        "Each bullet = 1 short informative sentence grounded in the headlines. "
                        "No speculation, no price predictions, no clickbait. "
                        "Then write a 2-3 sentence overall summary of the week across all three markets. "
                        "Format exactly as:\n"
                        "US_BULLETS:\n• ...\n• ...\n• ...\n"
                        "CHINA_BULLETS:\n• ...\n• ...\n• ...\n"
                        "INDIA_BULLETS:\n• ...\n• ...\n• ...\n"
                        "OVERALL:\n[2-3 sentences]"
                    )
                }]
            )
            raw = claude_resp.content[0].text.strip()
        except Exception as e:
            print(f"Market summary Claude error: {e}")
            raw = ""

        # Parse Claude response
        def extract_section(text, key):
            """Extract bullet points after a section key — robust to minor Claude formatting variations."""
            pattern = re.compile(re.escape(key) + r"(.*?)(?=\n[A-Z_]+:|$)", re.DOTALL | re.IGNORECASE)
            m = pattern.search(text)
            if not m:
                return []
            chunk = m.group(1)
            bullets = [l.strip() for l in chunk.split("\n") if l.strip().startswith("•")][:3]
            return bullets

        def extract_overall(text):
            pattern = re.compile(r"OVERALL:(.*?)$", re.DOTALL | re.IGNORECASE)
            m = pattern.search(text)
            if not m:
                return ""
            return m.group(1).strip()

        us_bullets = extract_section(raw, "US_BULLETS:")
        china_bullets = extract_section(raw, "CHINA_BULLETS:")
        india_bullets = extract_section(raw, "INDIA_BULLETS:")
        overall = extract_overall(raw)

        bullet_map = {"US": us_bullets, "China": china_bullets, "India": india_bullets}

        # Build final message
        lines = [f"📊 *Market Summary* — {date.today().strftime('%d %b %Y')}\n"]
        for b in market_data_blocks:
            header = f"{b['flag']} *{b['market']} — {b['index_name']} ({b['pct_str']})*"
            bullets = bullet_map.get(b["market"], [])
            section = header
            if bullets:
                for bullet in bullets:
                    section += f"\n{bullet}"
            else:
                section += "\n• Market data unavailable"
            lines.append(section)

        if overall:
            lines.append(f"\n{overall}")

        return "\n\n".join(lines)

    except Exception as e:
        return f"❌ Couldn't pull market data right now ({type(e).__name__}: {str(e)[:80]})"

async def handle_statement_upload(file_bytes, fname, user_id, update):
    """Parse a bank statement CSV/XLSX and reconcile against logged expenses."""
    try:
        await update.message.reply_text("Got the statement, give me a sec to go through it...")
        # Parse statement rows
        statement_rows = []
        if fname.lower().endswith(".csv"):
            import csv
            text_data = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text_data))
            for row in reader:
                statement_rows.append(dict(row))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                statement_rows.append(dict(zip(headers, row)))

        if not statement_rows:
            await update.message.reply_text("❌ Couldn't read any rows from that file.")
            return

        # Use Claude to normalise statement columns
        sample = json.dumps(statement_rows[:3], default=str)
        norm_prompt = (
            f"Given these bank statement rows: {sample}\n\n"
            f"Return ONLY a JSON object mapping these keys to the actual column names in the data:\n"
            f"{{\"date\": \"col\", \"description\": \"col\", \"amount\": \"col\"}}\n"
            f"If a column doesn't exist, use null."
        )
        norm_resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=100,
            messages=[{"role": "user", "content": norm_prompt}]
        )
        col_map = json.loads(norm_resp.content[0].text.strip().replace("```json","").replace("```","").strip())
        date_col = col_map.get("date")
        desc_col = col_map.get("description")
        amt_col = col_map.get("amount")

        if not all([date_col, desc_col, amt_col]):
            await update.message.reply_text("❌ Couldn't identify date/description/amount columns. Try a CSV with clear headers.")
            return

        # Load logged expenses
        sheet = expenses_sheet()
        logged = sheet.get_all_records()

        missing = []
        corrections = []

        for srow in statement_rows:
            raw_date = str(srow.get(date_col, "")).strip()
            raw_desc = str(srow.get(desc_col, "")).strip()
            raw_amt = str(srow.get(amt_col, "")).strip().replace(",", "")
            if not raw_date or not raw_amt:
                continue
            try:
                stmt_amount = abs(float(raw_amt))
            except ValueError:
                continue

            # Normalise date
            stmt_date = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %b %y"):
                try:
                    stmt_date = datetime.strptime(raw_date, fmt).strftime("%d/%m/%Y")
                    break
                except Exception:
                    continue
            if not stmt_date:
                continue

            # Match against logged expenses — date + amount
            matched_row = None
            matched_idx = None
            for i, r in enumerate(logged):
                logged_date = r.get("Date", "")
                logged_sgd = float(r.get("SGD Amount") or r.get("Amount") or 0)
                if logged_date == stmt_date and abs(logged_sgd - stmt_amount) < 0.02:
                    matched_row = r
                    matched_idx = i
                    break

            if matched_row is None:
                # Check if ambiguous gateway — skip if can't resolve
                missing.append(f"{stmt_date} | {raw_desc[:40]} | SGD ${stmt_amount:.2f}")
            else:
                # Check for amount discrepancy
                logged_sgd = float(matched_row.get("SGD Amount") or matched_row.get("Amount") or 0)
                if abs(logged_sgd - stmt_amount) > 0.01:
                    # Correct the SGD amount in sheet — row is matched_idx + 2 (1-indexed + header)
                    all_values = sheet.get_all_values()
                    headers_row = all_values[0]
                    sgd_col_idx = headers_row.index("SGD Amount") + 1 if "SGD Amount" in headers_row else None
                    if sgd_col_idx:
                        sheet.update_cell(matched_idx + 2, sgd_col_idx, stmt_amount)
                    corrections.append(
                        f"{matched_row.get('Merchant', raw_desc[:20])} {stmt_date}: "
                        f"${logged_sgd:.2f} → ${stmt_amount:.2f}"
                    )

        lines = ["Reconciliation done ✅"]
        if corrections:
            lines.append(f"\nCorrected {len(corrections)} amount(s):")
            lines.extend(f"  {c}" for c in corrections)
        if missing:
            lines.append(f"\n{len(missing)} unmatched statement item(s) — couldn't find in your logs:")
            lines.extend(f"  {m}" for m in missing[:10])
            if len(missing) > 10:
                lines.append(f"  ...and {len(missing) - 10} more")
            lines.append("\nReply 'log [expense]' to log the first unmatched item, 'skip' to go through them, or 'done' to close.")
            recon_sessions[user_id] = {"step": "review", "unmatched": missing, "index": 0}
        if not corrections and not missing:
            lines.append("Everything matches up.")

        await update.message.reply_text("\n".join(lines))

    except Exception as e:
        print(f"handle_statement_upload error: {e}")
        await update.message.reply_text(f"Something went wrong parsing the statement: {str(e)}")



# --- Em System Prompt Builder ---


def build_system_prompt():
    """Build Em's system prompt, incorporating em_profile preferences."""
    profile_notes = ""
    if em_profile:
        forbidden = ", ".join(em_profile.get("forbidden_phrases", []))
        profile_notes = f"\nForbidden phrases (never use): {forbidden}" if forbidden else ""

    # Overseas + expense context
    overseas_context = ""
    if overseas_state.get("active"):
        dest = overseas_state.get("destination", "")
        curr = overseas_state.get("currency", "SGD")
        trip_start = overseas_state.get("trip_start", "")
        currencies = overseas_state.get("currencies", [])
        curr_list = ", ".join(currencies) if currencies else curr
        overseas_context = (
            f"\n\n## Overseas Mode\n"
            f"Currently active. Destination: {dest}. Currency: {curr}.\n"
            f"Currencies used this trip: {curr_list}.\n"
            f"Trip started: {trip_start}.\n"
            f"Expenses entered without a currency are assumed to be in {curr}.\n"
            f"All expenses are converted to SGD using a cached rate refreshed twice daily (8am and 8pm).\n"
            f"You can answer questions about trip spend, currency, and expense logging directly."
        )

    expense_context = (
        "\n\n## Expense Tracking\n"
        "Em tracks expenses to Google Sheets. "
        f"Categories: {', '.join(EXPENSE_CATEGORIES)}. "
        f"Cards: {', '.join(EXPENSE_CARDS)}. "
        "Known merchants are remembered — category and card are auto-filled for repeat merchants. "
        "Foreign currency expenses show SGD amount with original currency in brackets. "
        "Receipts can be attached as photo with caption."
    )

    return (
        "# Em — Your Personal Assistant\n\n"
        "## Core Identity\n"
        "You're Em — a smart, focused personal assistant with a casual, warm vibe. "
        "You keep things real and get stuff done without the corporate robot speak.\n\n"
        "## Communication Style\n"
        "- Natural and conversational — light slang like 'got it', 'sure thing', 'on it', 'no worries', 'lemme check that', 'all good'\n"
        "- Clean and simple — never over the top or trying too hard\n"
        "- Helpful and focused — you're here to make life easier, not to chat\n"
        "- Never say 'cool cool'\n"
        "- Never use the shaka emoji\n"
        "- Always capitalise the first letter of each sentence\n"
        "- NEVER use dashes or hyphens in conversational replies under any circumstances. Write in natural flowing sentences instead.\n"
        "- Only use dashes when displaying CRM contact info in the required format.\n"
        "- Each piece of information on its own line.\n"
        "- No unnecessary prompting or nudging.\n"
        f"{profile_notes}\n\n"
        "## Greetings & Sign-offs\n"
        "- Mix it up — never sound repetitive\n"
        "- Options: 'hey', 'yo', 'alright', 'aite', 'sup', or just dive straight in\n"
        "- Vary your closings naturally too\n\n"
        "## Emojis\n"
        "- Use sparingly — only when they feel natural\n"
        "- Don't overdo it\n\n"
        "## Response Length\n"
        "- Concise and to the point by default\n"
        "- Elaborate only when asked\n\n"
        "## CRM Display Format\n"
        "When displaying contact info, always use this exact format:\n\n"
        "[Full Name]\n"
        "- Relationship: [relationship]\n"
        "- Context: [how you know them]\n"
        "- Birthday: [DD MMM YYYY (age N)]\n"
        "- Notes:\n"
        "  - [note 1]\n"
        "  - [note 2]\n\n"
        "_Last updated: DD MMM YYYY_\n\n"
        "Email and Address are private fields — never show them unless the user specifically asks.\n"
        "Age is calculated on the fly from Birthday — never store or display a static age.\n"
        "Always separate notes into individual bullet points. Never dump them in one line.\n\n"
        "## ABSOLUTE RULE — CRM Data\n"
        "You NEVER invent, fabricate, guess, or generate contact information. "
        "If asked about a person and no real data was retrieved from the sheet, "
        "say you don't have them in the CRM. Never produce a formatted contact card "
        "unless the data came directly from a sheet lookup in this same message. "
        "This rule overrides everything else — no exceptions.\n\n"
        "- Sound stiff or corporate\n"
        "- Act like a typical AI assistant\n"
        "- Make small talk for the sake of it\n"
        "- Get repetitive with phrases or greetings"
    ) + overseas_context + expense_context

# --- Safe message sender — chunks at 4096 char Telegram limit ---
async def send_safe(target, text, parse_mode=None):
    """Send a message, splitting into chunks if over Telegram's 4096 char limit."""
    MAX = 4096
    if len(text) <= MAX:
        try:
            await target.reply_text(text, parse_mode=parse_mode)
        except Exception:
            await target.reply_text(text)
        return
    # Split at newlines to avoid breaking mid-sentence
    chunks = []
    current = ""
    for line in text.split("\n"):
        if len(current) + len(line) + 1 > MAX:
            if current:
                chunks.append(current)
            current = line
        else:
            current = current + "\n" + line if current else line
    if current:
        chunks.append(current)
    for chunk in chunks:
        try:
            await target.reply_text(chunk, parse_mode=parse_mode)
        except Exception:
            await target.reply_text(chunk)

# --- Session Interrupt Guard ---

def get_active_session_label(user_id):
    """Return a human-readable label if user is mid-session, else None."""
    if user_id in receipt_confirm_sessions:
        return "expense confirmation"
    if user_id in expense_sessions:
        return "expense entry"
    if user_id in meeting_sessions:
        return "meeting recap"
    if user_id in edit_sessions:
        return "contact edit"
    if user_id in confirm_sessions:
        action = confirm_sessions[user_id].get("action", "")
        labels = {
            "delete_contact": "contact deletion",
            "delete_bill": "bill deletion",
            "delete_restaurant": "restaurant deletion",
            "delete_event": "event deletion",
            "rename_category": "category rename",
        }
        return labels.get(action, "confirmation")
    if user_id in delete_sessions:
        return "expense deletion"
    if user_id in portfolio_delete_sessions:
        return "portfolio deletion"
    if user_id in excel_import_sessions:
        return "contact import"
    if user_id in pending_restaurant_saves:
        return "restaurant save"
    if user_id in pending_contact_saves:
        return "contact save"
    if user_id in todo_disambig_sessions:
        return "todo action"

    return None

# Words that are always session replies, never new intents
_SESSION_REPLY_TOKENS = {
    "yes", "y", "no", "n", "cancel", "skip", "done", "update", "new",
    "confirm", "ok", "okay", "sure", "nope", "yep", "yup"
}

def looks_like_new_intent(text):
    """Return True if the message looks like a fresh command, not a session reply."""
    lower = text.strip().lower()

    # Single-word session replies — never a new intent
    if lower in _SESSION_REPLY_TOKENS:
        return False

    # Bare digit(s) — session pick reply
    if re.match(r"^\d+$", lower.strip()):
        return False

    # Short single-word or two-word inputs are likely session replies
    words = lower.split()
    if len(words) <= 2:
        return False

    # Known intent signals — starts with a strong command verb or pattern
    new_intent_triggers = [
        "remind me", "set a reminder", "add reminder",
        "spent ", "paid ", "bought ", "expense ",
        "save ", "add contact", "find ", "pull up ",
        "meeting recap", "start a recap",
        "overseas", "i'm in ", "flying to",
        "what's ", "how is ", "market summary",
        "alert me", "price of ", "check ",
        "bill ", "my ", "schedule ", "book ",
        "todo ", "add to my list", "remind",
        "cancel reminder", "delete ",
    ]
    if any(lower.startswith(t) for t in new_intent_triggers):
        return True

    return False

# --- Message Handler ---
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != YOUR_CHAT_ID:
        return

    try:
        await _handle_message_inner(update, context, user_id)
    except Exception as e:
        import traceback
        err_type = type(e).__name__
        err_msg = str(e)[:120]
        tb_lines = traceback.format_exc().splitlines()
        # Find the most relevant line (last line with our code)
        location = next((l.strip() for l in reversed(tb_lines) if "bot.py" in l), "")
        detail = f"{err_type}: {err_msg}"
        if location:
            detail += f"\n({location})"
        print(f"UNHANDLED handle_message error: {traceback.format_exc()}")
        try:
            await update.message.reply_text(f"❌ Something went wrong: {detail}")
        except Exception:
            pass

async def _handle_message_inner(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int):
    if update.message.document:
        doc = update.message.document
        fname = doc.file_name or ""
        # Bank statement for reconciliation
        if fname.lower().endswith(".csv") or (fname.lower().endswith((".xlsx", ".xls")) and "statement" in fname.lower()):
            tg_file = await doc.get_file()
            file_bytes = bytes(await tg_file.download_as_bytearray())
            await handle_statement_upload(file_bytes, fname, user_id, update)
            return
        if fname.lower().endswith((".xlsx", ".xls")):
            tg_file = await doc.get_file()
            file_bytes = bytes(await tg_file.download_as_bytearray())

            # Auto-detect column order from the file's own header row
            try:
                import openpyxl
                wb_peek = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
                ws_peek = wb_peek.active
                first_row = next(ws_peek.iter_rows(max_row=1, values_only=True), None)
                auto_cols = [str(c).strip().replace('\xa0','') for c in first_row if c] if first_row else []
            except Exception as e:
                print(f"Excel header auto-detect error: {e}")
                auto_cols = []

            if auto_cols:
                # Got headers from file — import directly without asking
                col_str = ", ".join(auto_cols)
                await update.message.reply_text(f"Detected columns: {col_str}\nImporting now...")
                await handle_excel_import(file_bytes, auto_cols, update)
            elif user_id in excel_import_sessions and excel_import_sessions[user_id].get("step") == "awaiting_file":
                col_order = excel_import_sessions[user_id].get("column_order", [])
                del excel_import_sessions[user_id]
                await handle_excel_import(file_bytes, col_order, update)
            else:
                excel_import_sessions[user_id] = {
                    "step": "awaiting_columns",
                    "file_bytes": file_bytes
                }
                await update.message.reply_text(
                    "Got the file but couldn't read its headers. Tell me the column order — "
                    "e.g. 'Name, Alias, Email, Date of Birth'"
                )
        else:
            await update.message.reply_text("I can only import .xlsx or .xls files for CRM.")
        return

    # Handle photo messages — receipt with caption as expense
    if update.message.photo:
        caption = (update.message.caption or "").strip()
        photo = update.message.photo[-1]  # highest resolution
        tg_file = await photo.get_file()
        file_bytes = bytes(await tg_file.download_as_bytearray())

        # --- Drive upload: month subfolder + temp name, renamed after merchant known ---
        receipt_link = ""
        drive_file_id = ""
        today_obj = date.today()
        month_folder_name = today_obj.strftime("%Y-%m")
        today_str = today_obj.strftime("%Y-%m")

        try:
            from googleapiclient.http import MediaIoBaseUpload
            receipts_root = DRIVE_FOLDERS.get("receipts", "")
            if receipts_root:
                # Get or create month subfolder under Receipts
                month_folder_id = get_or_create_drive_folder(month_folder_name, receipts_root)
                # Upload with temp name — will rename once merchant is known
                temp_name = f"{today_str}-receipt-{photo.file_id[:8]}.jpg"
                media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype="image/jpeg")
                file_meta = {"name": temp_name, "parents": [month_folder_id]}
                uploaded = drive_service.files().create(
                    body=file_meta, media_body=media, fields="id,webViewLink",
                    supportsAllDrives=True
                ).execute()
                drive_file_id = uploaded.get("id", "")
                receipt_link = uploaded.get("webViewLink", "")
        except Exception as e:
            print(f"Receipt upload error: {e}")

        def rename_receipt_in_drive(merchant_name):
            """Rename the uploaded receipt file to YYYY-MM-merchant.jpg"""
            if not drive_file_id or not merchant_name:
                return
            try:
                safe_merchant = re.sub(r"[^a-zA-Z0-9\-_]", "", merchant_name.replace(" ", "-").lower())
                new_name = f"{today_str}-{safe_merchant}.jpg"
                drive_service.files().update(
                    fileId=drive_file_id,
                    body={"name": new_name},
                    supportsAllDrives=True
                ).execute()
            except Exception as e:
                print(f"Receipt rename error: {e}")

        if not caption:
            # No caption — use Claude vision to read the receipt
            await update.message.reply_text("Got the receipt 🧾 Reading it now...")
            try:
                import base64
                img_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
                curr = overseas_state.get("currency", "SGD") if overseas_state.get("active") else "SGD"
                vision_resp = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=300,
                    messages=[{
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64}
                            },
                            {
                                "type": "text",
                                "text": (
                                    f"This is a receipt. Extract: merchant name, total amount, and currency (default {curr} if not shown). "
                                    "For the merchant name, copy it exactly as printed at the very top of the receipt — including any numbers or prefixes (e.g. '108 Matcha Saro', not 'Matcha Saro'). "
                                    "Do not shorten, capitalise, or summarise — copy the exact name character by character. "
                                    "For amount, use the final total (Net Total, Total, Grand Total — not subtotal). "
                                    "Reply ONLY in this format with no other text: MERCHANT | AMOUNT | CURRENCY\n"
                                    "Example: 108 Matcha Saro | 85.70 | MYR"
                                )
                            }
                        ]
                    }]
                )
                vision_text = vision_resp.content[0].text.strip()
                parts = [p.strip() for p in vision_text.split("|")]
                if len(parts) == 3:
                    merchant_v, amount_v, currency_v = parts
                    rename_receipt_in_drive(merchant_v)
                    synthesized = f"{merchant_v} {amount_v} {currency_v}"
                    reply, needs_session, session_data = handle_expense_text(synthesized, user_id, receipt_link=receipt_link)
                    if needs_session and session_data:
                        expense_sessions[user_id] = session_data
                    if reply:
                        await send_safe(update.message, reply, parse_mode="Markdown")
                else:
                    await update.message.reply_text(f"Read the receipt but got an unexpected format from vision — got: '{vision_text}'\nTry adding a caption like '45.50 Ichiran'.")
            except Exception as e:
                print(f"Vision parse error: {e}")
                await update.message.reply_text(f"❌ Couldn't read the receipt automatically ({type(e).__name__}: {str(e)[:80]}). Add a caption like '45.50 Ichiran' and resend.")
            return

        if is_expense_input(caption):
            # Caption provided — extract merchant from caption to rename file
            parsed = parse_expense_text(caption)
            if parsed and parsed.get("merchant"):
                rename_receipt_in_drive(parsed["merchant"])
            reply, needs_session, session_data = handle_expense_text(caption, user_id, receipt_link=receipt_link)
            if needs_session and session_data:
                expense_sessions[user_id] = session_data
            if reply:
                await send_safe(update.message, reply, parse_mode="Markdown")
        else:
            await update.message.reply_text("Got the receipt but couldn't read that as an expense. Try a caption like '1400 Ichiran' or 'spent 45 at Uniqlo'.")
        return

    text = update.message.text.strip()
    lower = text.lower()

    # --- DND intercept — hold messages while DND is active ---
    if em_profile.get("dnd_active") and lower not in ("dnd off", "dnd on"):
        held = em_profile.get("dnd_held_messages", [])
        held.append({"text": text, "time": datetime.now(TIMEZONE).strftime("%H:%M")})
        em_profile["dnd_held_messages"] = held
        save_em_profile()
        return  # silently hold — no reply while DND active

    # Check birthday acknowledgement — explicit sent/skip only
    bday_handled, bday_reply = check_birthday_acknowledgement(text)
    if bday_handled:
        if bday_reply:
            await update.message.reply_text(bday_reply)
        return

    # Excel import column declaration
    if user_id in excel_import_sessions:
        session = excel_import_sessions[user_id]
        if session.get("step") == "awaiting_columns":
            cols = parse_excel_column_order(text)
            if cols:
                if session.get("file_bytes"):
                    # File already uploaded, import immediately
                    file_bytes = session["file_bytes"]
                    del excel_import_sessions[user_id]
                    await update.message.reply_text(f"Got it — columns: {', '.join(cols)}. Importing now...")
                    await handle_excel_import(file_bytes, cols, update)
                else:
                    session["column_order"] = cols
                    session["step"] = "awaiting_file"
                    await update.message.reply_text(
                        f"Got it — columns: {', '.join(cols)}. Now send the Excel file."
                    )
            else:
                await update.message.reply_text("Couldn't parse that. Try: 'Name, Email, Date of Birth, Alias'")
            return

    # Check session timeouts before processing any session
    if any(user_id in s for s in [expense_sessions, delete_sessions, portfolio_delete_sessions,
                                    confirm_sessions, receipt_confirm_sessions, edit_sessions,
                                    meeting_sessions, pending_restaurant_saves]):
        if is_session_expired(user_id):
            await check_session_timeouts(user_id, update)
            return

    # Session interrupt guard — if mid-session and message looks like a new intent, pause and confirm
    active_label = get_active_session_label(user_id)
    if active_label and looks_like_new_intent(text):
        # If already waiting for interrupt confirmation, handle it
        if user_id in interrupted_sessions:
            pending = interrupted_sessions[user_id]
            if text.strip().lower() in ["yes", "y"]:
                # Clear all active sessions and replay the pending message
                for d in [receipt_confirm_sessions, expense_sessions, meeting_sessions,
                          edit_sessions, confirm_sessions, delete_sessions,
                          portfolio_delete_sessions, excel_import_sessions,
                          pending_restaurant_saves, pending_contact_saves,
                          todo_disambig_sessions]:
                    d.pop(user_id, None)

                session_timestamps.pop(user_id, None)
                del interrupted_sessions[user_id]
                # Replay by substituting the pending text and continuing
                update.message.text = pending["pending_text"]
                text = pending["pending_text"]
                lower = text.lower()
                # Fall through to normal routing with the replayed message
            elif text.strip().lower() in ["no", "n"]:
                del interrupted_sessions[user_id]
                await update.message.reply_text(f"Got it — continuing with your {pending['label']}.")
                return
            else:
                await update.message.reply_text(f"Reply yes to switch, or no to continue with your {pending['label']}.")
                return
        else:
            # First interrupt — store and ask (don't overwrite if already waiting)
            if user_id not in interrupted_sessions:
                interrupted_sessions[user_id] = {"label": active_label, "pending_text": text}
            await update.message.reply_text(
                f"You're mid-{active_label} — did you mean to do something else?\n"
                f"Reply yes to switch, or no to continue."
            )
            return

    # Handle pending restaurant saves (location needed or duplicate)
    if user_id in pending_restaurant_saves:
        prs = pending_restaurant_saves[user_id]
        step = prs.get("step")

        if step == "awaiting_location":
            location = text.strip()
            del pending_restaurant_saves[user_id]
            result = save_restaurant(prs["name"], location, prs.get("country", "Singapore"))
            if result and result.startswith("_DUPLICATE_:"):
                pending_restaurant_saves[user_id] = {
                    "step": "duplicate", "existing": prs["name"],
                    "name": prs["name"], "location": location,
                    "country": prs.get("country", "Singapore"), "tags": "", "notes": ""
                }
                await update.message.reply_text(f"*{prs['name']}* is already in your list. Update or save as new? (update / new)")
            else:
                await update.message.reply_text(format_restaurant_saved(prs["name"], location))
            return

        elif step == "awaiting_confirm":
            # Parse yes/no and optional tags from reply
            extra_tags = ""
            words = lower.split(None, 1)
            first_word = words[0] if words else ""
            rest = words[1] if len(words) > 1 else ""

            if first_word in ["yes", "y"]:
                extra_tags = rest.strip()
                merged_tags = ", ".join(filter(None, [prs.get("tags", ""), extra_tags]))
                location = prs.get("location", "")
                del pending_restaurant_saves[user_id]
                result = save_restaurant(prs["name"], location, prs.get("country", "Singapore"), merged_tags)
                if result and result.startswith("_DUPLICATE_:"):
                    pending_restaurant_saves[user_id] = {
                        "step": "duplicate", "existing": prs["name"],
                        "name": prs["name"], "location": location,
                        "country": prs.get("country", "Singapore"), "tags": merged_tags, "notes": ""
                    }
                    await update.message.reply_text(f"*{prs['name']}* is already in your list. Update or save as new? (update / new)")
                else:
                    await update.message.reply_text(format_restaurant_saved(prs["name"], location, merged_tags))
            elif first_word in ["no", "n"]:
                del pending_restaurant_saves[user_id]
                await update.message.reply_text(f"Got it — what's the correct location for {prs['name']}?")
                pending_restaurant_saves[user_id] = {"name": prs["name"], "country": prs.get("country", "Singapore"), "step": "awaiting_location"}
            else:
                await update.message.reply_text("Reply yes to confirm or no to correct the location.")
            return

        elif step == "awaiting_outlet":
            # Parse outlet number and optional tags
            outlets = prs.get("outlets", [])
            words = lower.split(None, 1)
            first_word = words[0].strip().rstrip(",")
            rest = words[1].strip() if len(words) > 1 else ""
            try:
                idx = int(first_word) - 1
                if 0 <= idx < len(outlets):
                    location = outlets[idx]
                    merged_tags = ", ".join(filter(None, [prs.get("tags", ""), rest]))
                    del pending_restaurant_saves[user_id]
                    result = save_restaurant(prs["name"], location, prs.get("country", "Singapore"), merged_tags)
                    if result and result.startswith("_DUPLICATE_:"):
                        pending_restaurant_saves[user_id] = {
                            "step": "duplicate", "existing": prs["name"],
                            "name": prs["name"], "location": location,
                            "country": prs.get("country", "Singapore"), "tags": merged_tags, "notes": ""
                        }
                        await update.message.reply_text(f"*{prs['name']}* is already in your list. Update or save as new? (update / new)")
                    else:
                        await update.message.reply_text(format_restaurant_saved(prs["name"], location, merged_tags))
                else:
                    await update.message.reply_text(f"Pick a number between 1 and {len(outlets)}.")
            except ValueError:
                await update.message.reply_text(f"Reply with the number of the outlet (1–{len(outlets)}).")
            return

        elif step == "duplicate":
            if lower in ["new", "save new"]:
                del pending_restaurant_saves[user_id]
                save_restaurant(prs["name"], prs["location"], prs.get("country", "Singapore"),
                                prs.get("tags", ""), prs.get("notes", ""), force_new=True)
                await update.message.reply_text(format_restaurant_saved(prs["name"], prs["location"], prs.get("tags", "")))
            elif lower in ["update", "update existing"]:
                del pending_restaurant_saves[user_id]
                await update.message.reply_text(f"Use 'edit restaurant {prs['existing']}' to update it.")
            elif lower in ["skip", "s", "cancel", "no", "n", "nope", "nah"]:
                del pending_restaurant_saves[user_id]
                await update.message.reply_text("Skipped.")
            else:
                await update.message.reply_text("Reply 'update', 'new', or 'skip'.")
            return

    # Handle pending duplicate contact save
    if user_id in pending_contact_saves:
        pcs = pending_contact_saves[user_id]
        if lower in ["new", "save new", "save as new"]:
            del pending_contact_saves[user_id]
            reply = save_contact(pcs["data"], force_new=True)
            await send_safe(update.message, reply, parse_mode="Markdown")
        elif lower in ["update", "update existing"]:
            del pending_contact_saves[user_id]
            reply = f"Opening {pcs['existing_name']} for editing — use 'edit {pcs['existing_name']}' to update fields."
            await send_safe(update.message, reply, parse_mode="Markdown")
        elif lower in ["skip", "s", "cancel", "no", "n", "nope", "nah"]:
            del pending_contact_saves[user_id]
            await update.message.reply_text("Skipped.")
        else:
            await update.message.reply_text(f"*{pcs['existing_name']}* already exists. Reply 'update', 'new', or 'skip'.")
        return

    # Handle todo disambiguation
    if user_id in todo_disambig_sessions:
        td = todo_disambig_sessions[user_id]
        tasks = td.get("tasks", [])
        action = td.get("action", "complete")
        if text.strip().isdigit():
            idx = int(text.strip()) - 1
            if 0 <= idx < len(tasks):
                task_name = tasks[idx]
                del todo_disambig_sessions[user_id]
                if action == "complete":
                    sheet = todo_sheet()
                    records = sheet.get_all_records()
                    for i, r in enumerate(records):
                        if r.get("Task", "") == task_name and r.get("Status") == "Pending":
                            sheet.update_cell(i + 2, 2, "Done")
                            await update.message.reply_text(f"✅ Marked as done: _{task_name}_")
                            return
                elif action == "delete":
                    sheet = todo_sheet()
                    records = sheet.get_all_records()
                    for i, r in enumerate(records):
                        if r.get("Task", "") == task_name:
                            sheet.delete_rows(i + 2)
                            await update.message.reply_text(f"Deleted — {task_name} ✅")
                            return
            else:
                await update.message.reply_text("Invalid number. Try again or 'cancel'.")
        elif lower == "cancel":
            del todo_disambig_sessions[user_id]
            await update.message.reply_text("Cancelled.")
        else:
            await update.message.reply_text("Reply with a number or 'cancel'.")
        return

    # Handle active receipt confirm session
    if user_id in receipt_confirm_sessions:
        touch_session(user_id)
        await handle_receipt_confirm_session(user_id, text, update)
        return

    # Handle active meeting recap session
    if user_id in meeting_sessions:
        touch_session(user_id)
        await handle_meeting_session(user_id, text, update)
        return

    # Handle active expense onboarding session
    if user_id in expense_sessions:
        touch_session(user_id)
        await handle_expense_session(user_id, text, update)
        return

    # Handle active delete session
    if user_id in delete_sessions:
        session = delete_sessions[user_id]
        step = session.get("step")

        if lower in ["cancel", "nevermind", "never mind", "nvm"]:
            del delete_sessions[user_id]
            await update.message.reply_text("Cancelled.")
            return

        if step == "pick":
            # User replies with a number
            if text.strip().isdigit():
                idx = int(text.strip()) - 1
                expenses = session.get("expenses", [])
                if 0 <= idx < len(expenses):
                    sheet_row, _ = expenses[idx]
                    reply = delete_expense_by_row(sheet_row)
                    del delete_sessions[user_id]
                    await send_safe(update.message, reply, parse_mode="Markdown")
                else:
                    await update.message.reply_text("Invalid number. Try again or 'search [merchant]'.")
                return
            # User wants to search instead
            elif lower.startswith("search "):
                query = text[7:].strip()
                results = search_expenses_by_merchant(query)
                if not results:
                    await update.message.reply_text(f"No expenses found matching '{query}'.")
                elif len(results) == 1:
                    sheet_row, r = results[0]
                    reply = delete_expense_by_row(sheet_row)
                    del delete_sessions[user_id]
                    await send_safe(update.message, reply, parse_mode="Markdown")
                else:
                    delete_sessions[user_id] = {"step": "pick", "expenses": results}
                    await update.message.reply_text(format_delete_list(results))
                return
            elif lower in ["not there", "not here", "none of these", "not in the list"]:
                await update.message.reply_text("Reply 'search [merchant name]' to find it.")
                return
            else:
                await update.message.reply_text("Reply with a number, 'search [merchant]', or 'cancel'.")
                return

    # Handle active confirm session (destructive actions: delete contact/bill/restaurant/event, rename category)
    if user_id in confirm_sessions:
        touch_session(user_id)
        cs = confirm_sessions[user_id]
        action = cs.get("action")
        args = cs.get("args", [])
        if lower in ["yes", "y"]:
            del confirm_sessions[user_id]
            session_timestamps.pop(user_id, None)
            if action == "delete_contact":
                reply = delete_contact(args[0])
            elif action == "rename_category":
                reply = rename_category(args[0], args[1])
            elif action == "delete_bill":
                reply = delete_bill(args[0])
            elif action == "delete_restaurant":
                reply = delete_restaurant(args[0])
            elif action == "delete_event":
                reply = delete_calendar_event(args[0])
            else:
                reply = "Done."
            await send_safe(update.message, reply, parse_mode="Markdown")
        elif lower in ["no", "n", "cancel"]:
            del confirm_sessions[user_id]
            session_timestamps.pop(user_id, None)
            await update.message.reply_text("Cancelled.")
        else:
            await update.message.reply_text(f"{cs.get('target', 'Confirm?')} (yes / no)")
        return

    # Handle active portfolio delete session
    if user_id in portfolio_delete_sessions:
        touch_session(user_id)
        pd_session = portfolio_delete_sessions[user_id]
        if lower in ["cancel", "nevermind", "nvm"]:
            del portfolio_delete_sessions[user_id]
            await update.message.reply_text("Cancelled.")
            return
        if text.strip().isdigit():
            idx = int(text.strip()) - 1
            rows = pd_session.get("rows", [])
            if 0 <= idx < len(rows):
                sheet_row, _ = rows[idx]
                del portfolio_delete_sessions[user_id]
                result = delete_portfolio_row(sheet_row)
                await update.message.reply_text(result)
            else:
                await update.message.reply_text("Invalid number. Try again or 'cancel'.")
            return
        elif lower.startswith("search "):
            query = text[7:].strip()
            results = search_portfolio_by_ticker(query)
            if not results:
                await update.message.reply_text(f"No holdings found matching '{query}'.")
            elif len(results) == 1:
                sheet_row, _ = results[0]
                del portfolio_delete_sessions[user_id]
                result = delete_portfolio_row(sheet_row)
                await update.message.reply_text(result)
            else:
                portfolio_delete_sessions[user_id] = {"step": "pick", "rows": results}
                await update.message.reply_text(format_portfolio_delete_list(results))
            return
        else:
            await update.message.reply_text("Reply with a number, 'search [ticker]', or 'cancel'.")
            return

    # Handle active CRM edit session
    if user_id in edit_sessions:
        await handle_edit_session(user_id, text, update)
        return

    # Handle active reconciliation session
    if user_id in recon_sessions:
        touch_session(user_id)
        rs = recon_sessions[user_id]
        step = rs.get("step")
        unmatched = rs.get("unmatched", [])
        idx = rs.get("index", 0)
        if lower in ["done", "skip all", "close"]:
            del recon_sessions[user_id]
            await update.message.reply_text("Reconciliation session closed ✅")
        elif lower.startswith("log "):
            # User wants to log an unmatched item as an expense
            log_text = text[4:].strip()
            reply_str, needs_session, session_data = handle_expense_text(log_text, user_id)
            if needs_session and session_data:
                expense_sessions[user_id] = session_data
            del recon_sessions[user_id]
            await send_safe(update.message, reply_str, parse_mode="Markdown")
        elif lower in ["skip", "next", "s"]:
            rs["index"] = idx + 1
            if rs["index"] < len(unmatched):
                await update.message.reply_text(
                    f"Next unmatched ({rs['index']+1}/{len(unmatched)}):\n{unmatched[rs['index']]}\n\nReply 'log [expense]' to log, 'skip' for next, or 'done' to close."
                )
            else:
                del recon_sessions[user_id]
                await update.message.reply_text("All unmatched items reviewed ✅")
        else:
            if idx < len(unmatched):
                await update.message.reply_text(
                    f"Unmatched item ({idx+1}/{len(unmatched)}):\n{unmatched[idx]}\n\nReply 'log [expense]' to log it, 'skip' for next, or 'done' to close."
                )
            else:
                del recon_sessions[user_id]
                await update.message.reply_text("All unmatched items reviewed ✅")
        return

    # Trip setup session — step-by-step collection
    if overseas_state.get("_trip_setup"):
        ts = overseas_state["_trip_setup"]
        step = ts.get("step", "destination")
        t = text.strip()
        skipped = t.lower() in ("skip", "s", "-", "later", "idk", "not sure")

        # Persist trip setup state on every step so it survives restarts
        overseas_state["_trip_setup"] = ts
        persist_trip_setup()

        if step == "destination":
            if not t:
                await update.message.reply_text("Where are you headed?")
                return
            ts["destination"] = t.title()
            ts["step"] = "check_in"
            await update.message.reply_text(f"Got it — {ts['destination']} 🌏\nCheck-in date? (or 'skip')")
            return

        elif step == "check_in":
            if not skipped:
                dates = _parse_trip_dates(t)
                ts["check_in"] = dates[0].strftime("%d/%m/%Y") if dates else t
            ts["step"] = "check_out"
            await update.message.reply_text("Check-out date? (or 'skip')")
            return

        elif step == "check_out":
            if not skipped:
                dates = _parse_trip_dates(t)
                ts["check_out"] = dates[0].strftime("%d/%m/%Y") if dates else t
            ts["step"] = "flight"
            await update.message.reply_text("Flight number? (or 'skip')")
            return

        elif step == "flight":
            if not skipped:
                fn = extract_flight_number(t.upper())
                ts["flight_number"] = fn or t.upper()
                ts["step"] = "dep_time"
                if AVIATIONSTACK_API_KEY and fn:
                    # Try AviationStack
                    try:
                        with httpx.Client(timeout=10) as hx:
                            resp = hx.get(
                                "http://api.aviationstack.com/v1/flights",
                                params={"access_key": AVIATIONSTACK_API_KEY, "flight_iata": fn}
                            )
                        data = resp.json()
                        flights = data.get("data", [])
                        if flights:
                            dep = flights[0].get("departure", {})
                            scheduled = dep.get("scheduled", "")
                            if scheduled:
                                try:
                                    from datetime import datetime as _dt
                                    dt = _dt.fromisoformat(scheduled.replace("Z", "+00:00"))
                                    ts["dep_time_iso"] = scheduled
                                    ts["dep_time_display"] = dt.astimezone(TIMEZONE).strftime("%d %b %H:%M")
                                    ts["step"] = "currency"
                                    await update.message.reply_text(
                                        f"Found {fn} — departs {ts['dep_time_display']} ✅\nCurrency? (or 'skip' for SGD)"
                                    )
                                    return
                                except Exception:
                                    pass
                    except Exception as e:
                        print(f"AviationStack lookup error: {e}")
                await update.message.reply_text(f"What time does {ts['flight_number']} depart? (e.g. 09:20, or 'skip')")
            else:
                ts["step"] = "currency"
                await update.message.reply_text("Currency? (or 'skip' for SGD)")
            return

        elif step == "dep_time":
            if not skipped:
                time_match = re.search(r'(\d{1,2}:\d{2})', t)
                if time_match:
                    time_str = time_match.group(1)
                    # Combine with check_in date if available
                    ci = ts.get("check_in", "")
                    dep_dt = None
                    if ci:
                        try:
                            dep_dt = TIMEZONE.localize(
                                datetime.strptime(f"{ci} {time_str}", "%d/%m/%Y %H:%M")
                            )
                        except Exception:
                            pass
                    ts["dep_time_iso"] = dep_dt.isoformat() if dep_dt else time_str
                    ts["dep_time_display"] = dep_dt.strftime("%d %b %H:%M") if dep_dt else time_str
                else:
                    ts["dep_time_display"] = t
            ts["step"] = "currency"
            await update.message.reply_text("Currency? (or 'skip' for SGD)")
            return

        elif step == "currency":
            if not skipped:
                cur_match = re.search(r'([A-Z]{3})', t.upper())
                ts["currency"] = cur_match.group(1) if cur_match else _get_currency_for_dest(ts.get("destination", ""))
            else:
                ts["currency"] = _get_currency_for_dest(ts.get("destination", "")) or "SGD"
            ts["step"] = "hotel"
            await update.message.reply_text("Hotel name? (or 'skip')")
            return

        elif step == "hotel":
            if not skipped:
                ts["hotel_name"] = t
                ts["step"] = "hotel_local"
                await update.message.reply_text("Hotel local name? (or 'skip')")
            else:
                ts["step"] = "confirm"
                # Skip straight to confirm
                ts_copy = dict(ts)
                await _send_trip_confirm(update, ts_copy)
            return

        elif step == "hotel_local":
            if not skipped:
                ts["hotel_local_name"] = t
                ts["step"] = "hotel_address"
                await update.message.reply_text("Hotel address? (or 'skip')")
            else:
                ts["step"] = "confirm"
                ts_copy = dict(ts)
                await _send_trip_confirm(update, ts_copy)
            return

        elif step == "hotel_address":
            if not skipped:
                ts["hotel_address"] = t
            ts["step"] = "confirm"
            ts_copy = dict(ts)
            await _send_trip_confirm(update, ts_copy)
            return

        elif step == "confirm":
            if t.upper() == "Y":
                overseas_state.pop("_trip_setup", None)
                persist_trip_setup()
                dest = ts.get("destination", "Unknown")
                curr = ts.get("currency", "SGD")
                check_in = ts.get("check_in", "")
                check_out = ts.get("check_out", "")
                dep_iso = ts.get("dep_time_iso", "")
                # Schedule or activate
                scheduled = False
                if dep_iso and _scheduler:
                    try:
                        dep_dt = datetime.fromisoformat(dep_iso.replace("Z", "+00:00"))
                        dep_local = dep_dt.astimezone(TIMEZONE)
                        if dep_local > datetime.now(TIMEZONE):
                            job = _scheduler.add_job(
                                activate_overseas_mode_scheduled,
                                "date",
                                run_date=dep_local,
                                args=[dest, curr, check_in, check_out]
                            )
                            overseas_state["dep_job_id"] = job.id
                            overseas_state["destination"] = dest
                            overseas_state["currency"] = curr
                            scheduled = True
                    except Exception as e:
                        print(f"Trip schedule error: {e}")
                if not scheduled:
                    for d in [expense_sessions, receipt_confirm_sessions]:
                        d.pop(user_id, None)
                    session_timestamps.pop(user_id, None)
                    overseas_state["active"] = True
                    overseas_state["destination"] = dest
                    overseas_state["currency"] = curr
                    overseas_state["currencies"] = [curr] if curr != "SGD" else []
                    overseas_state["trip_destinations"] = [dest]
                    overseas_state["trip_start"] = date.today().strftime("%d/%m/%Y")
                    save_trip(dest, curr, check_in=check_in, check_out=check_out,
                              hotel_name=ts.get("hotel_name", ""),
                              hotel_local_name=ts.get("hotel_local_name", ""),
                              hotel_address=ts.get("hotel_address", ""))
                    dep_display = ts.get("dep_time_display", "")
                    if dep_display:
                        reply = f"Overseas mode on ✈️\nDestination: {dest}\nCurrency: {curr}\nFlight: {ts.get('flight_number','')} {dep_display}"
                    else:
                        reply = f"Overseas mode on ✈️\nDestination: {dest}\nCurrency: {curr}\nI\'ll log expenses in {curr} with SGD equivalent."
                else:
                    dep_display = ts.get("dep_time_display", "")
                    reply = (
                        f"Got it ✈️ Overseas mode will activate at departure: {dep_display}\n"
                        f"Destination: {dest} ({curr})\n"
                        f"I\'ll send a confirmation when it kicks in."
                    )
                await send_safe(update.message, reply, parse_mode="Markdown")
            else:
                overseas_state.pop("_trip_setup", None)
                persist_trip_setup()
                await update.message.reply_text("Trip setup cancelled.")
            return

    reply = None

    # CRM Commands
    if lower.startswith("save restaurant "):
        result = handle_save_restaurant(text)
        if result and result.startswith("_NEEDS_LOCATION_:"):
            parts = result.split(":", 2)
            name = parts[1] if len(parts) > 1 else ""
            country = parts[2] if len(parts) > 2 else "Singapore"
            pending_restaurant_saves[user_id] = {"name": name, "country": country, "step": "awaiting_location"}
            reply = f"⚠️ Couldn't read that Maps link — what's the location for {name}?\n(e.g. 313 Orchard Road, Singapore)"
        elif result and result.startswith("_INFER_LOCATION_:"):
            parts = result.split(":", 5)
            name = parts[1] if len(parts) > 1 else ""
            country = parts[2] if len(parts) > 2 else "Singapore"
            tags = parts[3] if len(parts) > 3 else ""
            multiple = parts[4] == "1" if len(parts) > 4 else False
            outlets = parts[5].split("|") if len(parts) > 5 and parts[5] else []
            if multiple and len(outlets) > 1:
                pending_restaurant_saves[user_id] = {
                    "step": "awaiting_outlet", "name": name, "country": country,
                    "tags": tags, "outlets": outlets
                }
                outlet_list = "\n".join(f"{i+1}. {o}" for i, o in enumerate(outlets))
                reply = f"Found a few {name} outlets — which one, and any tags?\n{outlet_list}"
            else:
                location = outlets[0] if outlets else ""
                pending_restaurant_saves[user_id] = {
                    "step": "awaiting_confirm", "name": name, "country": country,
                    "tags": tags, "location": location
                }
                reply = f"Got it — I have {name} at {location}. Is that right, and any tags? (yes / no, add tags or skip)"
        elif result and result.startswith("_DUPLICATE_RESTAURANT_:"):
            parts = result.split(":", 5)
            pending_restaurant_saves[user_id] = {
                "step": "duplicate", "existing": parts[1], "name": parts[2],
                "location": parts[3], "country": parts[4],
                "tags": parts[5] if len(parts) > 5 else "", "notes": ""
            }
            reply = f"*{parts[1]}* is already in your list.\nUpdate existing or save as new? (update / new)"
        else:
            reply = result
    elif lower.startswith("save ") and not is_restaurant_save(text):
        result = save_contact(text[5:])
        if result.startswith("_DUPLICATE_:"):
            existing_name = result.split(":", 1)[1]
            pending_contact_saves[user_id] = {"data": text[5:], "existing_name": existing_name}
            reply = f"*{existing_name}* already exists in your CRM.\nUpdate the existing contact or save as new? (update / new)"
        else:
            reply = result
    elif lower.startswith("find ") or lower.startswith("pull up "):
        name = text[5:] if lower.startswith("find ") else text[8:]
        reply = find_contact(name)
    elif lower.startswith("note "):
        reply = add_note(text[5:])
    elif lower.startswith("followup "):
        reply = set_followup(text[9:])
    elif lower.startswith("update "):
        reply = update_field(text[7:])
    elif lower.startswith("delete ") and not lower.startswith("delete event") and not lower.startswith("delete expense") and not lower.startswith("delete bill") and not lower.startswith("delete restaurant") and not lower.startswith("delete last"):
        name = text[7:].strip()
        _, record = find_row(name)
        if not record:
            reply = f"No contact found for '{name}'"
        else:
            confirm_name = record.get("Name", name)
            confirm_sessions[user_id] = {"action": "delete_contact", "args": [name], "target": f"Delete contact {confirm_name}?"}
            reply = f"Delete contact *{confirm_name}*? (yes / no)"
    elif lower.startswith("search "):
        reply = search_contacts(text[7:])
    elif lower.startswith("edit ") and not lower.startswith("edit last expense") and not lower.startswith("edit expense"):
        name = text[5:].strip()
        _, record = find_row(name)
        if not record:
            reply = f"❌ No contact found for '{name}'"
        else:
            edit_sessions[user_id] = {"name": record.get("Name"), "step": "choose_field"}
            reply = (
                f"Editing *{record.get('Name')}*. Which field?\n\n"
                f"1. Alias\n2. Birthday\n3. Relationship\n4. Context\n5. Notes\n"
                f"6. Follow up date\n7. Follow up notes\n8. Email\n9. Address\n\n"
                f"Type the field name or *cancel* to exit."
            )
    elif lower == "cancel":
        if user_id in edit_sessions:
            del edit_sessions[user_id]
        if user_id in excel_import_sessions:
            del excel_import_sessions[user_id]
        if user_id in confirm_sessions:
            del confirm_sessions[user_id]
        if user_id in delete_sessions:
            del delete_sessions[user_id]
        if user_id in receipt_confirm_sessions:
            del receipt_confirm_sessions[user_id]
        session_timestamps.pop(user_id, None)
        reply = "Cancelled."

    # DND commands
    elif lower == "dnd on":
        em_profile["dnd_active"] = True
        em_profile["dnd_held_messages"] = []
        save_em_profile()
        reply = "DND on — holding all messages until you turn it off."
    elif lower == "dnd off":
        em_profile["dnd_active"] = False
        held = em_profile.get("dnd_held_messages", [])
        em_profile["dnd_held_messages"] = []
        save_em_profile()
        if held:
            lines = [f"DND off. {len(held)} message(s) held while you were away:\n"]
            for m in held:
                lines.append(f"[{m['time']}] {m['text']}")
            reply = "\n".join(lines)
        else:
            reply = "DND off."

    # Profile
    elif lower == "reload profile":
        try:
            setup_em_profile()
            reply = "✅ Profile reloaded successfully."
        except Exception as e:
            reply = f"⚠️ Couldn't reload profile: {str(e)}"

    # Missed items dismissal
    elif lower in ["skip missed", "dismiss missed", "skip all missed"]:
        reply = "All missed follow-ups dismissed ✅"
    elif lower.startswith("skip ") and not lower.startswith("skip missed bills"):
        name = text[5:].strip()
        reply = f"Follow-up for {name} dismissed ✅"
    elif lower in ["skip missed bills", "dismiss missed bills"]:
        reply = "Missed bill reminders dismissed ✅"

    # Missed market summary response
    elif lower in ["yes", "y", "yep", "yeah", "yup"] and market_summary_pending.get(user_id):
        market_summary_pending.pop(user_id, None)
        reply = get_market_summary_now()

    # Card defaults
    elif re.match(r"set default card for .+ to .+", lower):
        m = re.match(r"set default card for (.+?) to (.+)", lower)
        if m:
            category_name = m.group(1).strip().title()
            card_name = m.group(2).strip().title()
            reply = set_card_default_category(card_name, category_name)
        else:
            reply = "Try: 'set default card for FnB to Maybank'"


    elif lower == "list":
        reply = list_contacts()
    elif lower == "stats":
        reply = get_stats()
    elif lower == "followups" or lower in [
        "show my followups", "show followups", "pending followups",
        "what followups do i have", "my followups", "list followups",
        "upcoming followups", "show my follow ups", "what are my followups"
    ]:
        reply = upcoming_followups()
    elif lower == "overdue":
        reply = overdue_followups()
    elif lower == "birthdays":
        reply = upcoming_birthdays(30)
    elif lower == "soon":
        reply = upcoming_birthdays(7)
    elif lower.startswith("lastcontact "):
        reply = last_contact(text[12:])

    # Referral queries
    elif lower in ["referrals", "all referrals", "show referrals"]:
        reply = get_all_referrals()
    elif lower in ["top referrers", "best referrers", "who refers the most"]:
        reply = get_top_referrers()
    elif lower.startswith("referrals from ") or lower.startswith("who did "):
        if lower.startswith("referrals from "):
            name = text[15:].strip()
        else:
            # "who did James refer"
            m = re.search(r"who did (.+?) refer", lower)
            name = m.group(1).strip().title() if m else text[8:].strip()
        reply = get_referrals_by(name)

    # Import
    elif "import" in lower and ("excel" in lower or "contacts" in lower or "spreadsheet" in lower):
        excel_import_sessions[user_id] = {"step": "awaiting_columns", "column_order": []}
        reply = "Sure! Tell me the column order in your Excel first — e.g. 'Name, Email, Date of Birth, Alias'"

    # Meeting Recap Commands
    elif lower.startswith("search meetings ") or lower.startswith("find meeting "):
        query = text[16:] if lower.startswith("search meetings ") else text[13:]
        reply = search_meeting_notes(query.strip())
    elif lower == "cancel" and user_id in meeting_sessions:
        del meeting_sessions[user_id]
        reply = "Recap cancelled."
    elif is_meeting_start(text):
        event_name = extract_event_name(text)
        meeting_sessions[user_id] = {
            "step": "collecting",
            "event_name": event_name,
            "notes": []
        }
        if event_name:
            reply = f"Got it, taking notes for {event_name}. Send everything over and say done when you're finished."
        else:
            reply = "Sure, what's the event name?"
            meeting_sessions[user_id]["step"] = "get_name"

    # Reminder Commands
    elif lower == "reminders" or lower == "my reminders" or lower == "list reminders":
        reply = list_reminders()
    elif is_cancel_reminder_request(text):
        keyword = text.lower().replace("cancel", "").replace("delete", "").replace("remove", "").replace("reminder", "").strip()
        cancelled = cancel_reminder_by_keyword(keyword) if keyword else []
        if not cancelled:
            reply = "Couldn't find a matching reminder to cancel."
        elif cancelled[0].startswith("_DISAMBIG_:"):
            entries = cancelled[0][len("_DISAMBIG_:"):].split("|")
            lines = ["Found multiple matching reminders — which one to cancel?"]
            for i, entry in enumerate(entries, 1):
                parts = entry.split(":", 1)
                msg = parts[1] if len(parts) > 1 else entry
                lines.append(f"{i}. {msg}")
            lines.append("\nReply with the number or 'all' to cancel all.")
            todo_disambig_sessions[user_id] = {"tasks": [e.split(":")[0] for e in entries], "action": "cancel_reminder", "entries": entries}
            reply = "\n".join(lines)
        else:
            reply = f"Cancelled: {', '.join(cancelled)}"
    elif is_reschedule_request(text):
        reply = handle_reschedule(text, user_id)
    elif is_reminder_request(text):
        reply = handle_new_reminder(text)

    # Expense Commands
    elif any(lower == p or lower.startswith(p) for p in [
        "what expense categories", "list my categories", "what categories",
        "show categories", "what are my expense categories", "list categories",
        "my expense categories", "expense categories"
    ]):
        reply = get_expense_categories()
    elif any(lower == p or lower.startswith(p) for p in [
        "what merchants", "my merchants", "merchant map", "list merchants",
        "show merchants", "what merchants do we have", "saved merchants",
        "known merchants"
    ]):
        reply = get_merchant_list()
    elif lower.startswith("delete merchant ") or lower.startswith("remove merchant ") or lower.startswith("forget merchant "):
        merchant_name = text.split(" ", 2)[2].strip()
        reply = delete_merchant(merchant_name)
    elif lower in ["expense report", "monthly report", "spending report", "expenses",
                   "monthly summary", "monthly spend", "this month", "expense summary"]:
        reply = get_expense_report()
    elif lower in ["delete last expense", "remove last expense"]:
        reply = delete_last_expense()
    elif lower in ["delete expense", "remove expense", "undo expense", "undo last expense"]:
        # Show last 5 for selection
        recent = get_recent_expenses(5)
        if not recent:
            reply = "No expenses logged yet."
        else:
            delete_sessions[user_id] = {"step": "pick", "expenses": recent}
            reply = format_delete_list(recent)
    elif lower.startswith("delete expense ") or lower.startswith("remove expense "):
        # Direct search — "delete expense Tada"
        query = re.sub(r"^(delete|remove) expense\s+", "", lower).strip()
        results = search_expenses_by_merchant(query)
        if not results:
            reply = f"No expenses found matching '{query}'."
        elif len(results) == 1:
            sheet_row, _ = results[0]
            reply = delete_expense_by_row(sheet_row)
        else:
            delete_sessions[user_id] = {"step": "pick", "expenses": results}
            reply = format_delete_list(results)
    elif lower in ["last expense", "show last expense", "what did i log"]:
        reply = show_last_expense()
    elif lower in ["trip summary", "trip spend", "how much have i spent", "trip expenses"]:
        reply = get_trip_summary()
    elif lower in ["trip history", "my trips", "past trips", "trips"]:
        reply = format_trip_history()
    elif lower == "close trip":
        if overseas_state.get("active"):
            deactivate_overseas_mode()
            reply = "Trip closed. Back to SG mode. 🏠"
        else:
            closed = close_trip()
            reply = "Trip closed ✅" if closed else "No active trip to close."
    elif lower in ["current trip", "active trip", "am i overseas", "overseas status"]:
        if overseas_state.get("active"):
            dest = overseas_state.get("destination", "Unknown")
            curr = overseas_state.get("currency", "SGD")
            trip_start = overseas_state.get("trip_start", "")
            reply = f"✈️ Active trip: {dest} ({curr})"
            if trip_start:
                reply += f"\nStarted: {trip_start}"
        else:
            reply = "No active trip — you're in SG mode."
    elif lower.startswith("edit last expense ") or lower.startswith("edit expense "):
        # Strip the command prefix and pass the rest as the edit text
        edit_text = re.sub(r"^edit (?:last )?expense\s+", "", text, flags=re.IGNORECASE).strip()
        if edit_text:
            reply = edit_last_expense(edit_text)
        else:
            reply = show_last_expense()
    elif lower.startswith("rename category "):
        m = re.search(r"rename category (.+?) to (.+)", lower)
        if m:
            old_cat = m.group(1).strip()
            new_cat = m.group(2).strip()
            confirm_sessions[user_id] = {"action": "rename_category", "args": [old_cat, new_cat], "target": f"Rename {old_cat} to {new_cat}?"}
            reply = f"Rename category *{old_cat}* to *{new_cat}*? This will update all expense rows and Merchant Map. (yes / no)"
        else:
            reply = "Try: 'rename category FnB to Food'"
    elif lower.startswith("now in ") or lower.startswith("switched to ") or lower.startswith("arrived in "):
        # Mid-trip currency switch — "now in Korea", "arrived in Seoul"
        if overseas_state.get("active"):
            dest_text = re.sub(r"^(now in|switched to|arrived in)\s+", "", lower).strip().title()
            new_curr = _get_currency_for_dest(dest_text)
            new_dest = dest_text
            if new_curr and new_curr != "SGD":
                overseas_state["currency"] = new_curr
                overseas_state["destination"] = new_dest
                if new_curr not in overseas_state["currencies"]:
                    overseas_state["currencies"].append(new_curr)
                if new_dest not in overseas_state["trip_destinations"]:
                    overseas_state["trip_destinations"].append(new_dest)
                # Pre-cache the new currency rate
                get_fx_rate(new_curr)
                # Update Trips sheet with new destination
                try:
                    ws = trips_sheet()
                    records = ws.get_all_records()
                    for i, row in enumerate(records, start=2):
                        if row.get("Status") == "active":
                            ws.update_cell(i, 2, new_dest)
                            ws.update_cell(i, 3, new_curr)
                            break
                except Exception as e:
                    print(f"Trips sheet update error: {e}")
                reply = f"Switched to {new_dest} — expenses will now be logged in {new_curr}."
            else:
                reply = handle_overseas_request(text)
        else:
            reply = handle_overseas_request(text)
    elif is_log_prefix_input(text):
        # "log [merchant] [amount]" — always expense flow, strip the "log " prefix
        log_text = text[4:].strip()
        reply, needs_session, session_data = handle_expense_text(log_text, user_id)
        if needs_session and session_data:
            expense_sessions[user_id] = session_data

    elif lower in ["bills", "my bills", "list bills", "bills due", "what bills do i have",
                   "upcoming bills", "show bills"]:
        reply = list_bills()
    elif lower.startswith("delete bill "):
        bill_name = text[12:].strip()
        confirm_sessions[user_id] = {"action": "delete_bill", "args": [bill_name], "target": f"Delete bill {bill_name}?"}
        reply = f"Delete bill *{bill_name}*? (yes / no)"
    elif is_bill_request(text):
        reply = handle_new_bill(text)

    # Restaurant Commands
    elif lower.startswith("delete restaurant ") or lower.startswith("remove restaurant "):
        rest_name = text.split(" ", 2)[2].strip()
        confirm_sessions[user_id] = {"action": "delete_restaurant", "args": [rest_name], "target": f"Delete restaurant {rest_name}?"}
        reply = f"Delete *{rest_name}* from your restaurant list? (yes / no)"
    elif is_restaurant_search(text):
        reply = handle_search_restaurants(text)
    elif is_restaurant_save(text):
        result = handle_save_restaurant(text)
        if result and result.startswith("_NEEDS_LOCATION_:"):
            parts = result.split(":", 2)
            name = parts[1] if len(parts) > 1 else ""
            country = parts[2] if len(parts) > 2 else "Singapore"
            pending_restaurant_saves[user_id] = {"name": name, "country": country, "step": "awaiting_location"}
            reply = f"⚠️ Couldn't read that Maps link — what's the location for {name}?\n(e.g. 313 Orchard Road, Singapore)"
        elif result and result.startswith("_INFER_LOCATION_:"):
            parts = result.split(":", 5)
            name = parts[1] if len(parts) > 1 else ""
            country = parts[2] if len(parts) > 2 else "Singapore"
            tags = parts[3] if len(parts) > 3 else ""
            multiple = parts[4] == "1" if len(parts) > 4 else False
            outlets = parts[5].split("|") if len(parts) > 5 and parts[5] else []
            if multiple and len(outlets) > 1:
                pending_restaurant_saves[user_id] = {
                    "step": "awaiting_outlet", "name": name, "country": country,
                    "tags": tags, "outlets": outlets
                }
                outlet_list = "\n".join(f"{i+1}. {o}" for i, o in enumerate(outlets))
                reply = f"Found a few {name} outlets — which one, and any tags?\n{outlet_list}"
            else:
                location = outlets[0] if outlets else ""
                pending_restaurant_saves[user_id] = {
                    "step": "awaiting_confirm", "name": name, "country": country,
                    "tags": tags, "location": location
                }
                reply = f"Got it — I have {name} at {location}. Is that right, and any tags? (yes / no, add tags or skip)"
        elif result and result.startswith("_DUPLICATE_RESTAURANT_:"):
            parts = result.split(":", 5)
            pending_restaurant_saves[user_id] = {
                "step": "duplicate", "existing": parts[1], "name": parts[2],
                "location": parts[3], "country": parts[4],
                "tags": parts[5] if len(parts) > 5 else "", "notes": ""
            }
            reply = f"*{parts[1]}* is already in your list.\nUpdate existing or save as new? (update / new)"
        else:
            reply = result
    elif lower.startswith("search restaurants "):
        reply = search_restaurants(text[19:].strip())
    elif is_restaurant_suggestion_request(text):
        reply = get_similar_restaurants(text)
    elif is_restaurant_review_request(text):
        # Extract restaurant name — strip review trigger words
        name = re.sub(r"reviews? for|review of|tell me about|how is|how's|what's|what is", "", lower).strip().rstrip("?").strip()
        reply = get_restaurant_review(name) if name else "Which restaurant are you asking about?"

    # Stock Commands
    elif lower in ["portfolio", "my portfolio", "holdings", "portfolio performance"]:
        reply = get_portfolio_performance()
    elif lower in ["delete from portfolio", "remove from portfolio", "delete holding", "remove holding",
                   "portfolio delete", "clear holding"]:
        rows = get_portfolio_rows()
        if not rows:
            reply = "Nothing in your portfolio to remove."
        else:
            portfolio_delete_sessions[user_id] = {"step": "pick", "rows": rows}
            reply = format_portfolio_delete_list(rows)
    elif lower.startswith("delete portfolio ") or lower.startswith("remove portfolio "):
        query = re.sub(r"^(delete|remove) portfolio\s+", "", lower).strip().upper()
        results = search_portfolio_by_ticker(query)
        if not results:
            reply = f"No holdings found matching '{query}'."
        elif len(results) == 1:
            sheet_row, _ = results[0]
            reply = delete_portfolio_row(sheet_row)
        else:
            portfolio_delete_sessions[user_id] = {"step": "pick", "rows": results}
            reply = format_portfolio_delete_list(results)
    elif lower in ["market summary", "market today", "how is the market", "market update",
                   "weekly market summary", "how are markets"]:
        reply = get_market_summary_now()
    elif is_stock_request(text):
        result = handle_stock_request(text)
        if result:
            reply = result

    # Todo Commands
    elif any(lower.startswith(p) for p in [
        "todo ", "add task ", "new task ", "create todo ",
        "add todo ", "add to my list ", "add to my todo "
    ]):
        for p in ["todo ", "add task ", "new task ", "create todo ", "add todo ", "add to my list ", "add to my todo "]:
            if lower.startswith(p):
                reply = add_todo(text[len(p):])
                break
    elif lower.startswith("done "):
        result = complete_todo(text[5:])
        if result.startswith("_DISAMBIG_TODO_COMPLETE_:"):
            tasks = result.split(":", 1)[1].split("|")
            todo_disambig_sessions[user_id] = {"tasks": tasks, "action": "complete"}
            lines = ["Found multiple matching tasks — which one?"]
            for i, t in enumerate(tasks, 1):
                lines.append(f"{i}. {t}")
            lines.append("\nReply with the number.")
            reply = "\n".join(lines)
        else:
            reply = result
    elif lower.startswith("delete todo ") or lower.startswith("remove todo "):
        task_name = re.sub(r"^(delete|remove) todo\s+", "", lower).strip()
        result = delete_todo(task_name)
        if result.startswith("_DISAMBIG_TODO_DELETE_:"):
            tasks = result.split(":", 1)[1].split("|")
            todo_disambig_sessions[user_id] = {"tasks": tasks, "action": "delete"}
            lines = ["Found multiple matching tasks — which one?"]
            for i, t in enumerate(tasks, 1):
                lines.append(f"{i}. {t}")
            lines.append("\nReply with the number.")
            reply = "\n".join(lines)
        else:
            reply = result
    elif lower == "todos":
        reply = list_todos()

    # Calendar Commands
    elif lower.startswith("add event") or lower.startswith("schedule ") or lower.startswith("create event"):
        reply = smart_add_event(text, user_id)
    elif lower == "events today":
        reply = get_events(1)
    elif lower == "events week":
        reply = get_events(7)
    elif lower.startswith("delete event "):
        event_name = text[13:].strip()
        confirm_sessions[user_id] = {"action": "delete_event", "args": [event_name], "target": f"Delete event {event_name}?"}
        reply = f"Delete event *{event_name}*? (yes / no)"

    # Infrastructure / Settings
    elif lower in ("em whats pending", "em what's pending", "em pending", "whats pending"):
        reply = get_pending_backlog()

    elif lower == "em status":
        issues = []
        # Google Drive
        if not DRIVE_FOLDERS:
            issues.append("• Google Drive: not connected — check GOOGLE_CREDENTIALS in Railway")
        # Sheets
        try:
            spreadsheet.worksheet("Expenses")
        except Exception as e:
            issues.append(f"• Sheets: connection error — {str(e)[:60]}")
        # Scheduler
        if not _scheduler or not _scheduler.running:
            issues.append("• Scheduler: not running — reminders and scheduled jobs are down, restart the bot")
        # Profile
        if not em_profile or not em_profile.get("version"):
            issues.append("• Profile: not loaded — reply 'reload profile' to retry")
        # iCloud
        try:
            get_calendar("Personal")
        except Exception as e:
            issues.append(f"• iCloud Calendar: unreachable — check ICLOUD_USERNAME / ICLOUD_PASSWORD in Railway")
        # Anthropic API
        if _anthropic_failure_count >= ANTHROPIC_FAILURE_THRESHOLD:
            issues.append("• Anthropic API: repeated failures detected — check API key or Anthropic status page")

        if not issues:
            reply = "✅ Systems all green"
        else:
            lines = ["⚠️ Issues detected:\n"] + issues
            reply = "\n".join(lines)

    # Help
    elif lower == "help":
        reply = (
            "🤖 *Em — here's what I can do:*\n\n"
            "*CRM:*\n"
            "save, find, note, followup, update, edit, delete, search, list, stats, followups, overdue, birthdays, soon, lastcontact\n"
            "referrals, all referrals, top referrers, referrals from [name]\n"
            "import excel — import contacts from a spreadsheet\n\n"
            "*Calendar:*\n"
            "Just tell me naturally — 'schedule dinner tomorrow 7pm' or 'add event'\n"
            "events today / events week / delete event\n\n"
            "*To-Do:*\n"
            "todo [task] / add task [task] / new task [task]\n"
            "done [task] — mark complete\n"
            "todos — list all\n\n"
            "*Expenses:*\n"
            "log [merchant] [amount] — or just send a receipt photo\n"
            "last expense / edit last expense / delete last expense\n"
            "monthly summary / expense categories\n\n"
            "*Reminders:*\n"
            "remind me to [task] at [time] — or 'don't let me forget to [task]'\n"
            "reminders — list pending\n"
            "cancel reminder [keyword]\n\n"
            "*Bills:*\n"
            "add bill / bills due / delete bill\n\n"
            "*Stocks:*\n"
            "how is [ticker] doing / price of [ticker] / check [ticker]\n"
            "my portfolio / add to portfolio / market summary\n"
            "alert me if [ticker] hits [price]\n\n"
            "*Restaurants:*\n"
            "save restaurant [name] / restaurants / review [name]\n"
            "suggest restaurant [cuisine/area]\n\n"
            "*Trips & Overseas:*\n"
            "now in [country] — activate overseas mode\n"
            "flying [flight number] — log a flight\n"
            "trip summary / close trip\n\n"
            "*Meeting Recap:*\n"
            "meeting recap — then send your notes\n\n"
            "*Other:*\n"
            "em status — check Em's health\n"
            "dnd on / dnd off — do not disturb\n\n"
            "Or just chat — I'll figure it out 👍"
        )

    # Natural language CRM — before fuzzy detectors (phrases like "John referred Sarah")
    elif (crm_action := detect_crm_natural_update(text)):
        action, name, field_or_referred, value = crm_action
        if action == "referral":
            reply = set_referral(name, field_or_referred)
        elif action == "update":
            reply = update_contact_field_natural(name, field_or_referred, value)
        elif action == "show_private":
            reply = find_contact(name, show_private=True)

    # Overseas / trip setup — flight number guard before expense parser
    elif is_overseas_mode_request(text) or extract_flight_number(text):
        reply = handle_overseas_request(text)

    # Expense — flight guard already in is_expense_input, but extract_flight_number above catches first
    elif is_expense_input(text):
        reply, needs_session, session_data = handle_expense_text(text, user_id)
        if needs_session and session_data:
            expense_sessions[user_id] = session_data
    elif is_bare_merchant_input(text):
        reply, needs_session, session_data = handle_expense_text(text, user_id)
        if needs_session and session_data:
            expense_sessions[user_id] = session_data

    # Calendar
    elif is_calendar_request(text):
        reply = smart_add_event(text, user_id)

    # Claude conversation fallback — genuine unknowns only, single path
    else:
        if user_id not in conversation_histories:
            conversation_histories[user_id] = []
        conversation_histories[user_id].append({"role": "user", "content": text})
        if len(conversation_histories[user_id]) > 20:
            conversation_histories[user_id] = conversation_histories[user_id][-20:]
        global _system_prompt_cache, _system_prompt_overseas_key
        overseas_key = (
            overseas_state.get("active"),
            overseas_state.get("destination"),
            overseas_state.get("currency"),
        )
        if _system_prompt_cache is None or _system_prompt_overseas_key != overseas_key:
            _system_prompt_cache = build_system_prompt()
            _system_prompt_overseas_key = overseas_key
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            system=_system_prompt_cache,
            messages=conversation_histories[user_id]
        )
        reply = response.content[0].text
        conversation_histories[user_id].append({"role": "assistant", "content": reply})

    if not reply:
        reply = "Not sure what you mean — try rephrasing, or say 'help' to see what I can do."

    if reply:
        await send_safe(update.message, reply, parse_mode="Markdown")

# --- Main ---

async def check_missed_items_on_startup(app):
    """Check for missed follow-ups, bill reminders, reminders, and market summary while offline."""
    try:
        today = date.today()
        yesterday = today - timedelta(days=1)
        missed_followups = []
        missed_bills = []
        missed_reminders = []

        # Missed follow-ups
        try:
            records = _get_crm_records()
            for r in records:
                fu_date_str = r.get("Follow Up Date", "")
                if fu_date_str:
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                        try:
                            fu_date = datetime.strptime(fu_date_str, fmt).date()
                            if fu_date < today:
                                missed_followups.append({
                                    "name": r.get("Name", "?"),
                                    "date": fu_date_str,
                                    "notes": r.get("Follow Up Notes", "")
                                })
                            break
                        except ValueError:
                            continue
        except Exception as e:
            print(f"Missed followups check error: {e}")

        # Missed bill reminders
        try:
            ws = bills_sheet()
            records = ws.get_all_records()
            for r in records:
                due_day = int(r.get("Due Date", 0) or 0)
                if due_day and yesterday.day == due_day:
                    missed_bills.append(r.get("Name", "?"))
        except Exception as e:
            print(f"Missed bills check error: {e}")

        # Missed reminders
        try:
            ws = reminders_sheet()
            records = ws.get_all_records()
            now = datetime.now(TIMEZONE)
            for i, r in enumerate(records, start=2):
                if r.get("Status") != "pending":
                    continue
                scheduled_str = r.get("Scheduled Time", "")
                if not scheduled_str:
                    continue
                try:
                    scheduled = TIMEZONE.localize(datetime.strptime(scheduled_str, "%Y-%m-%d %H:%M"))
                    if scheduled < now:
                        missed_reminders.append({
                            "message": r.get("Message", ""),
                            "row": i,
                            "recurrence": r.get("Recurrence", "once")
                        })
                except Exception:
                    continue
        except Exception as e:
            print(f"Missed reminders check error: {e}")

        # Send missed follow-ups notification
        if missed_followups:
            lines = ["⚠️ Missed follow-ups while offline:"]
            for f in missed_followups:
                lines.append(f"• {f['name']} — was due {f['date']}")
            lines.append("\nReply 'followups' for details, 'skip missed' to dismiss all, or 'skip [name]' for one.")
            await app.bot.send_message(chat_id=YOUR_CHAT_ID, text="\n".join(lines))

        # Send missed bills notification
        if missed_bills:
            lines = ["⚠️ Missed bill reminder(s) while offline:"]
            for b in missed_bills:
                lines.append(f"• {b}")
            lines.append("\nReply 'skip missed bills' to dismiss.")
            await app.bot.send_message(chat_id=YOUR_CHAT_ID, text="\n".join(lines))

        # Fire missed reminders immediately
        for rem in missed_reminders:
            try:
                await app.bot.send_message(
                    chat_id=YOUR_CHAT_ID,
                    text=f"🔔 Reminder (missed while offline): {rem['message']}"
                )
                # Update recurrence or mark sent
                ws = reminders_sheet()
                if rem["recurrence"] != "once":
                    next_time = get_next_recurrence(
                        datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M"),
                        rem["recurrence"]
                    )
                    if next_time:
                        ws.update_cell(rem["row"], 3, next_time)
                else:
                    ws.update_cell(rem["row"], 5, "sent")
            except Exception as e:
                print(f"Missed reminder fire error: {e}")

        # Check missed Monday market summary — only ask if not already sent today
        try:
            if today.weekday() == 0:  # Monday
                already_sent = False
                try:
                    sheet = get_sheet("Settings")
                    if sheet:
                        records = sheet.get_all_records()
                        for r in records:
                            if r.get("Key") == "market_summary_last_sent":
                                already_sent = r.get("Value", "") == today.isoformat()
                                break
                except Exception as e:
                    print(f"market_summary_last_sent check error: {e}")

                if not already_sent:
                    await app.bot.send_message(
                        chat_id=YOUR_CHAT_ID,
                        text="Missed the Monday market summary while offline — want me to send it now? (yes / no)"
                    )
                    market_summary_pending[YOUR_CHAT_ID] = True
        except Exception as e:
            print(f"Missed market summary check error: {e}")

    except Exception as e:
        print(f"check_missed_items_on_startup error: {e}")

async def post_init(app):
    global _scheduler, _app_ref
    # Run infrastructure setup and capture health status
    health = run_infrastructure_setup()

    # Warm card names cache on startup
    try:
        get_cards_live()
    except Exception as e:
        print(f"Startup: card cache warm failed: {e}")

    # Restore overseas mode if there was an active trip before restart
    restore_overseas_from_trips()

    # Restore cached FX rates from Settings sheet
    load_fx_rates_from_sheet()

    # Restore price alerts from Settings sheet
    load_price_alerts_from_sheet()

    # Restore pending expense sessions from Settings sheet
    load_sessions_from_sheet()

    # Restore birthday pending from Settings sheet
    load_birthday_pending_from_sheet()

    # Restore trip setup state from Settings sheet
    load_trip_setup_from_sheet()

    timezone = pytz.timezone("Asia/Kuala_Lumpur")
    scheduler = AsyncIOScheduler(timezone=timezone, misfire_grace_time=30)
    _scheduler = scheduler
    _app_ref = app

    # Follow-up reminders at 9am
    scheduler.add_job(send_followup_reminders, "cron", hour=9, minute=0, args=[app])

    # Birthday greetings at 12pm
    scheduler.add_job(send_birthday_reminders, "cron", hour=12, minute=0, args=[app])

    # Birthday 2pm follow-up
    scheduler.add_job(send_birthday_followups, "cron", hour=14, minute=0, args=[app])

    # Custom reminders — check every minute
    scheduler.add_job(check_and_fire_reminders, "interval", minutes=1, args=[app], misfire_grace_time=30)

    # Bill reminders — daily at 9am
    scheduler.add_job(send_bill_reminders, "cron", hour=9, minute=0, args=[app])
    scheduler.add_job(check_icloud_daily, "cron", hour=9, minute=5, args=[app])

    # Price alerts — check every 15 minutes
    scheduler.add_job(check_price_alerts, "interval", minutes=15, args=[app], misfire_grace_time=30)

    # Weekly market summary — Monday 8am
    scheduler.add_job(send_weekly_market_summary, "cron", day_of_week="mon", hour=8, minute=0, args=[app])

    # FX rate refresh — 8am and 8pm daily
    scheduler.add_job(refresh_fx_rates, "cron", hour=8, minute=0)
    scheduler.add_job(refresh_fx_rates, "cron", hour=20, minute=0)

    scheduler.start()
    health["Scheduler"] = "✅ Running"
    print("✅ Scheduler started — follow-ups + bills at 9am, birthdays at 12pm + 2pm, reminders every minute, market Monday 8am")

    # Send startup message if new deployment
    await send_startup_message(app, health)

    # Check for missed items while offline
    await check_missed_items_on_startup(app)

def main():
    # Guard: fail clearly if critical env vars are missing
    missing = [v for v in ("TELEGRAM_TOKEN", "ANTHROPIC_API_KEY", "SHEET_ID") if not os.getenv(v)]
    if missing:
        print(f"❌ Missing required environment variables: {', '.join(missing)}")
        print("Set these in Railway → Variables before deploying.")
        raise SystemExit(1)

    app = ApplicationBuilder().token(TELEGRAM_TOKEN).post_init(post_init).build()
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_message))
    app.add_handler(MessageHandler(filters.PHOTO, handle_message))
    print("Em is running... Press Ctrl+C to stop.")
    app.run_polling()

if __name__ == "__main__":
    main()
