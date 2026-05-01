import re
import json
import asyncio
import httpx
from datetime import date, datetime
import state
from config import (
    MARKET_INDICES, MARKET_FLAGS_MAP, SGX_TICKER_MAP, HK_TICKER_MAP,
    SOURCE_LABELS, _OBSCURE_SOURCES, _HEADLINE_REJECT,
    ALPHA_VANTAGE_API_KEY, YOUR_CHAT_ID
)
from clients import client
from sheets import portfolio_sheet, get_sheet
from helpers import get_source_label

def persist_price_alerts_to_sheet():
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        data = json.dumps({k: v for k, v in state.price_alerts.items() if v.get("active")})
        for i, r in enumerate(records):
            if r.get("Key") == "price_alerts":
                sheet.update_cell(i + 2, 2, data)
                return
        sheet.append_row(["price_alerts", data])
    except Exception as e:
        print(f"persist_price_alerts_to_sheet error: {e}")

def load_price_alerts_from_sheet():
    try:
        sheet = get_sheet("Settings")
        if not sheet:
            return
        records = sheet.get_all_records()
        for r in records:
            if r.get("Key") == "price_alerts":
                raw = r.get("Value", "")
                if raw:
                    loaded = json.loads(raw)
                    state.price_alerts.update(loaded)
                    print(f"Restored {len(loaded)} price alert(s) from sheet")
                return
    except Exception as e:
        print(f"load_price_alerts_from_sheet error: {e}")

def normalise_ticker(ticker):
    t = ticker.strip().lower()
    if t in SGX_TICKER_MAP:
        return SGX_TICKER_MAP[t]
    if t in HK_TICKER_MAP:
        return HK_TICKER_MAP[t]
    upper = ticker.strip().upper()
    if upper in HK_TICKER_MAP:
        return HK_TICKER_MAP[upper]
    if upper in SGX_TICKER_MAP:
        return SGX_TICKER_MAP[upper]
    if "." in upper:
        return upper
    if re.match(r"^\d{4}$", upper):
        return f"{upper}.HK"
    return upper

async def fetch_weekly_change(ticker):
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1wk&range=1mo"
        async with httpx.AsyncClient(timeout=10, headers={"User-Agent": "Mozilla/5.0"}) as hx:
            resp = await hx.get(url)
        data = resp.json()
        result = data["chart"]["result"][0]
        closes = result["indicators"]["quote"][0].get("close", [])
        opens = result["indicators"]["quote"][0].get("open", [])
        pairs = [(o, c) for o, c in zip(opens, closes) if o is not None and c is not None]
        if len(pairs) >= 2:
            week_open, week_close = pairs[-2]
            if week_open:
                pct = (week_close - week_open) / week_open * 100
                return pct
    except Exception as e:
        print(f"fetch_weekly_change error for {ticker}: {e}")
    return None

async def fetch_price(ticker):
    ticker = normalise_ticker(ticker)
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=1y"
        async with httpx.AsyncClient(timeout=10, headers={"User-Agent": "Mozilla/5.0"}) as hx:
            resp = await hx.get(url)
        data = resp.json()
        result = data["chart"]["result"][0]
        meta = result["meta"]
        price = meta.get("regularMarketPrice", 0)
        prev_close = meta.get("chartPreviousClose", 0)
        currency = meta.get("currency", "USD")
        name = meta.get("longName") or meta.get("shortName") or ticker
        change = price - prev_close
        change_pct = (change / prev_close * 100) if prev_close else 0
        market_state = meta.get("marketState", "")
        exchange = meta.get("exchange", "")
        full_exchange = meta.get("fullExchangeName", "")
        closes = [c for c in result["indicators"]["quote"][0].get("close", []) if c is not None]
        week52_low = min(closes) if closes else None
        week52_high = max(closes) if closes else None
        suffix = ticker.split(".")[-1] if "." in ticker else ""
        exchange_flags = {
            "SI": "🇸🇬", "HK": "🇭🇰", "L": "🇬🇧", "AX": "🇦🇺",
            "T": "🇯🇵", "NS": "🇮🇳", "BO": "🇮🇳", "SS": "🇨🇳", "SZ": "🇨🇳",
        }
        flag = exchange_flags.get(suffix, "🇺🇸")
        return {
            "ticker": ticker, "name": name, "price": price, "prev_close": prev_close,
            "change": change, "change_pct": change_pct, "currency": currency,
            "market_state": market_state, "exchange": exchange,
            "fullExchangeName": full_exchange, "week52_low": week52_low,
            "week52_high": week52_high, "flag": flag,
        }
    except Exception as e:
        print(f"Yahoo Finance error for {ticker}: {e}")
    if ALPHA_VANTAGE_API_KEY:
        try:
            url = (
                f"https://www.alphavantage.co/query"
                f"?function=GLOBAL_QUOTE&symbol={ticker}&apikey={ALPHA_VANTAGE_API_KEY}"
            )
            async with httpx.AsyncClient(timeout=10) as hx:
                resp = await hx.get(url)
            data = resp.json()
            quote = data.get("Global Quote", {})
            if quote.get("05. price"):
                price = float(quote["05. price"])
                prev_close = float(quote["08. previous close"])
                change = float(quote["09. change"])
                change_pct = float(quote["10. change percent"].replace("%", ""))
                av_suffix = ticker.split(".")[-1] if "." in ticker else ""
                av_exchange_flags = {
                    "SI": "🇸🇬", "HK": "🇭🇰", "L": "🇬🇧", "AX": "🇦🇺",
                    "T": "🇯🇵", "NS": "🇮🇳", "BO": "🇮🇳", "SS": "🇨🇳", "SZ": "🇨🇳",
                }
                return {
                    "ticker": ticker, "name": ticker, "price": price,
                    "prev_close": prev_close, "change": change, "change_pct": change_pct,
                    "currency": "USD", "flag": av_exchange_flags.get(av_suffix, "🇺🇸"),
                }
        except Exception as e:
            print(f"Alpha Vantage fallback error for {ticker}: {e}")
    return None

async def _fetch_rss_headlines_for_stock(ticker, name):
    import xml.etree.ElementTree as ET
    gn_query = f"{name} stock".replace(" ", "+")
    yf_query = f"{ticker} stock".replace(" ", "+")
    rss_sources = [
        f"https://news.google.com/rss/search?q={gn_query}&hl=en&gl=US&ceid=US:en",
        f"https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US",
    ]
    headlines = []
    sources = []
    seen = set()

    async def fetch_one(url):
        try:
            async with httpx.AsyncClient(timeout=3) as hx:
                resp = await hx.get(url)
            root = ET.fromstring(resp.content)
            results = []
            for item in root.findall(".//item"):
                title = item.findtext("title", "")
                source = item.findtext("source", "") or title.split(" - ")[-1].strip()
                title = title.split(" - ")[0].strip()
                if title:
                    results.append((title, source))
            return results
        except Exception as e:
            print(f"RSS fetch error {url}: {e}")
            return []

    results = await asyncio.gather(*[fetch_one(url) for url in rss_sources], return_exceptions=True)
    for result in results:
        if isinstance(result, Exception):
            continue
        for title, source in result:
            if title not in seen and len(headlines) < 4:
                seen.add(title)
                headlines.append(title)
                sources.append(source)
    return headlines[:3], sources[:3]

def _generate_price_movement_summary(data):
    price = data.get("price", 0)
    change_pct = data.get("change_pct", 0)
    week52_low = data.get("week52_low")
    week52_high = data.get("week52_high")
    currency = data.get("currency", "")
    name = data.get("name", data.get("ticker", ""))
    direction = "up" if change_pct >= 0 else "down"
    sentences = [f"{name} is {direction} {abs(change_pct):.2f}% today, currently at {currency} {price:.2f}."]
    if week52_low and week52_high:
        position = (price - week52_low) / (week52_high - week52_low) * 100 if week52_high != week52_low else 50
        if position >= 75:
            range_desc = "trading near its 52-week high"
        elif position <= 25:
            range_desc = "trading near its 52-week low"
        else:
            range_desc = "trading in the middle of its 52-week range"
        sentences.append(f"It is {range_desc} ({currency} {week52_low:.2f} – {currency} {week52_high:.2f}).")
    return " ".join(sentences)

async def fetch_stock_summary(ticker, name, price_data=None):
    headlines, sources = await _fetch_rss_headlines_for_stock(ticker, name)
    labelled = [(h, get_source_label(s)) for h, s in zip(headlines, sources)]
    usable = [(h, l) for h, l in labelled if l]
    if not usable and price_data:
        return _generate_price_movement_summary(price_data), []
    if not usable:
        return None, []
    source_context = "; ".join(f"{h[:60]} [{l}]" for h, l in usable[:3])
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{
                "role": "user",
                "content": (
                    f"Headlines about {name} ({ticker}):\n{source_context}\n\n"
                    "Write 2-3 short factual sentences as one paragraph. "
                    "First sentence: where the stock sits in its 52-week range (near high, near low, or mid-range). "
                    "Remaining sentences: recent business developments or earnings from the headlines — factual only. "
                    "No timing advice, no buy/sell signals, no speculation, no phrases like 'investors should' or 'right time to buy'. "
                    "End with a single [SourceName] tag for the most relevant source. No source tag if no usable headlines."
                )
            }]
        )
        summary = resp.content[0].text.strip()
        return summary, [l for _, l in usable[:2]]
    except Exception as e:
        print(f"Stock summary Claude error: {e}")
        if price_data:
            return _generate_price_movement_summary(price_data), []
        return None, []

async def format_price(data, summary=None):
    if not data:
        return None
    flag = data.get("flag", "🌐")
    name = data.get("name", data["ticker"])
    ticker = data["ticker"]
    currency = data.get("currency", "")
    price = data.get("price", 0)
    change_pct = data.get("change_pct", 0)
    arrow = "▲" if change_pct >= 0 else "▼"
    market_state = data.get("market_state", "REGULAR")
    state_label = {"PRE": " (pre)", "POST": " (post)", "CLOSED": " (closed)", "REGULAR": ""}.get(market_state, "")
    week52_low = data.get("week52_low")
    week52_high = data.get("week52_high")
    range_line = f"52-week range: {currency} {week52_low:.2f} – {currency} {week52_high:.2f}" if week52_low and week52_high else ""
    if summary is None:
        summary, _ = await fetch_stock_summary(ticker, name, price_data=data)
    lines = [f"{flag} {name} ({ticker})"]
    lines.append(f"{currency} {price:.2f} {arrow} {abs(change_pct):.2f}%{state_label}")
    if range_line:
        lines.append(range_line)
    lines.append("")
    lines.append(f"{summary if summary else _generate_price_movement_summary(data)}")
    return "\n".join(lines)

def log_portfolio_buy(ticker, quantity, price, buy_date=None):
    try:
        sheet = portfolio_sheet()
        today = buy_date or date.today().strftime("%d/%m/%Y")
        sheet.append_row([ticker.upper(), str(quantity), str(price), today, ""])
        return True
    except Exception as e:
        print(f"log_portfolio_buy error for {ticker}: {e}")
        return False

def get_portfolio_holdings():
    try:
        sheet = portfolio_sheet()
        records = sheet.get_all_records()
        holdings = {}
        for r in records:
            ticker = r.get("Stock", "").upper()
            if not ticker:
                continue
            try:
                qty = float(r.get("Quantity", 0))
                price = float(r.get("Buy Price", 0))
            except (ValueError, TypeError):
                continue
            if ticker not in holdings:
                holdings[ticker] = {"total_qty": 0, "total_cost": 0}
            holdings[ticker]["total_qty"] += qty
            holdings[ticker]["total_cost"] += qty * price
        result = {}
        for ticker, h in holdings.items():
            if h["total_qty"] > 0:
                result[ticker] = {"qty": h["total_qty"], "avg_cost": h["total_cost"] / h["total_qty"]}
        return result
    except Exception as e:
        print(f"Error getting portfolio: {e}")
        return {}

async def get_portfolio_performance():
    holdings = get_portfolio_holdings()
    if not holdings:
        return "No holdings logged yet."
    lines = ["Portfolio:\n"]
    total_cost = 0
    total_value = 0
    for ticker, h in holdings.items():
        data = await fetch_price(ticker)
        avg = h["avg_cost"]
        qty = h["qty"]
        cost = avg * qty
        if data:
            current = data["price"]
            value = current * qty
            pnl = value - cost
            pnl_pct = (pnl / cost * 100) if cost else 0
            sign = "+" if pnl >= 0 else ""
            flag = "✅" if pnl >= 0 else "⚠️"
            lines.append(
                f"{flag} {ticker}: {qty:.0f} shares @ avg {data['currency']} {avg:.2f} "
                f"| now {data['currency']} {current:.2f} | {sign}{pnl_pct:.1f}% ({sign}{data['currency']} {pnl:.2f})"
            )
            total_cost += cost
            total_value += value
        else:
            lines.append(f"• {ticker}: {qty:.0f} shares @ avg {avg:.2f} (price unavailable)")
            total_cost += cost
    if total_cost > 0:
        total_pnl = total_value - total_cost
        total_pct = (total_pnl / total_cost * 100)
        sign = "+" if total_pnl >= 0 else ""
        lines.append(f"\nTotal P&L: {sign}{total_pct:.1f}% ({sign}${total_pnl:.2f})")
    return "\n".join(lines)

def get_portfolio_rows():
    try:
        ws = portfolio_sheet()
        records = ws.get_all_records()
        return [(i + 2, r) for i, r in enumerate(records)]
    except Exception as e:
        print(f"get_portfolio_rows error: {e}")
        return []

def format_portfolio_delete_list(rows):
    if not rows:
        return "No holdings in portfolio."
    lines = ["Which holding do you want to remove?\n"]
    for i, (_, r) in enumerate(rows, 1):
        ticker = r.get("Stock", "?")
        qty = r.get("Quantity", "?")
        price = r.get("Buy Price", "?")
        buy_date = r.get("Buy Date", "")
        lines.append(f"{i}. {ticker} — {qty} shares @ ${price} ({buy_date})")
    lines.append("\nReply with a number, 'search [ticker]', or 'cancel'.")
    return "\n".join(lines)

def delete_portfolio_row(sheet_row):
    try:
        ws = portfolio_sheet()
        all_vals = ws.get_all_values()
        if sheet_row - 1 < len(all_vals):
            row_data = all_vals[sheet_row - 1]
            ticker = row_data[0] if row_data else "?"
        ws.delete_rows(sheet_row)
        return f"Removed {ticker} from portfolio."
    except Exception as e:
        return f"Couldn't remove that holding: {str(e)}"

def search_portfolio_by_ticker(query):
    rows = get_portfolio_rows()
    q = query.upper()
    return [(row_idx, r) for row_idx, r in rows if q in r.get("Stock", "").upper()]

def set_price_alert(ticker, condition, price):
    ticker = ticker.upper()
    state.price_alerts[ticker] = {"condition": condition, "price": price, "active": True}
    persist_price_alerts_to_sheet()

def parse_stock_request(text):
    prompt = (
        f"Parse this stock market request: '{text}'\n\n"
        f"Return ONLY a JSON object with:\n"
        f"- intent: string (one of: price_check, set_alert, portfolio_add, portfolio_view, "
        f"stock_suggest, market_summary, price_alert_check)\n"
        f"- ticker: string (stock ticker symbol, uppercase, or empty)\n"
        f"- quantity: number (shares, or 0)\n"
        f"- price: number (price per share, or 0)\n"
        f"- alert_condition: string (above or below, or empty)\n"
        f"- alert_price: number (alert trigger price, or 0)\n"
        f"- criteria: string (for stock suggestions — describe what user wants)\n\n"
        f"Return ONLY the JSON."
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, ValueError) as e:
        print(f"parse_stock_request JSON error: {e} | raw: {raw[:100]}")
        return None
    if parsed.get("ticker"):
        parsed["ticker"] = normalise_ticker(parsed["ticker"])
    return parsed

def suggest_stocks(criteria):
    prompt = (
        f"Suggest 3 stocks based on this criteria: '{criteria}'\n\n"
        f"For each stock provide a qualitative summary. Flag concerns with a warning, all clear with a checkmark.\n\n"
        f"Format your response exactly like this for each stock (with a divider line between each):\n\n"
        f"TICKER — Company Name\n"
        f"[checkmark or warning] [2-3 sentence qualitative summary. No numbers unless asked.]\n\n"
        f"---\n\n"
        f"Keep it concise and honest. Flag anything concerning."
    )
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text.strip()

async def check_price_alerts(app):
    try:
        for ticker, alert in list(state.price_alerts.items()):
            if not alert.get("active"):
                continue
            data = await fetch_price(ticker)
            if not data:
                continue
            current = data["price"]
            condition = alert["condition"]
            trigger = alert["price"]
            triggered = (condition == "below" and current <= trigger) or \
                        (condition == "above" and current >= trigger)
            if triggered:
                msg = (
                    f"🔔 Price alert: {ticker} is now {data['currency']} {current:.2f} "
                    f"({condition} your target of {data['currency']} {trigger:.2f})"
                )
                await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg)
                state.price_alerts[ticker]["active"] = False
                persist_price_alerts_to_sheet()
    except Exception as e:
        print(f"Error checking price alerts: {e}")

async def send_weekly_market_summary(app):
    try:
        msg = await get_market_summary_now()
        await app.bot.send_message(chat_id=YOUR_CHAT_ID, text=msg, parse_mode="Markdown")
        try:
            sheet = get_sheet("Settings")
            if sheet:
                records = sheet.get_all_records()
                today_str = date.today().isoformat()
                for i, r in enumerate(records):
                    if r.get("Key") == "market_summary_last_sent":
                        sheet.update_cell(i + 2, 2, today_str)
                        break
                else:
                    sheet.append_row(["market_summary_last_sent", today_str])
        except Exception as e:
            print(f"market_summary_last_sent persist error: {e}")
    except Exception as e:
        print(f"Error sending weekly market summary: {e}")

def is_stock_request(text):
    lower = text.lower()
    exclusions = ["reminders", "reminder", "my bill", "credit card bill"]
    if any(e in lower for e in exclusions):
        return False
    share_patterns = [
        r"bought \d+ shares", r"sold \d+ shares", r"shares of [a-z]",
        r"shares @ ", r"shares at \$", r"\d+ shares"
    ]
    if any(re.search(p, lower) for p in share_patterns):
        return True
    explicit_triggers = [
        "pull up ", "look into ", "price of ",
        "alert me if", "alert if ", "add to portfolio",
        "suggest stocks", "stock ideas",
        "stock ", "ticker ", "p&l", "holdings",
        "how is the market", "market today",
        "weekly market", "how are markets", "portfolio performance",
        "my portfolio", "portfolio",
    ]
    if any(t in lower for t in explicit_triggers):
        return True
    reminder_prefixes = ["remind me", "notify me", "ping me", "alert me when", "alert me to", "send me"]
    if "market summary" in lower and not any(lower.startswith(p) or lower[:25].startswith(p) for p in reminder_prefixes):
        return True
    reminder_prefix = any(lower.startswith(p) or p + " " in lower[:20]
                          for p in ["ping me", "notify me", "remind me", "alert me when", "alert me to"])
    if "check " in lower and not any(e in lower for e in ["reminders", "reminder", "bill"]) and not reminder_prefix:
        return True
    if (("bought " in lower or "sold " in lower) and any(
        w in lower for w in ["shares", "stock", "equity", "position", "portfolio", "aapl", "tsla"]
    )):
        return True
    if re.search(r"what'?s\s+\S+\s+at\b", lower):
        return True
    if re.search(r"what is\s+\S+\s+at\b", lower):
        return True
    if re.search(r"how much is\s+\S+\s+(trading|at|worth)\b", lower):
        return True
    ticker_match = re.search(r'\b[A-Z]{2,5}\b', text)
    if ticker_match and any(w in lower for w in ["doing", "worth", "performing", "price", "target", "outlook"]):
        return True
    return False

async def handle_stock_request(text):
    try:
        parsed = parse_stock_request(text)
        if not parsed:
            return "Couldn't parse that stock request — try again with a ticker or clearer intent."
        intent = parsed.get("intent", "")
        ticker = parsed.get("ticker", "").upper()
        if intent == "price_check" and ticker:
            data, (rss_headlines, rss_sources) = await asyncio.gather(
                fetch_price(ticker),
                _fetch_rss_headlines_for_stock(ticker, ticker)
            )
            if data:
                exchange = data.get("exchange", "") or data.get("fullExchangeName", "")
                if "OTC" in exchange.upper() and "." not in normalise_ticker(ticker):
                    return (
                        f"⚠️ {ticker} appears to be an OTC/pink sheet listing — data may be unreliable.\n"
                        f"Try the primary exchange listing instead (e.g. 0700.HK for Tencent, 9988.HK for Alibaba)."
                    )
                name = data.get("name", ticker)
                labelled = [(h, get_source_label(s)) for h, s in zip(rss_headlines, rss_sources)]
                usable = [(h, l) for h, l in labelled if l]
                if usable:
                    source_context = "; ".join(f"{h[:60]} [{l}]" for h, l in usable[:3])
                    try:
                        resp = client.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=120,
                            messages=[{
                                "role": "user",
                                "content": (
                                    f"Headlines about {name} ({ticker}):\n{source_context}\n\n"
                                    "Write 2-3 short factual sentences as one paragraph. "
                                    "First sentence: where the stock sits in its 52-week range. "
                                    "Remaining sentences: recent business developments from headlines — factual only. "
                                    "No timing advice, no buy/sell signals, no speculation. "
                                    "End with a single [SourceName] tag for the most relevant source."
                                )
                            }]
                        )
                        summary = resp.content[0].text.strip()
                    except Exception:
                        summary = _generate_price_movement_summary(data)
                else:
                    summary = _generate_price_movement_summary(data)
                return await format_price(data, summary=summary)
            if ALPHA_VANTAGE_API_KEY:
                return f"Couldn't fetch {ticker} — Alpha Vantage and Yahoo Finance both failed. The ticker may be wrong, or try again in a moment."
            return f"Couldn't fetch {ticker}. Check the ticker and try again."
        elif intent == "set_alert" and ticker:
            condition = parsed.get("alert_condition", "below")
            alert_price = parsed.get("alert_price", 0)
            if not alert_price:
                return "What price should I alert you at? Try: 'alert me if AAPL drops below $180'"
            set_price_alert(ticker, condition, alert_price)
            return f"Alert set — I'll let you know when {ticker} goes {condition} ${alert_price:.2f}."
        elif intent == "portfolio_add" and ticker:
            qty = parsed.get("quantity", 0)
            price = parsed.get("price", 0)
            if not qty or not price:
                return "I need the quantity and price. Try: 'bought 100 AAPL at $180'"
            ok = log_portfolio_buy(ticker, qty, price)
            if ok:
                return f"Logged — {qty:.0f} {ticker} @ ${price:.2f}."
            return f"❌ Couldn't save that to the portfolio — try again."
        elif intent == "portfolio_view":
            return await get_portfolio_performance()
        elif intent == "stock_suggest":
            criteria = parsed.get("criteria", text)
            return suggest_stocks(criteria)
        elif intent == "market_summary":
            return "Pulling the latest market data, give me a sec..."
        else:
            if ticker:
                data = await fetch_price(ticker)
                if data:
                    return await format_price(data)
                return f"Couldn't find data for {ticker} — check the ticker and try again."
            return "Couldn't work out what stock you're asking about. Try 'price of AAPL' or 'what's DBS at'."
    except Exception as e:
        print(f"handle_stock_request error: {e}")
        return f"Something went wrong with that stock request — try again in a moment."

async def fetch_market_rss_headlines(market_name):
    try:
        queries = {
            "US": ["US stock market today", "Wall Street S&P 500", "Federal Reserve economy"],
            "China": ["China stock market", "Shanghai economy", "China trade economy"],
            "India": ["India Nifty stock market", "RBI India economy", "BSE Sensex"],
        }
        key = None
        for k in queries:
            if k in market_name:
                key = k
                break
        if not key:
            return []
        import xml.etree.ElementTree as ET
        headlines = []
        for q_text in queries[key]:
            if len(headlines) >= 3:
                break
            q = q_text.replace(" ", "+")
            url = f"https://news.google.com/rss/search?q={q}&hl=en-SG&gl=SG&ceid=SG:en"
            try:
                async with httpx.AsyncClient(timeout=5) as hx:
                    resp = await hx.get(url)
                root = ET.fromstring(resp.content)
                for item in root.findall(".//item"):
                    if len(headlines) >= 3:
                        break
                    title = item.findtext("title", "").split(" - ")[0].strip()
                    if not title or title in headlines:
                        continue
                    lower_title = title.lower()
                    if any(pattern in lower_title for pattern in _HEADLINE_REJECT):
                        continue
                    headlines.append(title)
            except Exception as e:
                print(f"RSS fetch error for {market_name} query '{q_text}': {e}")
                continue
        return headlines[:3]
    except Exception as e:
        print(f"fetch_market_rss_headlines error for {market_name}: {e}")
        return []

async def get_market_summary_now():
    try:
        market_data_blocks = []
        all_headlines = []
        for market, indices in MARKET_INDICES.items():
            flag = MARKET_FLAGS_MAP.get(market, "🌐")
            ticker, index_name = list(indices.items())[0]
            weekly_pct = await fetch_weekly_change(ticker)
            if weekly_pct is None:
                price_data = await fetch_price(ticker)
                weekly_pct = price_data["change_pct"] if price_data else 0
                arrow = "▲" if weekly_pct >= 0 else "▼"
                pct_str = f"Day {arrow} {abs(weekly_pct):.1f}%"
            else:
                arrow = "▲" if weekly_pct >= 0 else "▼"
                pct_str = f"Week {arrow} {abs(weekly_pct):.1f}%"
            headlines = await fetch_market_rss_headlines(market)
            all_headlines.extend(headlines)
            market_data_blocks.append({
                "market": market, "flag": flag, "index_name": index_name,
                "pct_str": pct_str, "weekly_pct": weekly_pct, "headlines": headlines,
            })
        data_summary = ""
        for b in market_data_blocks:
            data_summary += f"\n{b['flag']} {b['market']} — {b['index_name']} {b['pct_str']}\n"
            if b["headlines"]:
                data_summary += "Headlines: " + "; ".join(b["headlines"]) + "\n"
        try:
            claude_resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=500,
                messages=[{
                    "role": "user",
                    "content": (
                        f"Here is this week's market data:\n{data_summary}\n\n"
                        "For each market (US, China, India), write exactly 3 bullet points. "
                        "Each bullet = 1 short informative sentence grounded in the headlines. "
                        "No speculation, no price predictions, no clickbait. "
                        "Then write a 2-3 sentence overall summary of the week across all three markets. "
                        "Format exactly as:\n"
                        "US_BULLETS:\n• ...\n• ...\n• ...\n"
                        "CHINA_BULLETS:\n• ...\n• ...\n• ...\n"
                        "INDIA_BULLETS:\n• ...\n• ...\n• ...\n"
                        "OVERALL:\n[2-3 sentences]"
                    )
                }]
            )
            raw = claude_resp.content[0].text.strip()
        except Exception as e:
            print(f"Market summary Claude error: {e}")
            raw = ""

        def extract_section(text, key):
            import re as _re
            pattern = _re.compile(_re.escape(key) + r"(.*?)(?=\n[A-Z_]+:|$)", _re.DOTALL | _re.IGNORECASE)
            m = pattern.search(text)
            if not m:
                return []
            chunk = m.group(1)
            bullets = [l.strip() for l in chunk.split("\n") if l.strip().startswith("•")][:3]
            return bullets

        def extract_overall(text):
            import re as _re
            pattern = _re.compile(r"OVERALL:(.*?)$", _re.DOTALL | _re.IGNORECASE)
            m = pattern.search(text)
            if not m:
                return ""
            return m.group(1).strip()

        us_bullets = extract_section(raw, "US_BULLETS:")
        china_bullets = extract_section(raw, "CHINA_BULLETS:")
        india_bullets = extract_section(raw, "INDIA_BULLETS:")
        overall = extract_overall(raw)
        bullet_map = {"US": us_bullets, "China": china_bullets, "India": india_bullets}
        lines = [f"📊 *Market Summary* — {date.today().strftime('%d %b %Y')}\n"]
        for b in market_data_blocks:
            header = f"{b['flag']} *{b['market']} — {b['index_name']} ({b['pct_str']})*"
            bullets = bullet_map.get(b["market"], [])
            section = header
            if bullets:
                for bullet in bullets:
                    section += f"\n{bullet}"
            else:
                section += "\n• Market data unavailable"
            lines.append(section)
        if overall:
            lines.append(f"\n{overall}")
        return "\n\n".join(lines)
    except Exception as e:
        return f"❌ Couldn't pull market data right now ({type(e).__name__}: {str(e)[:80]})"

async def handle_statement_upload(file_bytes, fname, user_id, update):
    import io
    try:
        from sheets import expenses_sheet as _expenses_sheet
        await update.message.reply_text("Got the statement, give me a sec to go through it...")
        statement_rows = []
        if fname.lower().endswith(".csv"):
            import csv
            text_data = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text_data))
            for row in reader:
                statement_rows.append(dict(row))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                statement_rows.append(dict(zip(headers, row)))
        if not statement_rows:
            await update.message.reply_text("❌ Couldn't read any rows from that file.")
            return
        sample = json.dumps(statement_rows[:3], default=str)
        norm_prompt = (
            f"Given these bank statement rows: {sample}\n\n"
            f"Return ONLY a JSON object mapping these keys to the actual column names in the data:\n"
            f"{{\"date\": \"col\", \"description\": \"col\", \"amount\": \"col\"}}\n"
            f"If a column doesn't exist, use null."
        )
        norm_resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=100,
            messages=[{"role": "user", "content": norm_prompt}]
        )
        col_map = json.loads(norm_resp.content[0].text.strip().replace("```json", "").replace("```", "").strip())
        date_col = col_map.get("date")
        desc_col = col_map.get("description")
        amt_col = col_map.get("amount")
        if not all([date_col, desc_col, amt_col]):
            await update.message.reply_text("❌ Couldn't identify date/description/amount columns. Try a CSV with clear headers.")
            return
        sheet = _expenses_sheet()
        logged = sheet.get_all_records()
        missing = []
        corrections = []
        for srow in statement_rows:
            raw_date = str(srow.get(date_col, "")).strip()
            raw_desc = str(srow.get(desc_col, "")).strip()
            raw_amt = str(srow.get(amt_col, "")).strip().replace(",", "")
            if not raw_date or not raw_amt:
                continue
            try:
                stmt_amount = abs(float(raw_amt))
            except ValueError:
                continue
            stmt_date = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %b %y"):
                try:
                    stmt_date = datetime.strptime(raw_date, fmt).strftime("%d/%m/%Y")
                    break
                except Exception:
                    continue
            if not stmt_date:
                continue
            matched_row = None
            matched_idx = None
            for i, r in enumerate(logged):
                logged_date = r.get("Date", "")
                logged_sgd = float(r.get("SGD Amount") or r.get("Amount") or 0)
                if logged_date == stmt_date and abs(logged_sgd - stmt_amount) < 0.02:
                    matched_row = r
                    matched_idx = i
                    break
            if matched_row is None:
                missing.append(f"{stmt_date} | {raw_desc[:40]} | SGD ${stmt_amount:.2f}")
            else:
                logged_sgd = float(matched_row.get("SGD Amount") or matched_row.get("Amount") or 0)
                if abs(logged_sgd - stmt_amount) > 0.01:
                    all_values = sheet.get_all_values()
                    headers_row = all_values[0]
                    sgd_col_idx = headers_row.index("SGD Amount") + 1 if "SGD Amount" in headers_row else None
                    if sgd_col_idx:
                        sheet.update_cell(matched_idx + 2, sgd_col_idx, stmt_amount)
                    corrections.append(
                        f"{matched_row.get('Merchant', raw_desc[:20])} {stmt_date}: "
                        f"${logged_sgd:.2f} → ${stmt_amount:.2f}"
                    )
        lines = ["Reconciliation done ✅"]
        if corrections:
            lines.append(f"\nCorrected {len(corrections)} amount(s):")
            lines.extend(f"  {c}" for c in corrections)
        if missing:
            lines.append(f"\n{len(missing)} unmatched statement item(s) — couldn't find in your logs:")
            lines.extend(f"  {m}" for m in missing[:10])
            if len(missing) > 10:
                lines.append(f"  ...and {len(missing) - 10} more")
            lines.append("\nReply 'log [expense]' to log the first unmatched item, 'skip' to go through them, or 'done' to close.")
            state.recon_sessions[user_id] = {"step": "review", "unmatched": missing, "index": 0}
        if not corrections and not missing:
            lines.append("Everything matches up.")
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        print(f"handle_statement_upload error: {e}")
        await update.message.reply_text(f"Something went wrong parsing the statement: {str(e)}")
